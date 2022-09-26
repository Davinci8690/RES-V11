from tkinter import *
import tkinter as tk
import matplotlib.axes
import pandas
from PIL import Image, ImageTk
import numpy as np
import matplotlib.pyplot as plt
from pylab import plot, show, xlabel, ylabel
import webbrowser
from tkinter import messagebox
from tkinter import filedialog
import json
import requests
import io
import pandas as pd
from pandas import DataFrame
import openpyxl as xl
import time
import matplotlib as mpl

mpl.use("TkAgg")
from tkinter import ttk
import scipy.interpolate
from tkcalendar import DateEntry
from datetime import date
from datetime import datetime
from calendar import monthrange
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import math
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
# from tkinter.tix import Balloon
import numpy_financial as npf
from scipy.optimize import minimize, LinearConstraint, NonlinearConstraint
from tktooltip import ToolTip
import time as tm
from statistics import mean

root = Tk()
root.title("RES Tool V11.0")
root.geometry("850x650")
# w, h = root.winfo_screenwidth(), root.winfo_screenheight()
# root.geometry("%dx%d+0+0" % (w, h))
root.state("zoomed")
# root.attributes('-fullscreen', True)
# definition of variables
lst_irradiation_data = []
lst_time_data = []
lst_wind_speed = []
lst_sun_height = []
lst_air_temp = []
global IRR
global npv_df
global num
num = 0
global lst_var_sp_pv
global lst_var_sp_wt
lst_var_sp_pv=[0.04]*25
lst_var_sp_wt = [0.04]*25

my_notebook = ttk.Notebook(root, width=1500, height=850)
my_notebook.pack()


# definition of functions

def optimization():
    ####################################################################################################################
    global nb_of_iterations
    nb_of_iterations = 0
    lbl_itirations.config(text="No of iterations = " + str(nb_of_iterations))
    lbl_opt_status.config(text="Optimization status: Running...")

    ####################################################################################################################
    time_start_opt = tm.time()
    try:
        lst_opt_pv_sizes = []
        lst_opt_wt_sizes = []
        lst_opt_ST_sizes = []
        lst_opt_SS = []
        lst_opt_irr = []
        lst_opt_NPV = []
        lst_opt_PBT = []
        lst_opt_LCOE = []

        ###############################################################################################################
        def eqn(Inputs):
            x, y = Inputs
            return x ** 2 + (y + 2)

        def cons1(x):
            return x + 3

        def cons2(x):
            return x

        first_guess = [0, 0]
        cons1 = NonlinearConstraint(cons1, -200, 500)
        cons2 = NonlinearConstraint(cons2, -500, 850)
        lst_all_cons = [cons1, cons2]
        bound = scipy.optimize.Bounds(-2, 5, True)

        mymin = minimize(eqn, first_guess, method='SLSQP', constraints=lst_all_cons, bounds=bound)

        ###############################################################################################################

        def lower_self_sufficiency(Inputs):
            x, y, z = Inputs
            lst_result = download_data_optimisation(x, y, z)
            ss = lst_result[0]
            global opt_IRR
            opt_IRR = lst_result[1]
            npv = lst_result[2]
            pbt = lst_result[3]
            lcoe = lst_result[4]
            ###############################################################################################################
            ent_opt_selfsufficiency.delete(0, END)
            ent_opt_selfsufficiency.insert(END, str(round(ss)))
            ent_opt_irr.delete(0, END)
            ent_opt_irr.insert(END, str(round(lst_result[1])))
            ent_opt_PV_size.delete(0, END)
            ent_opt_PV_size.insert(END, str(round(x)))
            ent_opt_WT_size.delete(0, END)
            ent_opt_WT_size.insert(END, str(round(z)))
            ent_opt_ST_size.delete(0, END)
            ent_opt_ST_size.insert(END, str(round(y)))

            lst_opt_pv_sizes.append(x)
            lst_opt_wt_sizes.append(z)
            lst_opt_ST_sizes.append(y)
            lst_opt_SS.append(ss)
            lst_opt_irr.append(lst_result[1])
            lst_opt_NPV.append(lst_result[2])
            lst_opt_PBT.append(lst_result[3])
            lst_opt_LCOE.append(lst_result[4])

            ###############################################################################################################
            return -float(ss)

        def pv_size_opt_cons(opt_PV_size):
            return opt_PV_size

        def wt_size_opt_cons(opt_WT_size):
            return opt_WT_size

        def st_size_opt_cons(opt_ST_size):
            return opt_ST_size

        def irr_cons(Inputs):
            x, y, z = Inputs
            lst_result = download_data_optimisation(x, y, z)
            irr = lst_result[1]
            return irr

        # def irr_cons():
        #     return opt_IRR
        #######################################################################################################################
        if var_SI_unit_nom_pow.get() == 1:
            nominal_pow = float(ent_nominal_power.get())
        elif var_SI_unit_nom_pow.get() == 2:
            nominal_pow = float(ent_nominal_power.get()) * 1000
        elif var_SI_unit_nom_pow.get() == 3:
            nominal_pow = float(ent_nominal_power.get()) * 1000000
        else:
            messagebox.showerror("Error", "Please select the SI unit of nominal power")
            return

        selected_turbine = var_turbine.get()
        selected_turbine_para = para_WT_df[selected_turbine]
        sel_WT_nom_pow = selected_turbine_para[1]

        if var_SI_unit_storage_capacity.get() == 1:
            batt_capacity = float(ent_storage_capacity.get())
        if var_SI_unit_storage_capacity.get() == 2:
            batt_capacity = float(ent_storage_capacity.get()) * 1000
        if var_SI_unit_storage_capacity.get() == 3:
            batt_capacity = float(ent_storage_capacity.get()) * 1000000

        opt_nb_WT = int(ent_nb_turbines.get())

        ########################################################################################################################
        if ent_pv_size_opt_min.get() == "":
            min_opt_PV_size = nominal_pow
        else:
            min_opt_PV_size = float(ent_pv_size_opt_min.get())

        if ent_pv_size_opt_max.get() == "":
            max_opt_PV_size = np.inf
        else:
            max_opt_PV_size = float(ent_pv_size_opt_max.get())

        if ent_wt_size_opt_min.get() == "":
            min_opt_WT_size = opt_nb_WT
        else:
            min_opt_WT_size = float(ent_wt_size_opt_min.get())

        if ent_wt_size_opt_max.get() == "":
            max_opt_WT_size = np.inf
        else:
            max_opt_WT_size = float(ent_wt_size_opt_max.get())

        if ent_st_size_opt_min.get() == "":
            min_opt_ST_size = batt_capacity
        else:
            min_opt_ST_size = float(ent_st_size_opt_min.get())

        if ent_st_size_opt_max.get() == "":
            max_opt_ST_size = np.inf
        else:
            max_opt_ST_size = float(ent_st_size_opt_max.get())

        if ent_irr_opt_min.get() == "":
            min_opt_IRR = 8
        else:
            min_opt_IRR = int(ent_irr_opt_min.get())

        if ent_tol_opt.get() == "":
            min_opt_tol = 1e-8
        else:
            min_opt_tol = float(ent_tol_opt.get())

        ########################################################################################################################
        # first_guess = [min_opt_PV_size, min_opt_WT_size,min_opt_ST_size]
        first_guess = [min_opt_PV_size, min_opt_ST_size, min_opt_WT_size]

        pv_size_opt_cons = NonlinearConstraint(pv_size_opt_cons, 0, math.inf)
        wt_size_opt_cons = NonlinearConstraint(wt_size_opt_cons, 0, math.inf)
        st_size_opt_cons = NonlinearConstraint(st_size_opt_cons, 0, math.inf)
        irr_opt_con = NonlinearConstraint(irr_cons, min_opt_IRR, np.inf)
        # lst_all_cons = [pv_size_opt_cons, wt_size_opt_cons,st_size_opt_cons]
        # bound = scipy.optimize.Bounds(10000, math.inf, True)
        # bound = ((min_opt_PV_size,max_opt_PV_size), (min_opt_WT_size, max_opt_WT_size), (min_opt_ST_size, max_opt_ST_size))
        bound = (
        (min_opt_PV_size, max_opt_PV_size), (min_opt_ST_size, max_opt_ST_size), (min_opt_WT_size, max_opt_WT_size))
        # bound = ((0, np.inf), (0, sel_WT_nom_pow + 10000000), (0, batt_capacity + 10000000))

        mymin = minimize(lower_self_sufficiency, first_guess, bounds=bound, method="SLSQP", constraints=irr_opt_con,
                         tol=min_opt_tol)
        result = mymin.x
        opt_pv = result[0]
        opt_WT = result[1]
        opt_st = result[2]
        selfsuf = -mymin.fun

        ########################################################################################################################
        btn_opt_graph = Button(tab_opt, text="Graph",
                               command=lambda: plot_opt_graph(lst_opt_pv_sizes, lst_opt_wt_sizes, lst_opt_ST_sizes
                                                              , lst_opt_SS, lst_opt_irr))
        btn_opt_graph.place(x=100, y=10)

        btn_opt_SS_irr = Button(tab_opt, text="SS & IRR",
                                command=lambda: plot_5_graph(range(1, len(lst_opt_pv_sizes) + 1), lst_opt_SS,
                                                             lst_opt_irr, lst_opt_NPV, lst_opt_PBT, lst_opt_LCOE,
                                                             "Self Sufficiency", "IRR", "NPV", "PBT", "LCOE", "SS&IRR",
                                                             "Iterations", "Optimum resuls"))
        btn_opt_SS_irr.place(x=200, y=10)
        ########################################################################################################################
        #creating optimizing results df'
        lst_opt_wt_sizes= [round(each_num) for each_num in lst_opt_wt_sizes]
        opt_results_df = pd.DataFrame()
        opt_results_df.insert(0, "PV sizes(kW)", lst_opt_pv_sizes)
        opt_results_df.insert(1, "Number of WT",lst_opt_wt_sizes)
        opt_results_df.insert(2, "ST sizes(kWh)",lst_opt_ST_sizes)
        opt_results_df.insert(3, "Self Sufficiency", lst_opt_SS)
        opt_results_df.insert(4, "IRR(%)",lst_opt_irr)
        opt_results_df.insert(5, "NPV", lst_opt_NPV)
        opt_results_df.insert(6, "PBT", lst_opt_PBT)
        opt_results_df.insert(7, "LCOE", lst_opt_LCOE)

        btn_show_opt_results = Button(tab_opt, text="Show ", command = lambda: show_data(opt_results_df))
        btn_show_opt_results.place(x=300, y=10)
        btn_export_opt_results = Button(tab_opt, text="Export ", command= lambda: export_to_excel_with_name(opt_results_df, "Optimization results"))
        btn_export_opt_results.place(x=450, y=10)
        ########################################################################################################################
        time_end_opt = tm.time()
        total_time = round(time_end_opt - time_start_opt)
        lbl_PV_size_MW = Label(tab_opt, text="(MW) "+str(float(ent_opt_PV_size.get())/1000))
        lbl_PV_size_GW = Label(tab_opt, text="(GW) " + str(float(ent_opt_PV_size.get()) / 1000000))

        lbl_PV_size_MW.place(x=550, y=65)
        lbl_PV_size_GW.place(x=550,y=80)

        lbl_WT_size_MW = Label(tab_opt, text="(MW) " + str(sel_WT_nom_pow* float(ent_opt_WT_size.get())/1000))
        lbl_WT_size_GW = Label(tab_opt, text="(GW) " + str(sel_WT_nom_pow* float(ent_opt_WT_size.get())/1000000))
        lbl_WT_size_MW.place(x=755, y=65)
        lbl_WT_size_GW.place(x=755, y=80)

        lbl_ST_size_MW = Label(tab_opt, text="(MW) "+str(float(ent_opt_ST_size.get())/1000))
        lbl_ST_size_GW = Label(tab_opt, text="(GW) " + str(float(ent_opt_ST_size.get()) / 1000000))

        lbl_ST_size_MW.place(x=960, y=65)
        lbl_ST_size_GW.place(x=960, y=80)


        lbl_opt_status.config(text="Optimization status: Completed :)")
        messagebox.showinfo("Completed", "Optimization Completed and total time = " + str(total_time / 60) + " mins")

        return
    except:
        messagebox.showerror("Error", "Error while optimizing")


def download_data_optimisation(opt_PV_size, opt_ST_size, opt_nb_WT):
    # if ent_start_year.get() > ent_end_year.get():
    #     messagebox.showerror("Date error", "Starting Year should be less than end year")
    #     return
    # if len(ent_start_year.get()) != 4 or len(ent_end_year.get()) != 4:
    #     messagebox.showerror("Date error", "starting year or Ending year is not entered properly")
    #     return
    # if int(ent_start_year.get()) < 2005 or int(ent_end_year.get()) > 2016:
    #     messagebox.showerror("Date error", "Incorrect value. Please, enter an integer between 2005 and 2016")
    #     return
    # if ent_load_type.get() == "":
    #     messagebox.showerror("Error", "Please select excel file with hourly load values")
    #     return
    #
    # latitude = str(ent_latitude.get())
    # longitude = str(ent_longitude.get())
    global nb_of_iterations
    nb_of_iterations += 1
    lbl_itirations.config(text="No of iterations = " + str(nb_of_iterations))

    #

    start_year = str(ent_start_year.get())
    end_year = str(ent_end_year.get())
    total_years = int(end_year) - int(start_year)
    if var_load.get() == "Load of 1 year and increase %":
        lst_load_all_years = create_load_list_frm_1st_year(total_years)
        if lst_load_all_years == None:
            return
    elif var_load.get() == "Load of 25 years":
        lst_load_all_years = create_load_list_frm_25years(total_years)
        if lst_load_all_years == None:
            return
    ########################################################################################################################
    try:

        # progress bar start
        newWindow = Toplevel(root)
        newWindow.title("Progress Bar")
        newWindow.geometry('400x100+400+250')

        newWindow.protocol("WM_DELETE_WINDOW", disable_event)

        # progress_bar = ttk.Progressbar(newWindow, orient=HORIZONTAL, length=400, mode="determinate")
        # progress_bar.pack()
        lbl_progress_bar = Label(newWindow, text="")
        lbl_progress_bar.pack()

        # progress_bar["value"] += 2

        lbl_progress_bar.config(text="Starting new itiration")
        root.update()
    except:
        messagebox.showerror("Error")
    #######################################################################################################################

    # url_first_part = "https://re.jrc.ec.europa.eu/api/seriescalc?"
    # final_url = url_first_part + "lat=" + latitude + "&lon=" + longitude + "&startyear=" + start_year \
    #             + "&endyear=" + end_year + "&optimalangles=1&outputformat=json&browser=1"
    # res = requests.get(final_url)
    # res_json_file = io.StringIO(res.text)
    # src = json.load(res_json_file)
    # output = src['outputs']
    # output_hourly = output["hourly"]
    # data = pd.DataFrame(output_hourly)
    # data.index = pd.to_datetime(data["time"], format='%Y%m%d:%H%M', utc=True)
    # data = data.drop("Int", axis=1)

    data = data_for_opt[data_for_opt.columns[0:5]]
    # data_raw = data.drop("time", axis=1)
    ########################################################################################################################
    try:
        data.insert(5, "Load(kW)", lst_load_all_years)
    except:
        messagebox.showerror("Error", "Error while appending load data")

    ##############################################################o###########################################################
    # data_mean = data_raw.resample('M').mean()
    # data_sum = data_raw.resample('M').sum()

    # inserting delat G in the dataframe
    g0 = ent_G0.get()
    lst_g0 = []
    col_go = data["G(i)"]

    for each_row in col_go:
        lst_g0.append(each_row - float(g0))

    data.insert(6, "Delta G", lst_g0)

    # inserting cell temp in the dataframe
    lst_cellTemp = []
    noct = ent_noct.get()
    noct_Tref = ent_Trif_noct.get()
    delta_NOCT = float(noct) - float(noct_Tref)
    col_Ta = data["T2m"]
    col_deltaG0 = data["Delta G"]

    for i in range(0, len(col_Ta)):
        cell_temp = col_Ta[i] + delta_NOCT / 800 * col_deltaG0[i]
        lst_cellTemp.append(cell_temp)

    data.insert(7, "Cell Temp", lst_cellTemp)

    # inserting DC power column in DF
    nominal_pow = opt_PV_size

    lst_DC_power = []
    eta_dirt = float(ent_etaDirt.get())
    eta_mismatch = float(ent_etaMM.get())
    eta_ref = float(ent_etaRef.get())
    eta_cable = float(ent_etaCable.get())

    gamma_th = float(ent_Gamma.get())
    temp_STC = float(ent_T_STC.get())
    col_cell_temp = data["Cell Temp"]

    for i in range(0, len(col_deltaG0)):
        DC_pow = nominal_pow * col_deltaG0[i] / 1000 * (
                1 + gamma_th * (col_cell_temp[i] - temp_STC)) * eta_dirt * eta_ref * eta_mismatch * eta_cable
        if DC_pow < 0.001:
            DC_pow = 0
        lst_DC_power.append(DC_pow)

    data.insert(8, "DC Power", lst_DC_power)

    # progress_bar["value"] = 5
    # lbl_progress_bar.config(text="Calculating Invertor Efficiency")
    # root.update()
    # calculation of Pac by formula
    if nominal_pow != 0:
        p_0 = 0.002 * nominal_pow
        k_linear = 0.005
        k_quad = (0.02 / nominal_pow)
        lst_AC_power = [None] * len(lst_DC_power)
        lst_AC_DC_eff = [None] * len(lst_DC_power)
        for i in range(0, len(lst_DC_power)):
            if lst_DC_power[i] != 0:
                b = 1 + k_linear
                a = k_quad
                c = lst_DC_power[i] - p_0
                num_AC_power = -b + ((b ** 2) + 4 * a * c) ** (0.5)
                de_AC_power = 2 * a
                lst_AC_power[i] = num_AC_power / de_AC_power
            else:
                lst_AC_power[i] = 0
        for i in range(0, len(lst_DC_power)):
            if lst_DC_power[i] != 0:
                lst_AC_DC_eff[i] = lst_AC_power[i] / lst_DC_power[i]
            else:
                lst_AC_DC_eff[i] = 0
    else:
        lst_AC_power = [0] * len(lst_DC_power)
        lst_AC_DC_eff = [0] * len(lst_DC_power)

    data.insert(9, "DC/AC effficiency", lst_AC_DC_eff)

    data.insert(10, "Power in ALternate Current(kW)", lst_AC_power)
    # # creating dataframe to find invertor efficiency
    #
    # # insertinf Pac% col in invertor DF
    # inv_df = pd.DataFrame()
    # inv_perc = 0.0
    # lst_Pac_percentage = []
    # while inv_perc < 120:
    #     lst_Pac_percentage.append(inv_perc)
    #     inv_perc += 0.25
    #
    # inv_df.insert(0, "Pac%", lst_Pac_percentage)
    #
    # # inserting AC power col in inverter DF
    # lst_Pac = []
    # col_Pac_per = inv_df["Pac%"]
    # for i in range(0, len(col_Pac_per)):
    #     powerAC = nominal_pow * col_Pac_per[i] / 100
    #     lst_Pac.append(powerAC)
    #
    # inv_df.insert(1, "Pac(kW)", lst_Pac)
    #
    # # inserting inverter efficiency in invertor DF
    #
    # lst_eta_inv = []
    # col_Pac = inv_df["Pac(kW)"]
    #
    # if nominal_pow != 0:
    #     noload_loss = 0.7 / 100 * nominal_pow
    #     # noload_loss = float(ent_noloadloss.get())
    #     linear_loss = noload_loss / nominal_pow
    #     # linear_loss = float(ent_linearloss.get())
    #     quadratic_loss = linear_loss / nominal_pow
    #     # quadratic_loss = float(ent_Q_loss.get())
    # else:
    #     noload_loss = float(ent_noloadloss.get())
    #     linear_loss = float(ent_linearloss.get())
    #     quadratic_loss = float(ent_Q_loss.get())
    # for i in range(0, len(col_Pac)):
    #     eta_inv = col_Pac[i] / (
    #                 col_Pac[i] + noload_loss + (linear_loss * col_Pac[i]) + (quadratic_loss * (col_Pac[i]) ** 2)) * 100
    #     lst_eta_inv.append(eta_inv)
    #
    # inv_df.insert(2, "eta_inv", lst_eta_inv)
    #
    # # inserting DC power in invertor DF
    # try:
    #     lst_dc_pow = []
    #     col_eta_inv = inv_df["eta_inv"]
    #     for i in range(0, len(col_eta_inv)):
    #         pow_dc_inv = col_Pac[i] / col_eta_inv[i]
    #         lst_dc_pow.append(pow_dc_inv)
    #
    #     inv_df.insert(3, "Pdc", lst_dc_pow)
    # except:
    #     messagebox.showerror("Error", "Python found a error while Calculating DC power")
    #
    # # inserting Load factor in invertor DF
    # lst_load_factor = []
    # for i in range(0, len(col_Pac)):
    #     load_factor = col_Pac[i] / nominal_pow * 100
    #     lst_load_factor.append(load_factor)
    #
    # inv_df.insert(4, "Load Factor (%)", lst_load_factor)
    #
    # # inserting DC/AC inv efficiency in DF
    # try:
    #     lst_AC_DC_eff = []
    #     col_DC_pow = data["DC Power"]
    #     # col_loadFactor = inv_df["Load Factor (%)"]
    #     # lst_load_factor2 =[]
    #     # for i in range (0, len(col_loadFactor)):
    #     #     lst_load_factor2.append(col_loadFactor[i])
    #     for i in range(0, len(col_DC_pow)):
    #         if nominal_pow != 0:
    #             dc_nompow_ratio = (col_DC_pow[i] / nominal_pow) * 100
    #         else:
    #             dc_nompow_ratio = 0
    #         y_interp = scipy.interpolate.interp1d(lst_load_factor, lst_eta_inv)
    #         DCAC_eff = y_interp(dc_nompow_ratio)
    #
    #         lst_AC_DC_eff.append((DCAC_eff))
    #
    #     data.insert(9, "DC/AC effficiency", lst_AC_DC_eff)
    # except:
    #     messagebox.showerror("Error", "Python found a error while Calculating inverter efficiency ")
    #
    # # inserting DC/AC inv efficiency in DF
    # progress_bar["value"] = 10
    # lbl_progress_bar.config(text="Calculating AC power produced by PV")
    # root.update()
    #
    # lst_AC_pow = []
    # col_DCAC_eff = data["DC/AC effficiency"]
    # for i in range(0, len(col_DC_pow)):
    #     AC_pow = col_DC_pow[i] * col_DCAC_eff[i] / 100
    #     lst_AC_pow.append(AC_pow)
    #
    # data.insert(10, "Power in ALternate Current(kW)", lst_AC_pow)
    ########################################################################################################################
    yearly_data_df = data.resample('Y').sum()
    yearly_data_df = yearly_data_df["Power in ALternate Current(kW)"]
    yearly_data_df = yearly_data_df.div(nominal_pow)
    lbl_SI_PV_prod = "(kWh/kW/year)"

    # if var_SI_unit_nom_pow.get()==1:
    #     yearly_data_df=yearly_data_df
    #     lbl_SI_PV_prod = "(kWh)"
    # elif var_SI_unit_nom_pow.get()==2:
    #     yearly_data_df=yearly_data_df.div(1000)
    #     lbl_SI_PV_prod = "(MWh)"
    # elif var_SI_unit_nom_pow.get()==3:
    #     yearly_data_df=yearly_data_df.div(1000000)
    #     lbl_SI_PV_prod = "(GWh)"
    # else:
    #     messagebox.showerror("Error","Please select the SI unit of nominal power")
    #     return
    yearly_data_df = yearly_data_df.round()

    # Calcualtion of wind turbine production
    try:
        lst_wind_speed_hubHeight = []
        lst_wind_turbine_prod = []
        selected_turbine = var_turbine.get()
        selected_turbine_para = para_WT_df[selected_turbine]
        height_of_rotor = selected_turbine_para[0]
        terrain_rough = float(ent_terrain_roug.get())
        sel_WT_nom_pow = selected_turbine_para[1]
        col_wind_speed = data["WS10m"]
        nb_WT = opt_nb_WT
        # ref_height = float(ent_mes_height.get())
        # numerator = math.log(height_of_rotor / terrain_rough)
        # denominator = math.log(ref_height / terrain_rough)
        #
        # int_col_wind_speed2 = power_curve_df.columns.get_loc(selected_turbine) - 1
        # lst_power_PC = power_curve_df[selected_turbine].tolist()
        # lst_wind_speed_PC = power_curve_df.iloc[:, int_col_wind_speed2].tolist()
        # all_WT_df = pd.DataFrame()
        # col_time = data["time"]
        #
        # lst_time_data = []
        # lst_wind_speed = []
        # for o in range (0, len(col_time)):
        #     lst_time_data.append(col_time[o])
        #     lst_wind_speed.append(col_wind_speed[o])
        # lst_time_data = data["time"].tolist
        # all_WT_df.insert(0,"Time",lst_time_data)
        # all_WT_df = data[["time"]]
        # # all_WT_df.insert(1, "Wind Speed(m/s)", lst_wind_speed)
        # for i in range(0, len(col_wind_speed)):
        #     corrected_wind_speed = col_wind_speed[i] * (numerator / denominator)
        #     wind_pow_interp = scipy.interpolate.interp1d(lst_wind_speed_PC, lst_power_PC)
        #     wind_turbine_prod = wind_pow_interp(corrected_wind_speed)
        #     wind_turbine_prod = wind_turbine_prod * nb_WT
        #     lst_wind_speed_hubHeight.append(corrected_wind_speed)
        #     lst_wind_turbine_prod.append(wind_turbine_prod)

        old_nb_wt = int(ent_nb_turbines.get())

        lst_old_wind_speed = data_for_opt[data_for_opt.columns[11:12]]
        lst_old_Wt_prod = data_for_opt[data_for_opt.columns[12:13]]
        lst_new_Wt_prod = (lst_old_Wt_prod / old_nb_wt) * nb_WT

        data.insert(11, "Wind speed at Hub height(m/s)", lst_old_wind_speed)
        data.insert(12, "Wind Turbine Active Power(kW)", lst_new_Wt_prod)

        # progress_bar["value"] = 15
        # lbl_progress_bar.config(text="Caculating AC power produced by Wind Turbine")
        # root.update()
        # int_WT= 1
        # for j in range (0, int(len(power_curve_df.columns)/2)):
        #     selected_turbine_name_int_WT = para_WT_df.columns[j]
        #     selected_turbine_para_int_WT = para_WT_df[selected_turbine_name_int_WT]
        #     height_of_rotor_int_WT = selected_turbine_para_int_WT[0]
        #     numerator_int_WT = math.log(height_of_rotor_int_WT / terrain_rough)
        #     denominator_int_WT = math.log(ref_height / terrain_rough)
        #     lbl_progress_bar.config(text="Calculating AC power produced by "+str(selected_turbine_name_int_WT))
        #     root.update()
        #     int_col_wind_speed_int_WT = power_curve_df.columns.get_loc(selected_turbine_name_int_WT) - 1
        #     lst_power_PC_int_WT = power_curve_df[selected_turbine_name_int_WT].tolist()
        #     lst_wind_speed_PC_int_WT = power_curve_df.iloc[:, int_col_wind_speed_int_WT].tolist()
        #
        #     lst_wind_turbine_prod_int_WT = []
        #     for t in range (0, len(col_wind_speed)):
        #         corrected_wind_speed_int_WT = col_wind_speed[t] * (numerator_int_WT / denominator_int_WT)
        #         wind_pow_interp_int_WT = scipy.interpolate.interp1d(lst_wind_speed_PC_int_WT, lst_power_PC_int_WT)
        #         wind_turbine_prod_int_WT = wind_pow_interp_int_WT(corrected_wind_speed_int_WT)
        #         wind_turbine_prod_int_WT = wind_turbine_prod_int_WT * nb_WT
        #         lst_wind_turbine_prod_int_WT.append(wind_turbine_prod_int_WT)
        #
        #     all_WT_df.insert(j+1,para_WT_df.columns[j],lst_wind_turbine_prod_int_WT)
        #     int_WT +=2

        # all_WT_df = all_WT_df.resample('Y').sum()

        # all_WT_df = all_WT_df.round()
    except:
        messagebox.showerror("Error", "Python found error while calculating the Wind turbine production")
    ########################################################################################################Commenting on 21-07 ends
    # Battery storage calculations#####################################################################################
    # progress_bar["value"] = 20
    # lbl_progress_bar.config(text="Limitation on Maximum generated power")
    # root.update()

    try:
        lst_overall_prod = []
        col_PV_prod = data["Power in ALternate Current(kW)"]
        col_WT_prod = data["Wind Turbine Active Power(kW)"]
        for i in range(0, len(col_PV_prod)):
            overall_prod = col_PV_prod[i] + col_WT_prod[i]
            lst_overall_prod.append(overall_prod)

        data.insert(13, "Overall Production(kW)", lst_overall_prod)

        lst_over_prod_with_PS = []
        col_ovr_prod = data["Overall Production(kW)"]

        if var_peak_shaving.get() == "Limitation on Maximum generated power":
            if var_SI_unit_PS.get() == 1:
                lim_max_gen = float(ent_max_gen.get())
            elif var_SI_unit_PS.get() == 2:
                lim_max_gen = float(ent_max_gen.get()) * 1000
            elif var_SI_unit_PS.get() == 3:
                lim_max_gen = float(ent_max_gen.get()) * 1000000
            else:
                messagebox.showerror("Error", "Error in converting SI unit of peak shaving")
                return

        for i in range(0, len(col_ovr_prod)):

            if var_peak_shaving.get() == "Limitation on Maximum generated power":
                if col_ovr_prod[i] > lim_max_gen:
                    overall_prod_with_PS = lim_max_gen
                else:
                    overall_prod_with_PS = col_ovr_prod[i]
            else:
                overall_prod_with_PS = col_ovr_prod[i]

            lst_over_prod_with_PS.append(overall_prod_with_PS)

        data.insert(14, "Overall production with Peak Shaving(kW)", lst_over_prod_with_PS)

        lst_wastage_PS_gen = []
        col_over_prod_with_PS = data["Overall production with Peak Shaving(kW)"]
        # progress_bar["value"] = 25
        # lbl_progress_bar.config(text="Calculating Overall(PV+WT) production")
        # root.update()
        for i in range(0, len(col_over_prod_with_PS)):
            wastage_PS_gen = col_ovr_prod[i] - col_over_prod_with_PS[i]
            lst_wastage_PS_gen.append(wastage_PS_gen)

        data.insert(15, "Not produced energy due to peak shaving(kW)", lst_wastage_PS_gen)

        lst_balance = []
        col_load = data["Load(kW)"]
        col_overall_prod = data["Overall Production(kW)"]
        for i in range(0, len(col_over_prod_with_PS)):
            balance = col_over_prod_with_PS[i] - col_load[i]
            lst_balance.append(balance)

        data.insert(16, "Balance(kW)", lst_balance)
    except:
        messagebox.showerror("Error", "Python found error while calculating the battery storage")
    # Battery df#####################################################################################

    try:
        col_balance = data["Balance(kW)"]
        lst_SOC = []
        lst_pow_discharge = []
        lst_pow_charge = []
        lst_batt_usage_DC_side = []
        lst_batt_usage_AC_side = []
        lst_batt_discharge = []
        lst_batt_charge = []
        lst_balance_with_storage = []
        lst_absorption_frm_grid = []
        lst_injection_in_grid = []
        discharge_eff = float(ent_discharge_eff.get()) / 100
        charge_eff = float(ent_charge_eff.get()) / 100
        min_SOC = float(ent_min_soc.get())
        max_SOC = float(ent_max_soc.get())

        batt_capacity = opt_ST_size

        for i in range(0, len(col_balance)):
            if i == 0:
                SOC = min_SOC
            else:
                SOC = ((lst_SOC[i - 1] / 100) - (lst_batt_usage_DC_side[i - 1] / batt_capacity)) * 100

            lst_SOC.append(SOC)

            avlb_pow_discharge = ((lst_SOC[i] - min_SOC) / 100) * batt_capacity
            avlb_pow_charge = ((max_SOC - lst_SOC[i]) / 100) * batt_capacity

            if col_balance[i] < 0:
                lst1 = [avlb_pow_discharge, -col_balance[i] / discharge_eff, batt_capacity]
                batt_usage_DC_side = min(lst1)
                batt_usage_AC_side = batt_usage_DC_side * discharge_eff
            else:
                lst1 = [avlb_pow_charge, col_balance[i] * charge_eff, batt_capacity]
                batt_usage_DC_side = - min(lst1)
                batt_usage_AC_side = batt_usage_DC_side / charge_eff
            # progress_bar["value"] = 30
            # lbl_progress_bar.config(text="Calculating SOC% of battery")
            # root.update()
            lst_pow_discharge.append(avlb_pow_discharge)
            lst_pow_charge.append(avlb_pow_charge)
            lst_batt_usage_DC_side.append(batt_usage_DC_side)
            lst_batt_usage_AC_side.append(batt_usage_AC_side)

            if lst_batt_usage_AC_side[i] >= 0:
                batt_discharge = lst_batt_usage_AC_side[i]
            else:
                batt_discharge = 0

            if lst_batt_usage_AC_side[i] <= 0:
                batt_charge = -lst_batt_usage_AC_side[i]
            else:
                batt_charge = 0

            balance_with_storage = col_balance[i] + lst_batt_usage_AC_side[i]
            if balance_with_storage < 0:
                absortionfrom_grid = - balance_with_storage
            else:
                absortionfrom_grid = 0

            if balance_with_storage >= 0:
                injectioin_grid = balance_with_storage
            else:
                injectioin_grid = 0

            lst_batt_discharge.append(batt_discharge)
            lst_batt_charge.append(batt_charge)
            lst_balance_with_storage.append(balance_with_storage)
            lst_absorption_frm_grid.append(absortionfrom_grid)
            lst_injection_in_grid.append(injectioin_grid)

        # progress_bar["value"] = 35
        # lbl_progress_bar.config(text="Calculation of Grid exchange")
        # root.update()
        batt_df = data[["Power in ALternate Current(kW)", "Wind Turbine Active Power(kW)", "Overall Production(kW)",
                        "Overall production with Peak Shaving(kW)", "Not produced energy due to peak shaving(kW)",
                        "Balance(kW)"]]
        batt_df.insert(6, "State of charge (%)", lst_SOC)
        batt_df.insert(7, "Available power to charge(kW)", lst_pow_charge)
        batt_df.insert(8, "Available power to discharge(kW)", lst_pow_discharge)
        batt_df.insert(9, "Actual battery usage inside battery (DC side) ", lst_batt_usage_DC_side)
        batt_df.insert(10, "Actual battery usage outside battery (AC side) ", lst_batt_usage_AC_side)
        batt_df.insert(11, "Battery Discharge(kW)", lst_batt_discharge)
        batt_df.insert(12, "Battery Charge(kW)", lst_batt_charge)
        batt_df.insert(13, "Balance with Storage (kW)", lst_balance_with_storage)
        ###Inserting Battery df into data df#################################################################################
        data.insert(17, "State of charge (%)", lst_SOC)
        data.insert(18, "Available power to charge(kW)", lst_pow_charge)
        data.insert(19, "Available power to discharge(kW)", lst_pow_discharge)
        data.insert(20, "Actual battery usage inside battery (DC side) ", lst_batt_usage_DC_side)
        data.insert(21, "Actual battery usage outside battery (AC side) ", lst_batt_usage_AC_side)
        data.insert(22, "Battery Discharge(kW)", lst_batt_discharge)
        data.insert(23, "Battery Charge(kW)", lst_batt_charge)
        data.insert(24, "Balance with Storage (kW)", lst_balance_with_storage)
        data.insert(25, "Absorption from grid(kW)", lst_absorption_frm_grid)
        data.insert(26, "Injection in grid(kW)", lst_injection_in_grid)
    except:
        messagebox.showerror("Error", "Python found error while calculating the battery soc")

    ####################################################################################
    try:
        # Limitation on injection into grid
        if var_peak_shaving.get() == "Limitation on Maximum power injection":
            if var_SI_unit_PS.get() == 1:
                lim_max_inj = float(ent_max_inj.get())
            elif var_SI_unit_PS.get() == 2:
                lim_max_inj = float(ent_max_inj.get()) * 1000
            elif var_SI_unit_PS.get() == 3:
                lim_max_inj = float(ent_max_inj.get()) * 1000000
            else:
                messagebox.showerror("Error", "Error in converting SI unit of peak shaving")
                return

        col_inj_grid = data["Injection in grid(kW)"]
        lst_inj_with_PS = []
        for i in range(0, len(col_inj_grid)):
            if var_peak_shaving.get() == "Limitation on Maximum power injection":
                if col_inj_grid[i] <= lim_max_inj:
                    inj_with_PS = col_inj_grid[i]
                else:
                    inj_with_PS = lim_max_inj
            else:
                inj_with_PS = col_inj_grid[i]

            lst_inj_with_PS.append(inj_with_PS)

        data.insert(27, "Injection in grid with Peak shaving(kW)", lst_inj_with_PS)

        col_inj_grid_with_PS = data["Injection in grid with Peak shaving(kW)"]
        lst_wastage_PS_inj = []
        for i in range(0, len(col_inj_grid_with_PS)):
            wastage_PS_inj = col_inj_grid[i] - col_inj_grid_with_PS[i]
            lst_wastage_PS_inj.append(wastage_PS_inj)

        data.insert(28, "Not produced energy due to peak shaving of injection(kW)", lst_wastage_PS_inj)

        lst_overall_prod_after_both_PS = []
        col_waste_PS_gen = data["Not produced energy due to peak shaving(kW)"]
        col_waste_PS_inj = data["Not produced energy due to peak shaving of injection(kW)"]
        for i in range(0, len(col_inj_grid_with_PS)):
            overall_prod_after_both_PS = col_ovr_prod[i] - col_waste_PS_gen[i] - col_waste_PS_inj[i]
            lst_overall_prod_after_both_PS.append(overall_prod_after_both_PS)

        data.insert(29, "PV+WT production after peak shaving (any kind of peak shaving)",
                    lst_overall_prod_after_both_PS)
        col_absorption_frm_grid = data["Absorption from grid(kW)"]
        lst_grid_exchange = []
        for i in range(0, len(col_absorption_frm_grid)):
            if col_absorption_frm_grid[i] == 0:
                grid_exchange = col_inj_grid_with_PS[i]
            else:
                grid_exchange = col_absorption_frm_grid[i]

            lst_grid_exchange.append(grid_exchange)

        data.insert(30, "Grid Exchange(kW)", lst_grid_exchange)
        # progress_bar["value"] = 40
        # lbl_progress_bar.config(text="Calculating Self Sufficiency")
        # root.update()
        try:
            lst_self_sufficiency = []

            for i in range(0, len(col_absorption_frm_grid)):
                if col_absorption_frm_grid[i] == 0:
                    self_sufficiency = col_load[i]
                else:
                    self_sufficiency = col_load[i] - col_absorption_frm_grid[i]

                lst_self_sufficiency.append(self_sufficiency)

            data.insert(31, "Self Suffiency", lst_self_sufficiency)

        except:
            messagebox.showerror("Error", "Error while calculating self sufficiency")
    except:
        messagebox.showerror("Error", "Error while calculating injection into grid with peak shaving")

    ##Ripartition of peak shaving between PV and WT##################################################################################
    try:
        lst_quota_PV = []
        for i in range(0, len(col_ovr_prod)):
            if col_ovr_prod[i] == 0:
                quota_PV = 0
            else:
                quota_PV = (col_PV_prod[i] / col_ovr_prod[i]) * 100

            lst_quota_PV.append(quota_PV)

        data.insert(32, "Quota of production of PV with respect to (PV+WT) [%]", lst_quota_PV)

        lst_quota_WT = []
        col_overall_prod_after_both_PS = data["PV+WT production after peak shaving (any kind of peak shaving)"]

        for i in range(0, len(col_overall_prod_after_both_PS)):
            if col_overall_prod_after_both_PS[i] == 0:
                quota_WT = 0
            else:
                quota_WT = (col_WT_prod[i] / col_ovr_prod[i]) * 100

            lst_quota_WT.append(quota_WT)

        data.insert(33, "Quota of production of WT with respect to (PV+WT) [%]", lst_quota_WT)

        load = float(ent_load.get())
        lst_PV_prod_after_PS = []
        lst_WT_prod_after_PS = []
        lst_inj_grid_PV = []
        lst_inj_grid_WT = []
        lst_self_consumption_PV = []
        lst_self_consumption_WT = []
        lst_load = []

        col_quota_PV = data["Quota of production of PV with respect to (PV+WT) [%]"]
        col_quota_WT = data["Quota of production of WT with respect to (PV+WT) [%]"]
        col_self_suffiency = data["Self Suffiency"]

        for i in range(0, len(col_quota_PV)):
            PV_prod_after_PS = (col_quota_PV[i] / 100) * col_overall_prod_after_both_PS[i]
            PV_prod_after_WT = (col_quota_WT[i] / 100) * col_overall_prod_after_both_PS[i]
            inj_grid_PV = (col_quota_PV[i] / 100) * col_inj_grid_with_PS[i]
            inj_grid_WT = (col_quota_WT[i] / 100) * col_inj_grid_with_PS[i]
            self_consumption_PV = (col_quota_PV[i] / 100) * col_self_suffiency[i]
            self_consumption_WT = (col_quota_WT[i] / 100) * col_self_suffiency[i]

            lst_PV_prod_after_PS.append(PV_prod_after_PS)
            lst_WT_prod_after_PS.append(PV_prod_after_WT)
            lst_inj_grid_PV.append(inj_grid_PV)
            lst_inj_grid_WT.append(inj_grid_WT)
            lst_self_consumption_PV.append(self_consumption_PV)
            lst_self_consumption_WT.append(self_consumption_WT)
            lst_load.append(load)

        data.insert(34, "Production of PV after peak shaving(kW)", lst_PV_prod_after_PS)
        data.insert(35, "Production of WT after peak shaving(kW)", lst_WT_prod_after_PS)
        data.insert(36, "Injection in the grid from PV(kW)", lst_inj_grid_PV)
        data.insert(37, "Injection in the grid from WT(kW)", lst_inj_grid_WT)
        data.insert(38, "Self-consumtion for PV(kW)", lst_self_consumption_PV)
        data.insert(39, "Self-consumtion for WT(kW)", lst_self_consumption_WT)
        # progress_bar["value"] = 45
        # lbl_progress_bar.config(text="Calculating Self Consumption")
        # root.update()
    except:
        messagebox.showerror("Error", "Error while calculating Ripartition of peak shaving between PV and WT")

    ####################################################################################
    # Display Montly data
    # frame_MD = LabelFrame(tab_PV_results, text="Show Monthly Data", padx=10, pady=10)
    # frame_MD.place(x=10, y=10, height=150, width=400)
    # # frame_MD.place(row=7, column=0, pady=10, columnspan=2)
    # btn_irrdata = Button(frame_MD, text="Irr Data", width=10,
    #                      command=lambda: display_irradiation_heatmap(data_sum, total_years, int(start_year)))
    # # btn_irrdata = Button(frame_MD, text="Show Irr Data", command=lambda : display_heat_map(data_sum, total_years, int(start_year), "G(i)"))
    # btn_irrdata.place(x=10, y=10)
    # btn_tempdata = Button(frame_MD, text="Temp Data", width=10,
    #                       command=lambda: display_temp_heatmap(data_mean, total_years, int(start_year)))
    # # btn_tempdata = Button(frame_MD, text="Show Temp Data",command=lambda: display_heat_map(data_mean, total_years, int(start_year), "T2m"))
    # btn_tempdata.place(x=130, y=10)
    # btn_winddata = Button(frame_MD, text="Wind Data", width=10,
    #                       command=lambda: display_wind_heatmap(data_mean, total_years, int(start_year)))
    # # btn_winddata = Button(frame_MD, text="Show Wind Data",command=lambda: display_heat_map(data_mean, total_years, int(start_year), "WS10m"))
    # btn_winddata.place(x=250, y=10)
    #
    # var_SI_unit = IntVar()
    # btn_PV_data = Button(frame_MD, text="PV+WT production", width=15,
    #                      command=lambda: show_PV_barGraph_monthly(data, total_years, start_year, var_SI_unit.get()))
    # btn_PV_data.place(x=30, y=50)
    # btn_WT_data = Button(frame_MD, text="Exchange with Grid", width=15,
    #                      command=lambda: show_grid_ex_barGraph_monthly(data, total_years, start_year,
    #                                                                    var_SI_unit.get()))
    # btn_WT_data.place(x=160, y=50)
    # rad_SI_unit_kw = Radiobutton(frame_MD, text="kWh/month", variable=var_SI_unit, value=1)
    # rad_SI_unit_kw.place(x=10, y=90)
    # rad_SI_unit_Mw = Radiobutton(frame_MD, text="MWh/month", variable=var_SI_unit, value=2)
    # rad_SI_unit_Mw.place(x=120, y=90)
    # rad_SI_unit_Gw = Radiobutton(frame_MD, text="GWh/month", variable=var_SI_unit, value=3)
    # rad_SI_unit_Gw.place(x=230, y=90)
    #
    # btn_graph_eff = Button(frame_para, text="Graph",
    #                        command=lambda: plot_graph(lst_load_factor, lst_eta_inv, "Load Factor (%) = Pac/Pac-nom",
    #                                                   "Efficiency of the DC/AC converter (%)", ""))
    # # btn_graph_eff.grid(row=5, column=0,pady=10, columnspan= 4)
    # btn_graph_eff.place(x=50, y=111)
    #
    # frame_all_data = LabelFrame(tab_PV_results, text="All Data", padx=10, pady=10)
    # # frame_all_data.grid(row=6, column=1, pady=10, columnspan=2, rowspan=1)
    # # frame_all_data.grid(row=0, column=6, columnspan=2,  sticky=tk.NW)
    # # frame_all_data.place(x=500, y=500, height=75, width= 200)
    # btn_show_data = Button(frame_all_data, text="Show Data", command=lambda: show_data(data))
    # # btn_show_data.place(x=10, y=5)
    #
    # btn_export_excel = Button(tab_Allresults, text="Export", command=lambda: export_to_excel(data))
    # btn_export_excel.place(x=0, y=0)
    #
    # frame_batt_data = LabelFrame(tab_PV_results, text="Battery Data", padx=10, pady=10)
    # frame_batt_data.grid(row=1, column=6, sticky=tk.NW)
    # progress_bar["value"] = 50
    # lbl_progress_bar.config(text="Creating Daily graphs")
    # root.update()
    # # btn_show_batt_data = Button(frame_batt_data, text="Show Battery Data", command=lambda: show_data(batt_df))
    # # btn_show_batt_data = Button(frame_batt_data, text="Show Battery Data", command=lambda: show_data_new_win(batt_df,tab_storage))
    # # btn_show_batt_data.grid(row=0, column=0)
    #
    # frame_daily_data = LabelFrame(tab_PV_results, text="Daily Trend", padx=5, pady=5)
    # # frame_daily_data.grid(row=0, column=3, pady=10, columnspan=2, rowspan=6, sticky=tk.NW)
    # frame_daily_data.place(x=600, y=10, height=150, width=620)
    # startingday = date(int(start_year), 1, 1)
    # endingday = date(int(end_year), 12, 31)
    # cal = DateEntry(frame_daily_data, selectmode="day", year=int(start_year), month=1, day=1)
    # cal.place(x=10, y=30)

    # btn_daily_PV_prod = Button(frame_daily_data, text="Production", width=15,
    #                            command=lambda: show_daily_trend(cal.get_date(), data, 'PV production'))
    # btn_daily_PV_prod.place(x=360, y=10)
    #
    # btn_daily_irr_values = Button(frame_daily_data, text="Irradiance", width=15,
    #                               command=lambda: show_daily_trend(cal.get_date(), data, 'Irradiance'))
    # btn_daily_irr_values.place(x=120, y=10)
    #
    # btn_daily_temp_values = Button(frame_daily_data, text="Cell Temperature", width=15,
    #                                command=lambda: show_daily_trend(cal.get_date(), data, 'Cell Temperature'))
    # btn_daily_temp_values.place(x=240, y=10)
    #
    # btn_daily_grid_exchange = Button(frame_daily_data, text="Grid Exchange", width=15,
    #                                  command=lambda: show_daily_trend(cal.get_date(), data, 'Grid Exchange'))
    # btn_daily_grid_exchange.place(x=480, y=10)
    #
    # btn_monthly_trend = Button(frame_daily_data, text="Daily Production", width=15,
    #                            command=lambda: show_daily_trend_ofmonth(cal.get_date(), data))
    # btn_monthly_trend.place(x=120, y=40)
    #
    # btn_monthly_grid_ex = Button(frame_daily_data, text="Daily Grid Exchange",
    #                              command=lambda: show_grid_exchange_ofmonth(cal.get_date(), data))
    # btn_monthly_grid_ex.place(x=360, y=40)
    #
    # btn_supply_load = Button(frame_daily_data, text="Load Supply", width=15,
    #                          command=lambda: plot_pieChart(cal.get_date(), data, "Load Supply"))
    # btn_supply_load.place(x=240, y=40)
    #
    # btn_energy_gen = Button(frame_daily_data, text="Energy Generation", width=15,
    #                         command=lambda: plot_pieChart(cal.get_date(), data, "Energy Generation"))
    # btn_energy_gen.place(x=480, y=40)

    # appending yearly PV production in PV results tab ####################################################
    # lbl_year =Label(frame_yearly_data, text="Year",font=("Helvatical bold", 8))
    # lbl_year.grid(row=0, column=0)
    # lbl_Value = Label(frame_yearly_data, text="Specific PV Production "+ lbl_SI_PV_prod,font=("Helvatical bold", 8))
    # lbl_Value.grid(row=0, column=1)
    # for i in range(0, len(yearly_data_df)):
    #     str_year = int(start_year)+i
    #     str_value = yearly_data_df[i]
    #
    #     row_pos = i +1
    #     Label(frame_yearly_data, text=str_year).grid(row=row_pos, column=0)
    #     Label(frame_yearly_data, text=str_value).grid(row=row_pos, column=1)
    ########################################################################################################
    # show_data_new_win_XY(data, tab_Allresults, 600, 1200, 10, 30)  # inserting battery data in battery storage tab

    ########################################################################################################Commenting on 21-07
    # WT_id = 0
    # for each_col in all_WT_df.columns:
    #     selected_turbine_name_int_WT2 = para_WT_df.columns[WT_id]
    #     selected_turbine_nom_pow = para_WT_df[selected_turbine_name_int_WT2][1]
    #     all_WT_df[selected_turbine_name_int_WT2] = all_WT_df[selected_turbine_name_int_WT2].div(selected_turbine_nom_pow)
    #     WT_id +=1
    #
    # all_WT_df = all_WT_df.div(nb_WT)
    # float_WT_ageing = float(ent_ageing_WT.get())/100
    # lst_25_yrs = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25]
    # all_WT_25yrs_df = replicate_data_for_25yrs(total_years+1,all_WT_df )
    # production_with_ageing(all_WT_25yrs_df,float_WT_ageing)
    # all_WT_25yrs_df.insert(0,"Years", lst_25_yrs)
    # show_data_new_win(all_WT_25yrs_df,tab_WT_results,500,800)#inserting Wind turbine data in Wind turbine tab
    ########################################################################################################Commenting on 21-07 ends
    # appending yearly WT production in WT results tab#########################################################
    #     frame_all_wt = LabelFrame(tab_WT_results, text="All WT results")
    #     frame_all_wt.grid(row=0, column=0, sticky=tk.NW)
    #
    #     lbl_year_WT = Label(frame_all_wt, text="Year", font=("Helvatical bold", 8))
    #     lbl_year_WT.grid(row=0, column=0)
    #
    #
    #     for i in range(0, len(all_WT_df)):
    #         row_pos = i + 1
    #         str_year = int(start_year) + i
    #         Label(frame_all_wt, text=str_year).grid(row=row_pos, column=0)
    #
    #     try:
    #         for int_WT in range(0, len(all_WT_df.columns)-1):
    #             Wt_prod = all_WT_df.iloc[:, int_WT ]
    #             WT_name = all_WT_df.columns[int_WT ]
    #             col_pos = int_WT + 1
    #             lbl_Value_WT = Label(frame_all_wt, text=WT_name + "" + lbl_SI_PV_prod, font=("Helvatical bold", 8))
    #             lbl_Value_WT.grid(row=0, column=col_pos)
    #             selected_turbine_name_int_WT2 = para_WT_df.columns[int_WT]
    #             selected_turbine_para_int_WT2 = para_WT_df[selected_turbine_name_int_WT2]
    #             nominal_pow_WT = selected_turbine_para_int_WT2[1]
    #
    #             for i in range(0, len(Wt_prod)):
    #                 # str_year = int(start_year)+i
    #                 str_value = Wt_prod[i]/nominal_pow_WT
    #                 row_pos = i + 1
    #                 # Label(frame_all_wt, text=str_year).grid(row=row_pos, column=0)
    #                 Label(frame_all_wt, text=str_value).grid(row=row_pos, column=col_pos)
    #     except:
    #         messagebox.showerror("Error","Python found error while appending the Wind turbine production")
    # progress_bar["value"] = 55
    # lbl_progress_bar.config(text="Calculating Results")
    # root.update()
    try:
        # results_df = produce_results_df(data, sel_WT_nom_pow)
        results_df = produce_results_df_for_opt(data, nominal_pow, sel_WT_nom_pow, opt_nb_WT, batt_capacity)
    except:
        messagebox.showerror("error", "error while producing results df")
    #
    # update_results_in_main_wind(data, results_df,nominal_pow,)
    ###########################################################################################################

    # updating values of results DF in Optimization window
    global lst_result_values_opt
    lst_result_values_opt = []

    all_data_sum = data.resample('Y').sum()
    col_load_opt = all_data_sum["Load(kW)"]
    col_absorption_frm_grid_opt = all_data_sum["Absorption from grid(kW)"]
    annual_self_sufficiency = ((col_load_opt[0] - col_absorption_frm_grid_opt[0]) / col_load_opt[0]) * 100
    # col_pv_prod_without_ps = all_data_sum[""]
    col_pv_prod_with_ps = all_data_sum['Production of PV after peak shaving(kW)'].div(1000)
    # col_WT_prod_without_ps= all_data_sum[""]
    col_pv_prod_without_ps = all_data_sum["Power in ALternate Current(kW)"].div(1000)
    if nominal_pow != 0:
        col_pv_productivity = col_pv_prod_without_ps.div(nominal_pow / 1000)
    else:
        col_pv_productivity = col_pv_prod_without_ps
    col_WT_prod_with_ps = all_data_sum["Production of WT after peak shaving(kW)"].div(1000)
    results_col = results_df[str(start_year)]
    if nominal_pow != 0:
        productivity_pv_with_ps = (col_pv_prod_with_ps[0] / nominal_pow) * 1000
    else:
        productivity_pv_with_ps = 0

    if sel_WT_nom_pow != 0:
        productivity_WT_with_ps = (col_WT_prod_with_ps[0] / sel_WT_nom_pow) * 1000
    else:
        productivity_WT_with_ps = 0

    ent_pv_prodcutivity_opt.configure(state="normal")
    ent_pv_prodcutivity_opt.delete(0, END)
    ent_pv_prodcutivity_opt.insert(END, str(round(col_pv_productivity[0])))
    # lst_result_values.append(round(col_pv_productivity[0]))
    #
    ent_pv_annual_prod_opt.configure(state="normal")
    ent_pv_annual_prod_opt.delete(0, END)
    ent_pv_annual_prod_opt.insert(END, str(round(col_pv_prod_without_ps[0], 3)))
    lst_result_values_opt.append(round(col_pv_prod_without_ps[0]))
    #
    ent_pv_productivity_with_ps_opt.configure(state="normal")
    ent_pv_productivity_with_ps_opt.delete(0, END)
    ent_pv_productivity_with_ps_opt.insert(END, str(round(productivity_pv_with_ps)))
    # lst_result_values.append(round(productivity_pv_with_ps))
    #
    ent_Annual_pv_prod_with_ps_opt.configure(state="normal")
    ent_Annual_pv_prod_with_ps_opt.delete(0, END)
    ent_Annual_pv_prod_with_ps_opt.insert(END, str(round(col_pv_prod_with_ps[0], 3)))
    lst_result_values_opt.append(round(col_pv_prod_with_ps[0]))
    #
    ent_WT_prodcutivity_opt.configure(state="normal")
    ent_WT_prodcutivity_opt.delete(0, END)
    ent_WT_prodcutivity_opt.insert(END, str(results_col[10]))
    # # lst_result_values.append(results_col[10])
    #
    ent_WT_annual_prod_opt.configure(state="normal")
    ent_WT_annual_prod_opt.delete(0, END)
    ent_WT_annual_prod_opt.insert(END, str(float(results_col[10]) * nb_WT * sel_WT_nom_pow / 1000))
    lst_result_values_opt.append(float(results_col[10]) * nb_WT * sel_WT_nom_pow / 1000)
    #
    ent_WT_productivity_with_ps_opt.configure(state="normal")
    ent_WT_productivity_with_ps_opt.delete(0, END)
    if nb_WT != 0:
        ent_WT_productivity_with_ps_opt.insert(END, str(round(productivity_WT_with_ps / nb_WT)))
        # lst_result_values.append(round(productivity_WT_with_ps/nb_WT))
    else:
        ent_WT_productivity_with_ps_opt.insert(END, str(0))
        # lst_result_values.append(0)
    #
    ent_Annual_WT_prod_with_ps_opt.configure(state="normal")
    ent_Annual_WT_prod_with_ps_opt.delete(0, END)
    ent_Annual_WT_prod_with_ps_opt.insert(END, str(round(col_WT_prod_with_ps[0])))
    lst_result_values_opt.append(round(col_WT_prod_with_ps[0]))
    #
    ent_dischrged_energy_opt.configure(state="normal")
    ent_dischrged_energy_opt.delete(0, END)
    ent_dischrged_energy_opt.insert(END, str(results_col[20]))
    lst_result_values_opt.append(results_col[20])
    #
    ent_chrged_energy_opt.configure(state="normal")
    ent_chrged_energy_opt.delete(0, END)
    ent_chrged_energy_opt.insert(END, str(results_col[19]))
    lst_result_values_opt.append(results_col[19])
    #
    ent_self_consumption_opt.configure(state="normal")
    ent_self_consumption_opt.delete(0, END)
    ent_self_consumption_opt.insert(END, str(results_col[22]))
    # # lst_result_values.append(results_col[22])
    #
    ent_self_sufficiency_opt.configure(state="normal")
    ent_self_sufficiency_opt.delete(0, END)
    ent_self_sufficiency_opt.insert(END, str(results_col[23]))
    # # lst_result_values.append(results_col[23])
    #
    ent_absorption_from_grid_opt.configure(state="normal")
    ent_absorption_from_grid_opt.delete(0, END)
    ent_absorption_from_grid_opt.insert(END, str(results_col[24]))
    # # lst_result_values.append(results_col[24])
    #
    ent_injection_to_grid_opt.configure(state="normal")
    ent_injection_to_grid_opt.delete(0, END)
    ent_injection_to_grid_opt.insert(END, str(results_col[25]))
    # # lst_result_values.append(results_col[25])
    #
    ent_annual_load_opt.configure(state="normal")
    ent_annual_load_opt.delete(0, END)
    ent_annual_load_opt.insert(END, str(results_col[14]))
    lst_result_values_opt.append(results_col[14])
    #
    var_results_SI_units_opt.set(lst_SI_units_opt[1])
    # #######################################################################################################

    ########################################################################################################
    data_sum_fr_25 = data.resample("Y").sum()

    yrs_df = replicate_data_for_25yrs(total_years + 1, data_sum_fr_25)
    ###########################################################################################################

    # creatio of NPV table
    npv_df = pd.DataFrame()
    # elec_sp_1 = float(ent_ele_sp.get())
    # elec_sp_2 = float(ent_ele_sp2.get())
    ageing = float(ent_ageing.get()) / 100
    ageing_WT = float(ent_ageing_WT.get()) / 100
    # elec_sp_1_WT = float(ent_ele_sp_WT.get())
    # elec_sp_2_WT = float(ent_ele_sp2_WT.get())
    price_self_consumed_energy = float(ent_value_SC_energy.get())
    lst_NPV_PV_prod = yrs_df["Production of PV after peak shaving(kW)"].to_list()
    lst_NPV_PV_annual_inj = yrs_df["Injection in the grid from PV(kW)"].to_list()
    lst_NPV_self_consumption_PV = yrs_df["Self-consumtion for PV(kW)"].to_list()
    lst_NPV_self_consumption_WT = yrs_df["Self-consumtion for WT(kW)"].to_list()
    lst_NPV_PV_prod_age = []
    lst_NPV_PV_annual_inj_age = []
    lst_NPV_self_consumption_PV_age = []
    lst_price_elec_nt_purchased = []
    lst_grid_inj_price = []
    lst_selling_price = []
    lst_selling_price_WT = []
    lst_grid_inj_price_WT = []
    lst_price_self_consumption_PV = []
    lst_price_self_consumption_WT = []
    lst_NPV_years = []
    lst_grid_inj_price.append(0)
    lst_NPV_PV_prod_age.append(0)
    lst_NPV_PV_annual_inj_age.append(0)
    lst_selling_price.append(0)
    lst_price_self_consumption_PV.append(0)
    lst_NPV_self_consumption_PV_age.append(0)
    lst_price_elec_nt_purchased.append(0)
    lst_grid_inj_price_WT.append(0)
    lst_selling_price_WT.append(0)
    lst_price_self_consumption_WT.append(0)

    lst_NPV_WT_prod = yrs_df["Production of WT after peak shaving(kW)"].to_list()
    lst_NPV_WT_annual_inj = yrs_df["Injection in the grid from WT(kW)"].to_list()

    lst_NPV_WT_prod_age = []
    lst_NPV_WT_annual_inj_age = []
    lst_NPV_self_consumption_WT_age = []

    lst_NPV_ann_inj = yrs_df["Injection in grid with Peak shaving(kW)"].to_list()
    lst_NPV_ann_abs = yrs_df["Absorption from grid(kW)"].to_list()

    lst_NPV_WT_prod_age.append(0)
    lst_NPV_WT_annual_inj_age.append(0)
    lst_NPV_self_consumption_WT_age.append(0)

    if (lst_NPV_ann_inj[0] - lst_NPV_ann_abs[0]) > 0:
        annual_surplus = (lst_NPV_ann_inj[0] - lst_NPV_ann_abs[0]) / 1000
    else:
        annual_surplus = 0

    # To find battery lifetime
    lst_negetive_cash_flow_batt = []
    lst_tax_red_st = []

    batt_discharge_NPV = yrs_df["Battery Discharge(kW)"][0]
    nb_of_cycles_by_batt = batt_discharge_NPV / batt_capacity
    if nb_of_cycles_by_batt != 0:
        theoritical_lifetime = 10000 / nb_of_cycles_by_batt
    else:
        theoritical_lifetime = 1000000

    max_lifetime = min(10, theoritical_lifetime)
    validity_of_batt = int(ent_val_ST.get())
    investment_cost_batt = float(ent_ins_cost_ST.get())

    ##########################################################################################################################
    # entry_set_and_disable(ent_size, var_ent_size, nominal_pow / 1000)
    # entry_set_and_disable(ent_prod, var_ent_prod, round(lst_NPV_PV_prod[0] / 1000))
    # entry_set_and_disable(ent_grid_inj, var_ent_grid_inj, lst_NPV_PV_annual_inj[0] / 1000)
    # entry_set_and_disable(ent_life, var_ent_life, 25)
    # entry_set_and_disable(ent_spec_prod, var_ent_spec_prod, results_col[8])

    # entry_set_and_disable(ent_size_WT, var_ent_size_WT, sel_WT_nom_pow * nb_WT / 1000)
    # entry_set_and_disable(ent_prod_WT, var_ent_prod_WT, round(lst_NPV_WT_prod[0] / 1000))
    # entry_set_and_disable(ent_grid_inj_WT, var_ent_grid_inj_WT, lst_NPV_WT_annual_inj[0] / 1000)
    # entry_set_and_disable(ent_life_WT, var_ent_life_WT, 25)
    # entry_set_and_disable(ent_spec_prod_WT, var_ent_spec_prod_WT, ent_WT_productivity_with_ps.get())

    # entry_set_and_disable(ent_size_ST, var_ent_size_ST, batt_capacity / 1000)
    # entry_set_and_disable(ent_SC_energy_PV, var_ent_SC_energy_PV, round(lst_NPV_self_consumption_PV[0] / 1000))
    # entry_set_and_disable(ent_SC_energy_WT, var_ent_SC_energy_WT, round(lst_NPV_self_consumption_WT[0] / 1000))

    # entry_set_and_disable(ent_annual_surplus, var_ent_annual_surplus, annual_surplus)
    # entry_set_and_disable(ent_NM_energy, var_ent_NM_energy, min(lst_NPV_ann_inj[0], lst_NPV_ann_abs[0]) / 1000)
    #######################################################################################################################
    for i in range(0, 26):
        # lst_NPV_years.append(int(start_year)+i)
        lst_NPV_years.append(i)
    for i in range(0, 25):
        pv_prod_after_ageing = lst_NPV_PV_prod[i] * (1 - ageing * i) / 1000
        lst_NPV_PV_prod_age.append(pv_prod_after_ageing)
        NPV_PV_annual_inj_age = lst_NPV_PV_annual_inj[i] * (1 - ageing * i) / 1000
        lst_NPV_PV_annual_inj_age.append(NPV_PV_annual_inj_age)
        Wt_prod_after_ageing = lst_NPV_WT_prod[i] * (1 - ageing_WT * i) / 1000
        lst_NPV_WT_prod_age.append(Wt_prod_after_ageing)
        NPV_WT_annual_inj_age = lst_NPV_WT_annual_inj[i] * (1 - ageing_WT * i) / 1000
        lst_NPV_WT_annual_inj_age.append(NPV_WT_annual_inj_age)

    # for i in range(0, 25):
    #     if i < 5:
    #         lst_selling_price.append(elec_sp_1)
    #         lst_selling_price_WT.append(elec_sp_1_WT)
    #     else:
    #         lst_selling_price.append(elec_sp_2)
    #         lst_selling_price_WT.append(elec_sp_2_WT)
    for i in range(0, 25):
        lst_selling_price.append(float(lst_var_sp_pv[i]))
        lst_selling_price_WT.append(float(lst_var_sp_wt[i]))

    # lst_selling_price = lst_var_sp_pv
    # lst_selling_price_WT = lst_var_sp_wt
    for i in range(0, 25):
        if var_dd_SC.get() == "Yes":
            grid_inj_price = lst_NPV_PV_annual_inj_age[i] * lst_selling_price[i] * 1000
            grid_inj_price_WT = lst_NPV_WT_annual_inj_age[i] * lst_selling_price_WT[i] * 1000
        else:
            grid_inj_price = lst_NPV_PV_prod_age[i] * lst_selling_price[i] * 1000
            grid_inj_price_WT = lst_NPV_WT_prod_age[i] * lst_selling_price_WT[i] * 1000

        lst_grid_inj_price.append(grid_inj_price)
        lst_grid_inj_price_WT.append(grid_inj_price_WT)

    npv_df.insert(0, "Years", lst_NPV_years)
    npv_df.insert(1, "Grid injection of PV($/y)", lst_grid_inj_price)
    npv_df.insert(2, "PV production (MWh/y)", lst_NPV_PV_prod_age)
    npv_df.insert(3, "PV annual injection (MWh/y)", lst_NPV_PV_annual_inj_age)
    npv_df.insert(4, "Price of sold  electricity from PV ($/kWh)", lst_selling_price)

    for i in range(0, 25):
        lst_price_elec_nt_purchased.append(price_self_consumed_energy)
        if var_dd_SC.get() == "Yes":
            NPV_self_consumption_PV_age = lst_NPV_self_consumption_PV[i] * (1 - ageing * i) / 1000
            lst_NPV_self_consumption_PV_age.append(NPV_self_consumption_PV_age)
            NPV_self_consumption_WT_age = lst_NPV_self_consumption_WT[i] * (1 - ageing_WT * i) / 1000
            lst_NPV_self_consumption_WT_age.append(NPV_self_consumption_WT_age)
        else:
            lst_NPV_self_consumption_PV_age.append(0)
            lst_NPV_self_consumption_WT_age.append(0)

        total_price_self_consumption_PV = lst_price_elec_nt_purchased[i + 1] * lst_NPV_self_consumption_PV_age[
            i + 1] * 1000
        lst_price_self_consumption_PV.append(total_price_self_consumption_PV)
        total_price_self_consumption_WT = lst_price_elec_nt_purchased[i + 1] * lst_NPV_self_consumption_WT_age[
            i + 1] * 1000
        lst_price_self_consumption_WT.append(total_price_self_consumption_WT)

    npv_df.insert(5, "Self Consumption of PV($/y)", lst_price_self_consumption_PV)
    npv_df.insert(6, "PV annual self consumption (MWh/y)", lst_NPV_self_consumption_PV_age)
    npv_df.insert(7, "Price of  electricity not purchased from the grid due to PV self consumption ($/kWh)",
                  lst_price_elec_nt_purchased)

    no_of_years_tax_PV = float(ent_val_PV.get())
    no_of_years_tax_WT = float(ent_val_WT.get())

    # pv_size = float(ent_size.get()) * 1000
    pv_size = nominal_pow
    # wt_size = float(ent_size_WT.get()) * 1000
    wt_size = sel_WT_nom_pow * nb_WT

    tax_reduction_pv = float(ent_tax_red_PV.get())
    tax_reduction_wt = float(ent_tax_red_WT.get())

    investment_cost_PV = float(ent_ins_cost.get())
    investment_cost_WT = float(ent_ins_cost_WT.get())
    investment_cost_ST = float(ent_ins_cost_ST.get())

    OandM_cost_PV = float(ent_om.get())
    OandM_cost_WT = float(ent_om_WT.get())

    total_inv_cost_PV = investment_cost_PV * pv_size
    total_inv_cost_WT = investment_cost_WT * wt_size
    total_inv_cost_ST = investment_cost_ST * batt_capacity

    # lbl_tax_red_PV_cost.configure(text="$" + str(total_inv_cost_PV))
    # lbl_tax_red_WT_cost.configure(text="$" + str(total_inv_cost_WT))
    # lbl_tax_red_ST_cost.configure(text="$" + str(total_inv_cost_ST))

    lst_total_tax_reduction_PV = []
    lst_total_positive_cash_flow_PV = []
    lst_total_negetive_cash_flow_PV = []
    lst_total_tax_reduction_WT = []
    lst_total_positive_cash_flow_WT = []
    lst_total_negetive_cash_flow_WT = []

    # total_negetive_cash_flow_PV = -total_inv_cost_PV
    # total_negetive_cash_flow_WT = - total_inv_cost_WT

    lst_total_tax_reduction_PV.append(0)
    lst_total_positive_cash_flow_PV.append(0)
    lst_total_tax_reduction_WT.append(0)
    lst_total_positive_cash_flow_WT.append(0)
    lst_total_negetive_cash_flow_PV.append(-total_inv_cost_PV)
    lst_total_negetive_cash_flow_WT.append(- total_inv_cost_WT)

    for i in range(0, 25):
        if i < no_of_years_tax_PV:
            total_tax_reduction_PV = tax_reduction_pv * total_inv_cost_PV / 100
        else:
            total_tax_reduction_PV = 0

        if i < no_of_years_tax_WT:
            total_tax_reduction_WT = tax_reduction_wt * total_inv_cost_WT / 100
        else:
            total_tax_reduction_WT = 0

        lst_total_tax_reduction_PV.append(total_tax_reduction_PV)
        total_positive_cash_flow_PV = lst_total_tax_reduction_PV[i + 1] + lst_price_self_consumption_PV[i + 1] + \
                                      lst_grid_inj_price[i + 1]
        lst_total_positive_cash_flow_PV.append(total_positive_cash_flow_PV)

        lst_total_tax_reduction_WT.append(total_tax_reduction_WT)
        total_positive_cash_flow_WT = lst_total_tax_reduction_WT[i + 1] + lst_price_self_consumption_WT[i + 1] + \
                                      lst_grid_inj_price_WT[i + 1]
        lst_total_positive_cash_flow_WT.append(total_positive_cash_flow_WT)

        total_negetive_cash_flow_PV = -OandM_cost_PV * pv_size
        # total_negetive_cash_flow_WT= - OandM_cost_WT*wt_size
        total_negetive_cash_flow_WT = - OandM_cost_WT * lst_NPV_WT_prod_age[1] * 1000

        lst_total_negetive_cash_flow_PV.append(total_negetive_cash_flow_PV)
        lst_total_negetive_cash_flow_WT.append(total_negetive_cash_flow_WT)

    npv_df.insert(8, "Tax reduction for PV ($/y)", lst_total_tax_reduction_PV)
    npv_df.insert(9, "Positive Cash Flows of PV  ($/y)", lst_total_positive_cash_flow_PV)
    npv_df.insert(10, "Negative Cash Flows of PV ($/y)", lst_total_negetive_cash_flow_PV)

    npv_df.insert(11, "Grid injection of WT ($/y)", lst_grid_inj_price_WT)
    npv_df.insert(12, "WT production (MWh/y)", lst_NPV_WT_prod_age)
    npv_df.insert(13, "WT annual injection (MWh/y)", lst_NPV_WT_annual_inj_age)
    npv_df.insert(14, "Price of sold  electricity from WT ($/kWh)", lst_selling_price_WT)

    npv_df.insert(15, "Self Consumption of WT($/y)", lst_price_self_consumption_WT)
    npv_df.insert(16, "WT annual self consumption (MWh/y)", lst_NPV_self_consumption_WT_age)
    npv_df.insert(17, "Price of  electricity not purchased from the grid due to WT self consumption ($/kWh)",
                  lst_price_elec_nt_purchased)

    npv_df.insert(18, "Tax reduction for WT ($/y)", lst_total_tax_reduction_WT)
    npv_df.insert(19, "Positive Cash Flows of WT ($/y)", lst_total_positive_cash_flow_WT)
    npv_df.insert(20, "Negative Cash Flows of WT ($/y)", lst_total_negetive_cash_flow_WT)

    for i in range(0, 26):
        if i / max_lifetime == int(i / int(ent_val_ST.get())):
            batt_expense = -batt_capacity * investment_cost_batt
        else:
            batt_expense = 0

        lst_negetive_cash_flow_batt.append(batt_expense)

    lst_tax_red_st.append(0)
    for i in range(0, 25):
        if i < int(ent_val_ST.get()):
            tax_reduction_batt = float(ent_tax_red_ST.get()) * total_inv_cost_ST / 100
        else:
            tax_reduction_batt = 0
        lst_tax_red_st.append(tax_reduction_batt)

    npv_df.insert(21, "Negative cash flow related to battery investment/replacement ($/y)", lst_negetive_cash_flow_batt)
    npv_df.insert(22, "Tax reduction for storage", lst_tax_red_st)

    lst_total_positive_cash_flow = []
    lst_total_negetive_cash_flow = []
    lst_not_actualised_cash_flow = []
    lst_total_actual_cash_flow = []
    lst_NPV = []
    discount_rate = float(ent_rate.get()) / 100

    for i in range(0, 26):
        total_positive_cash_flow = lst_total_positive_cash_flow_PV[i] + lst_total_positive_cash_flow_WT[i] + \
                                   lst_tax_red_st[i]
        total_negetive_cash_flow = lst_negetive_cash_flow_batt[i] + lst_total_negetive_cash_flow_WT[i] + \
                                   lst_total_negetive_cash_flow_PV[i]
        not_actualised_cash_flow = total_positive_cash_flow + total_negetive_cash_flow
        total_actual_cash_flow = not_actualised_cash_flow / ((1 + discount_rate) ** i)

        if i == 0:
            NPV = total_actual_cash_flow
        else:
            NPV = lst_NPV[i - 1] + total_actual_cash_flow

        lst_total_positive_cash_flow.append(total_positive_cash_flow)
        lst_total_negetive_cash_flow.append(total_negetive_cash_flow)
        lst_not_actualised_cash_flow.append(not_actualised_cash_flow)
        lst_total_actual_cash_flow.append(total_actual_cash_flow)
        lst_NPV.append(NPV)

    npv_df.insert(23, "Positive Cash Flows ($/y)", lst_total_positive_cash_flow)
    npv_df.insert(24, "Negative Cash Flows ($/y)", lst_total_negetive_cash_flow)
    npv_df.insert(25, "Cash flow ( not actualized)", lst_not_actualised_cash_flow)
    npv_df.insert(26, "Total Cash Flows($/y)", lst_total_actual_cash_flow)
    npv_df.insert(27, "NPV", lst_NPV)

    IRR = npf.irr(lst_not_actualised_cash_flow) * 100
    NPV_future = lst_NPV[25] / 1000000
    PBT = sum(npv_df.NPV < 0)

    # entry_set_and_disable(ent_irr, var_ent_irr, round(IRR, 2))
    # entry_set_and_disable(ent_NPV, var_ent_NPV, round(NPV_future))
    # entry_set_and_disable(ent_PBT, var_ent_PBT, PBT)

    lst_million_dollor_NPV = [x / 1000000 for x in lst_NPV]

    # frame_NPV_graph = LabelFrame(tab_fin_ana, text="NPV data")
    # frame_NPV_graph.place(x=10, y=550, height=65, width=420)

    # btn_NPV_graph = Button(frame_NPV_graph, text="NPV Graph", command=lambda : plot_graph(lst_NPV_years,lst_million_dollor_NPV,"Years","NPV[M$]","NPV Graph" ))
    # btn_NPV_graph = Button(frame_NPV_graph, text="NPV Graph",
    #                        command=lambda: plot_bar_and_line_graph(npv_df))
    #
    # btn_NPV_graph.place(x=10, y=10)
    #
    # btn_show_NPV_data = Button(frame_NPV_graph, text="Show NPV data", command=lambda: show_data(npv_df))
    # btn_show_NPV_data.place(x=90, y=10)
    #
    # btn_export_NPV_data = Button(frame_NPV_graph, text="Export NPV data",
    #                              command=lambda: export_to_excel_with_name(npv_df, "NPV_data"))
    # btn_export_NPV_data.place(x=190, y=10)
    ########################################################################################################
    # creation of LCOE df
    lcoe_df = pd.DataFrame()
    lst_discounted_pv_prod = []
    lst_discounted_wt_prod = []
    lst_yearly_cost_PV = []
    lst_yearly_cost_WT = []
    lst_NPV_PV_farm = []
    lst_NPV_WT_farm = []

    pv_rate = float(ent_rate.get()) / 100
    wt_rate = float(ent_rate_WT.get()) / 100
    # wt_prod = float(ent_prod_WT.get()) * 1000
    wt_prod = lst_NPV_WT_prod[0]
    # lst_PV_productivity = lst_NPV_PV_prod_age / pv_size
    if pv_size != 0:

        lst_PV_productivity = [(x * 1000) / pv_size for x in lst_NPV_PV_prod_age]
    else:
        lst_PV_productivity = [x * 0 for x in lst_NPV_PV_prod_age]
    # lst_WT_productivity = lst_NPV_WT_prod_age/ wt_size
    if wt_size != 0:
        lst_WT_productivity = [(x * 1000) / wt_size for x in lst_NPV_WT_prod_age]
    else:
        lst_WT_productivity = [x * 0 for x in lst_NPV_WT_prod_age]

    for i in range(0, 26):
        discounted_pv_prod = lst_NPV_PV_prod_age[i] / ((1 + pv_rate) ** i)
        discounted_wt_prod = lst_NPV_WT_prod_age[i] / ((1 + wt_rate) ** i)
        lst_discounted_pv_prod.append(discounted_pv_prod)
        lst_discounted_wt_prod.append(discounted_wt_prod)

        if i == 0:
            yearly_cost_PV = investment_cost_PV * pv_size
            NPV_PV_farm = yearly_cost_PV
            yearly_cost_WT = investment_cost_WT * wt_size
            NPV_WT_farm = yearly_cost_WT
        else:
            yearly_cost_PV = OandM_cost_PV * pv_size
            NPV_PV_farm = yearly_cost_PV / ((1 + pv_rate) ** i)
            yearly_cost_WT = OandM_cost_WT * wt_prod
            NPV_WT_farm = yearly_cost_WT / ((1 + wt_rate) ** i)
        lst_yearly_cost_PV.append(yearly_cost_PV)
        lst_NPV_PV_farm.append(NPV_PV_farm)
        lst_yearly_cost_WT.append(yearly_cost_WT)
        lst_NPV_WT_farm.append(NPV_WT_farm)

    lcoe_df.insert(0, "Years", lst_NPV_years)
    lcoe_df.insert(1, "NPV of Total Cost of PV plant($/y)", lst_NPV_PV_farm)
    lcoe_df.insert(2, "Yearly Cost of PV plant($/y)", lst_yearly_cost_PV)
    lcoe_df.insert(3, "Discounted PV Production of PV plant (MWh/y)", lst_discounted_pv_prod)
    lcoe_df.insert(4, "PV annual production (MWh/y)", lst_NPV_PV_prod_age)
    lcoe_df.insert(5, "PV Specific Production (kWh/kW/y)", lst_PV_productivity)
    lcoe_df.insert(6, "NPV of Total Cost of WT farm($/y)", lst_NPV_WT_farm)
    lcoe_df.insert(7, "Yearly Cost of WT farm($/y)", lst_yearly_cost_WT)
    lcoe_df.insert(8, "Discounted PV Production of WT farm(MWh/y)", lst_discounted_wt_prod)
    lcoe_df.insert(9, "WT annual production (MWh/y)", lst_NPV_WT_prod_age)
    lcoe_df.insert(10, "WT Specific Production (kWh/kW/y)", lst_WT_productivity)

    sum_NPV_PV_farm = lcoe_df["NPV of Total Cost of PV plant($/y)"].sum()
    sum_discounted_pv_prod = lcoe_df["Discounted PV Production of PV plant (MWh/y)"].sum()

    sum_NPV_WT_farm = lcoe_df["NPV of Total Cost of WT farm($/y)"].sum()
    sum_discounted_wt_prod = lcoe_df["Discounted PV Production of WT farm(MWh/y)"].sum()

    lcoe_PV = sum_NPV_PV_farm / (sum_discounted_pv_prod * 1000)
    lcoe_WT = sum_NPV_WT_farm / (sum_discounted_wt_prod * 1000)

    loce_total_plant = (sum_NPV_PV_farm + sum_NPV_WT_farm) / ((sum_discounted_wt_prod + sum_discounted_pv_prod) * 1000)
    rounded_loce_total_plant = round(loce_total_plant, 2)

    # entry_set_and_disable(ent_LCOE, var_ent_LCOE, rounded_loce_total_plant)
    ########################################################################################################
    newWindow.destroy()
    ss = results_col[23]
    lst_result = [annual_self_sufficiency, IRR, NPV_future, PBT, rounded_loce_total_plant]
    # messagebox.showinfo("Completed", "Download and calculation of data completed")
    return lst_result


def save_sp_pv(lst_ent_var_sp):
    lst_var_sp_pv.clear()
    for i in range(0, len(lst_ent_var_sp)):
        sp = lst_ent_var_sp[i].get()
        lst_var_sp_pv.append(sp)
    return


def save_sp_wt(lst_ent_var_sp):
    lst_var_sp_wt.clear()
    for i in range(0, len(lst_ent_var_sp)):
        sp = float(lst_ent_var_sp[i].get())
        lst_var_sp_wt.append(sp)
    return



def variable_selling_price_wt():
    try:
        sp_window = Toplevel(root)
        sp_window.title("Variable selling price of WT")
        sp_window.geometry('490x550')

        global lst_var_sp_wt
        if len(lst_var_sp_wt) == 0:
            lst_var_sp_wt = [0.04]*25

        frm_var_sp = LabelFrame(sp_window, width = 200, height=520)
        frm_var_sp.place(x=10, y=10)
        frm_var_sp2 = LabelFrame(sp_window, width=200, height=520)
        frm_var_sp2.place(x=250, y=10)

        lbl_sp_1 = Label(frm_var_sp, text="Year 1")
        lbl_sp_2 = Label(frm_var_sp, text="Year 2")
        lbl_sp_3 = Label(frm_var_sp, text="Year 3")
        lbl_sp_4 = Label(frm_var_sp, text="Year 4")
        lbl_sp_5 = Label(frm_var_sp, text="Year 5")
        lbl_sp_6 = Label(frm_var_sp, text="Year 6")
        lbl_sp_7 = Label(frm_var_sp, text="Year 7")
        lbl_sp_8 = Label(frm_var_sp, text="Year 8")
        lbl_sp_9 = Label(frm_var_sp, text="Year 9")
        lbl_sp_10 = Label(frm_var_sp, text="Year 10")
        lbl_sp_11 = Label(frm_var_sp, text="Year 11")
        lbl_sp_12 = Label(frm_var_sp, text="Year 12")
        lbl_sp_13 = Label(frm_var_sp, text="Year 13")
        lbl_sp_14 = Label(frm_var_sp, text="Year 14")
        lbl_sp_15 = Label(frm_var_sp, text="Year 15")

        lbl_sp_16 = Label(frm_var_sp2, text="Year 16")
        lbl_sp_17 = Label(frm_var_sp2, text="Year 17")
        lbl_sp_18 = Label(frm_var_sp2, text="Year 18")
        lbl_sp_19 = Label(frm_var_sp2, text="Year 19")
        lbl_sp_20 = Label(frm_var_sp2, text="Year 20")
        lbl_sp_21 = Label(frm_var_sp2, text="Year 21")
        lbl_sp_22 = Label(frm_var_sp2, text="Year 22")
        lbl_sp_23 = Label(frm_var_sp2, text="Year 23")
        lbl_sp_24 = Label(frm_var_sp2, text="Year 24")
        lbl_sp_25 = Label(frm_var_sp2, text="Year 25")

        lbl_sp_1.place(x=10, y=10)
        lbl_sp_2.place(x=10, y=40)
        lbl_sp_3.place(x=10, y=70)
        lbl_sp_4.place(x=10, y=100)
        lbl_sp_5.place(x=10, y=130)
        lbl_sp_6.place(x=10, y=160)
        lbl_sp_7.place(x=10, y=190)
        lbl_sp_8.place(x=10, y=220)
        lbl_sp_9.place(x=10, y=250)
        lbl_sp_10.place(x=10, y=280)
        lbl_sp_11.place(x=10, y=310)
        lbl_sp_12.place(x=10, y=340)
        lbl_sp_13.place(x=10, y=370)
        lbl_sp_14.place(x=10, y=400)
        lbl_sp_15.place(x=10, y=430)

        lbl_sp_16.place(x=10, y=10)
        lbl_sp_17.place(x=10, y=40)
        lbl_sp_18.place(x=10, y=70)
        lbl_sp_19.place(x=10, y=100)
        lbl_sp_20.place(x=10, y=130)
        lbl_sp_21.place(x=10, y=160)
        lbl_sp_22.place(x=10, y=190)
        lbl_sp_23.place(x=10, y=220)
        lbl_sp_24.place(x=10, y=250)
        lbl_sp_25.place(x=10, y=280)

        ent_sp_1_wt = Entry(frm_var_sp, width=10)
        ent_sp_1_wt.insert(END, str(lst_var_sp_wt[0]))
        ent_sp_2_wt = Entry(frm_var_sp, width=10)
        ent_sp_2_wt.insert(END, str(lst_var_sp_wt[1]))
        ent_sp_3_wt = Entry(frm_var_sp, width=10)
        ent_sp_3_wt.insert(END, str(lst_var_sp_wt[2]))
        ent_sp_4_wt = Entry(frm_var_sp, width=10)
        ent_sp_4_wt.insert(END, str(lst_var_sp_wt[3]))
        ent_sp_5_wt = Entry(frm_var_sp, width=10)
        ent_sp_5_wt.insert(END, str(lst_var_sp_wt[4]))
        ent_sp_6_wt = Entry(frm_var_sp, width=10)
        ent_sp_6_wt.insert(END, str(lst_var_sp_wt[5]))
        ent_sp_7_wt = Entry(frm_var_sp, width=10)
        ent_sp_7_wt.insert(END, str(lst_var_sp_wt[6]))
        ent_sp_8_wt = Entry(frm_var_sp, width=10)
        ent_sp_8_wt.insert(END, str(lst_var_sp_wt[7]))
        ent_sp_9_wt = Entry(frm_var_sp, width=10)
        ent_sp_9_wt.insert(END, str(lst_var_sp_wt[8]))
        ent_sp_10_wt = Entry(frm_var_sp, width=10)
        ent_sp_10_wt.insert(END, str(lst_var_sp_wt[9]))
        ent_sp_11_wt = Entry(frm_var_sp, width=10)
        ent_sp_11_wt.insert(END, str(lst_var_sp_wt[10]))
        ent_sp_12_wt = Entry(frm_var_sp, width=10)
        ent_sp_12_wt.insert(END, str(lst_var_sp_wt[11]))
        ent_sp_13_wt = Entry(frm_var_sp, width=10)
        ent_sp_13_wt.insert(END, str(lst_var_sp_wt[12]))
        ent_sp_14_wt = Entry(frm_var_sp, width=10)
        ent_sp_14_wt.insert(END, str(lst_var_sp_wt[13]))
        ent_sp_15_wt = Entry(frm_var_sp, width=10)
        ent_sp_15_wt.insert(END, str(lst_var_sp_wt[14]))

        ent_sp_16_wt = Entry(frm_var_sp2, width=10)
        ent_sp_16_wt.insert(END, str(lst_var_sp_wt[15]))
        ent_sp_17_wt = Entry(frm_var_sp2, width=10)
        ent_sp_17_wt.insert(END, str(lst_var_sp_wt[16]))
        ent_sp_18_wt = Entry(frm_var_sp2, width=10)
        ent_sp_18_wt.insert(END, str(lst_var_sp_wt[17]))
        ent_sp_19_wt = Entry(frm_var_sp2, width=10)
        ent_sp_19_wt.insert(END, str(lst_var_sp_wt[18]))
        ent_sp_20_wt = Entry(frm_var_sp2, width=10)
        ent_sp_20_wt.insert(END, str(lst_var_sp_wt[19]))
        ent_sp_21_wt = Entry(frm_var_sp2, width=10)
        ent_sp_21_wt.insert(END, str(lst_var_sp_wt[20]))
        ent_sp_22_wt = Entry(frm_var_sp2, width=10)
        ent_sp_22_wt.insert(END, str(lst_var_sp_wt[21]))
        ent_sp_23_wt = Entry(frm_var_sp2, width=10)
        ent_sp_23_wt.insert(END, str(lst_var_sp_wt[22]))
        ent_sp_24_wt = Entry(frm_var_sp2, width=10)
        ent_sp_24_wt.insert(END, str(lst_var_sp_wt[23]))
        ent_sp_25_wt = Entry(frm_var_sp2, width=10)
        ent_sp_25_wt.insert(END, str(lst_var_sp_wt[24]))

        ent_sp_1_wt.place(x=60, y=10)
        ent_sp_2_wt.place(x=60, y=40)
        ent_sp_3_wt.place(x=60, y=70)
        ent_sp_4_wt.place(x=60, y=100)
        ent_sp_5_wt.place(x=60, y=130)
        ent_sp_6_wt.place(x=60, y=160)
        ent_sp_7_wt.place(x=60, y=190)
        ent_sp_8_wt.place(x=60, y=220)
        ent_sp_9_wt.place(x=60, y=250)
        ent_sp_10_wt.place(x=60, y=280)
        ent_sp_11_wt.place(x=60, y=310)
        ent_sp_12_wt.place(x=60, y=340)
        ent_sp_13_wt.place(x=60, y=370)
        ent_sp_14_wt.place(x=60, y=400)
        ent_sp_15_wt.place(x=60, y=430)

        ent_sp_16_wt.place(x=60, y=10)
        ent_sp_17_wt.place(x=60, y=40)
        ent_sp_18_wt.place(x=60, y=70)
        ent_sp_19_wt.place(x=60, y=100)
        ent_sp_20_wt.place(x=60, y=130)
        ent_sp_21_wt.place(x=60, y=160)
        ent_sp_22_wt.place(x=60, y=190)
        ent_sp_23_wt.place(x=60, y=220)
        ent_sp_24_wt.place(x=60, y=250)
        ent_sp_25_wt.place(x=60, y=280)

        lbl_sp_1_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_2_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_3_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_4_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_5_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_6_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_7_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_8_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_9_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_10_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_11_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_12_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_13_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_14_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_15_des = Label(frm_var_sp, text="$/kWh")

        lbl_sp_16_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_17_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_18_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_19_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_20_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_21_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_22_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_23_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_24_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_25_des = Label(frm_var_sp2, text="$/kWh")

        lbl_sp_1_des.place(x=130, y=10)
        lbl_sp_2_des.place(x=130, y=40)
        lbl_sp_3_des.place(x=130, y=70)
        lbl_sp_4_des.place(x=130, y=100)
        lbl_sp_5_des.place(x=130, y=130)
        lbl_sp_6_des.place(x=130, y=160)
        lbl_sp_7_des.place(x=130, y=190)
        lbl_sp_8_des.place(x=130, y=220)
        lbl_sp_9_des.place(x=130, y=250)
        lbl_sp_10_des.place(x=130, y=280)
        lbl_sp_11_des.place(x=130, y=310)
        lbl_sp_12_des.place(x=130, y=340)
        lbl_sp_13_des.place(x=130, y=370)
        lbl_sp_14_des.place(x=130, y=400)
        lbl_sp_15_des.place(x=130, y=430)

        lbl_sp_16_des.place(x=130, y=10)
        lbl_sp_17_des.place(x=130, y=40)
        lbl_sp_18_des.place(x=130, y=70)
        lbl_sp_19_des.place(x=130, y=100)
        lbl_sp_20_des.place(x=130, y=130)
        lbl_sp_21_des.place(x=130, y=160)
        lbl_sp_22_des.place(x=130, y=190)
        lbl_sp_23_des.place(x=130, y=220)
        lbl_sp_24_des.place(x=130, y=250)
        lbl_sp_25_des.place(x=130, y=280)

        lst_ent_var_sp = [ent_sp_1_wt,ent_sp_2_wt,ent_sp_3_wt,ent_sp_4_wt,ent_sp_5_wt,ent_sp_6_wt,ent_sp_7_wt,ent_sp_8_wt,ent_sp_9_wt,
                          ent_sp_10_wt,ent_sp_11_wt,ent_sp_12_wt,ent_sp_13_wt,ent_sp_14_wt,ent_sp_15_wt,
                          ent_sp_16_wt,ent_sp_17_wt,ent_sp_18_wt,ent_sp_19_wt,ent_sp_20_wt,ent_sp_21_wt,
                          ent_sp_22_wt,ent_sp_23_wt,ent_sp_24_wt,ent_sp_25_wt]
        btn_save_sp = Button(frm_var_sp, text="Save", command= lambda:save_sp_wt(lst_ent_var_sp) )
        btn_save_sp.place(x=80, y=470)
    except:
        messagebox.showerror("Error","Error while saving selling price")

def variable_selling_price_pv():
    try:
        sp_window = Toplevel(root)
        sp_window.title("Variable selling price")
        sp_window.geometry('490x550')

        global lst_var_sp_pv
        if len(lst_var_sp_pv) == 0:
            lst_var_sp_pv = [0.04]*25

        frm_var_sp = LabelFrame(sp_window, width = 200, height=520)
        frm_var_sp.place(x=10, y=10)
        frm_var_sp2 = LabelFrame(sp_window, width = 200, height=520)
        frm_var_sp2.place(x=250, y=10)

        lbl_sp_1 = Label(frm_var_sp, text="Year 1")
        lbl_sp_2 = Label(frm_var_sp, text="Year 2")
        lbl_sp_3 = Label(frm_var_sp, text="Year 3")
        lbl_sp_4 = Label(frm_var_sp, text="Year 4")
        lbl_sp_5 = Label(frm_var_sp, text="Year 5")
        lbl_sp_6 = Label(frm_var_sp, text="Year 6")
        lbl_sp_7 = Label(frm_var_sp, text="Year 7")
        lbl_sp_8 = Label(frm_var_sp, text="Year 8")
        lbl_sp_9 = Label(frm_var_sp, text="Year 9")
        lbl_sp_10 = Label(frm_var_sp, text="Year 10")
        lbl_sp_11 = Label(frm_var_sp, text="Year 11")
        lbl_sp_12 = Label(frm_var_sp, text="Year 12")
        lbl_sp_13 = Label(frm_var_sp, text="Year 13")
        lbl_sp_14 = Label(frm_var_sp, text="Year 14")
        lbl_sp_15 = Label(frm_var_sp, text="Year 15")

        lbl_sp_16 = Label(frm_var_sp2, text="Year 16")
        lbl_sp_17 = Label(frm_var_sp2, text="Year 17")
        lbl_sp_18 = Label(frm_var_sp2, text="Year 18")
        lbl_sp_19 = Label(frm_var_sp2, text="Year 19")
        lbl_sp_20 = Label(frm_var_sp2, text="Year 20")
        lbl_sp_21 = Label(frm_var_sp2, text="Year 21")
        lbl_sp_22 = Label(frm_var_sp2, text="Year 22")
        lbl_sp_23 = Label(frm_var_sp2, text="Year 23")
        lbl_sp_24 = Label(frm_var_sp2, text="Year 24")
        lbl_sp_25 = Label(frm_var_sp2, text="Year 25")

        lbl_sp_1.place(x=10, y=10)
        lbl_sp_2.place(x=10, y=40)
        lbl_sp_3.place(x=10, y=70)
        lbl_sp_4.place(x=10, y=100)
        lbl_sp_5.place(x=10, y=130)
        lbl_sp_6.place(x=10, y=160)
        lbl_sp_7.place(x=10, y=190)
        lbl_sp_8.place(x=10, y=220)
        lbl_sp_9.place(x=10, y=250)
        lbl_sp_10.place(x=10, y=280)
        lbl_sp_11.place(x=10, y=310)
        lbl_sp_12.place(x=10, y=340)
        lbl_sp_13.place(x=10, y=370)
        lbl_sp_14.place(x=10, y=400)
        lbl_sp_15.place(x=10, y=430)

        lbl_sp_16.place(x=10, y=10)
        lbl_sp_17.place(x=10, y=40)
        lbl_sp_18.place(x=10, y=70)
        lbl_sp_19.place(x=10, y=100)
        lbl_sp_20.place(x=10, y=130)
        lbl_sp_21.place(x=10, y=160)
        lbl_sp_22.place(x=10, y=190)
        lbl_sp_23.place(x=10, y=220)
        lbl_sp_24.place(x=10, y=250)
        lbl_sp_25.place(x=10, y=280)

        ent_sp_1 = Entry(frm_var_sp, width=10)
        ent_sp_1.insert(END, str(lst_var_sp_pv[0]))
        ent_sp_2 = Entry(frm_var_sp, width=10)
        ent_sp_2.insert(END, str(lst_var_sp_pv[1]))
        ent_sp_3 = Entry(frm_var_sp, width=10)
        ent_sp_3.insert(END, str(lst_var_sp_pv[2]))
        ent_sp_4 = Entry(frm_var_sp, width=10)
        ent_sp_4.insert(END, str(lst_var_sp_pv[3]))
        ent_sp_5 = Entry(frm_var_sp, width=10)
        ent_sp_5.insert(END, str(lst_var_sp_pv[4]))
        ent_sp_6 = Entry(frm_var_sp, width=10)
        ent_sp_6.insert(END, str(lst_var_sp_pv[5]))
        ent_sp_7 = Entry(frm_var_sp, width=10)
        ent_sp_7.insert(END, str(lst_var_sp_pv[6]))
        ent_sp_8 = Entry(frm_var_sp, width=10)
        ent_sp_8.insert(END, str(lst_var_sp_pv[7]))
        ent_sp_9 = Entry(frm_var_sp, width=10)
        ent_sp_9.insert(END, str(lst_var_sp_pv[8]))
        ent_sp_10 = Entry(frm_var_sp, width=10)
        ent_sp_10.insert(END, str(lst_var_sp_pv[9]))
        ent_sp_11 = Entry(frm_var_sp, width=10)
        ent_sp_11.insert(END, str(lst_var_sp_pv[10]))
        ent_sp_12 = Entry(frm_var_sp, width=10)
        ent_sp_12.insert(END, str(lst_var_sp_pv[11]))
        ent_sp_13 = Entry(frm_var_sp, width=10)
        ent_sp_13.insert(END, str(lst_var_sp_pv[12]))
        ent_sp_14 = Entry(frm_var_sp, width=10)
        ent_sp_14.insert(END, str(lst_var_sp_pv[13]))
        ent_sp_15 = Entry(frm_var_sp, width=10)
        ent_sp_15.insert(END, str(lst_var_sp_pv[14]))

        ent_sp_16 = Entry(frm_var_sp2, width=10)
        ent_sp_16.insert(END, str(lst_var_sp_pv[15]))
        ent_sp_17 = Entry(frm_var_sp2, width=10)
        ent_sp_17.insert(END, str(lst_var_sp_pv[16]))
        ent_sp_18 = Entry(frm_var_sp2, width=10)
        ent_sp_18.insert(END, str(lst_var_sp_pv[17]))
        ent_sp_19 = Entry(frm_var_sp2, width=10)
        ent_sp_19.insert(END, str(lst_var_sp_pv[18]))
        ent_sp_20 = Entry(frm_var_sp2, width=10)
        ent_sp_20.insert(END, str(lst_var_sp_pv[19]))
        ent_sp_21 = Entry(frm_var_sp2, width=10)
        ent_sp_21.insert(END, str(lst_var_sp_pv[20]))
        ent_sp_22 = Entry(frm_var_sp2, width=10)
        ent_sp_22.insert(END, str(lst_var_sp_pv[21]))
        ent_sp_23 = Entry(frm_var_sp2, width=10)
        ent_sp_23.insert(END, str(lst_var_sp_pv[22]))
        ent_sp_24 = Entry(frm_var_sp2, width=10)
        ent_sp_24.insert(END, str(lst_var_sp_pv[23]))
        ent_sp_25 = Entry(frm_var_sp2, width=10)
        ent_sp_25.insert(END, str(lst_var_sp_pv[24]))

        ent_sp_1.place(x=60, y=10)
        ent_sp_2.place(x=60, y=40)
        ent_sp_3.place(x=60, y=70)
        ent_sp_4.place(x=60, y=100)
        ent_sp_5.place(x=60, y=130)
        ent_sp_6.place(x=60, y=160)
        ent_sp_7.place(x=60, y=190)
        ent_sp_8.place(x=60, y=220)
        ent_sp_9.place(x=60, y=250)
        ent_sp_10.place(x=60, y=280)
        ent_sp_11.place(x=60, y=310)
        ent_sp_12.place(x=60, y=340)
        ent_sp_13.place(x=60, y=370)
        ent_sp_14.place(x=60, y=400)
        ent_sp_15.place(x=60, y=430)

        ent_sp_16.place(x=60, y=10)
        ent_sp_17.place(x=60, y=40)
        ent_sp_18.place(x=60, y=70)
        ent_sp_19.place(x=60, y=100)
        ent_sp_20.place(x=60, y=130)
        ent_sp_21.place(x=60, y=160)
        ent_sp_22.place(x=60, y=190)
        ent_sp_23.place(x=60, y=220)
        ent_sp_24.place(x=60, y=250)
        ent_sp_25.place(x=60, y=280)

        lbl_sp_1_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_2_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_3_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_4_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_5_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_6_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_7_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_8_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_9_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_10_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_11_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_12_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_13_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_14_des = Label(frm_var_sp, text="$/kWh")
        lbl_sp_15_des = Label(frm_var_sp, text="$/kWh")

        lbl_sp_16_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_17_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_18_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_19_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_20_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_21_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_22_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_23_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_24_des = Label(frm_var_sp2, text="$/kWh")
        lbl_sp_25_des = Label(frm_var_sp2, text="$/kWh")



        lbl_sp_1_des.place(x=130, y=10)
        lbl_sp_2_des.place(x=130, y=40)
        lbl_sp_3_des.place(x=130, y=70)
        lbl_sp_4_des.place(x=130, y=100)
        lbl_sp_5_des.place(x=130, y=130)
        lbl_sp_6_des.place(x=130, y=160)
        lbl_sp_7_des.place(x=130, y=190)
        lbl_sp_8_des.place(x=130, y=220)
        lbl_sp_9_des.place(x=130, y=250)
        lbl_sp_10_des.place(x=130, y=280)
        lbl_sp_11_des.place(x=130, y=310)
        lbl_sp_12_des.place(x=130, y=340)
        lbl_sp_13_des.place(x=130, y=370)
        lbl_sp_14_des.place(x=130, y=400)
        lbl_sp_15_des.place(x=130, y=430)

        lbl_sp_16_des.place(x=130, y=10)
        lbl_sp_17_des.place(x=130, y=40)
        lbl_sp_18_des.place(x=130, y=70)
        lbl_sp_19_des.place(x=130, y=100)
        lbl_sp_20_des.place(x=130, y=130)
        lbl_sp_21_des.place(x=130, y=160)
        lbl_sp_22_des.place(x=130, y=190)
        lbl_sp_23_des.place(x=130, y=220)
        lbl_sp_24_des.place(x=130, y=250)
        lbl_sp_25_des.place(x=130, y=280)

        lst_ent_var_sp = [ent_sp_1,ent_sp_2,ent_sp_3,ent_sp_4,ent_sp_5,ent_sp_6,ent_sp_7,ent_sp_8,ent_sp_9,
                          ent_sp_10,ent_sp_11,ent_sp_12,ent_sp_13,ent_sp_14,ent_sp_15,ent_sp_16,ent_sp_17
                          ,ent_sp_18,ent_sp_19,ent_sp_20,ent_sp_21,ent_sp_22,ent_sp_23,ent_sp_24,ent_sp_25]
        btn_save_sp = Button(frm_var_sp, text="Save", command= lambda:save_sp_pv(lst_ent_var_sp) )
        btn_save_sp.place(x=80, y=470)
    except:
        messagebox.showerror("Error","Error while saving selling price")

def change_SI_wind_farm():
    if var_SI_wind_farm.get() == "kWh":
        ent_nom_pow_WF.configure(state="normal")
        ent_nom_pow_WF.delete(0, END)
        nom_pow_WF = nom_pow_wind_farm
        ent_nom_pow_WF.insert(END, str(nom_pow_WF))
        ent_nom_pow_WF.config(state="disabled")

    elif var_SI_wind_farm.get() == "MWh":
        ent_nom_pow_WF.configure(state="normal")
        ent_nom_pow_WF.delete(0, END)
        nom_pow_WF = nom_pow_wind_farm / 1000
        ent_nom_pow_WF.insert(END, str(nom_pow_WF))
        ent_nom_pow_WF.config(state="disabled")

    elif var_SI_wind_farm.get() == "GWh":
        ent_nom_pow_WF.configure(state="normal")
        ent_nom_pow_WF.delete(0, END)
        nom_pow_WF = nom_pow_wind_farm / 1000000
        ent_nom_pow_WF.insert(END, str(nom_pow_WF))
        ent_nom_pow_WF.config(state="disabled")
    return


def plot_5_graph(x_values, y1_values, y2_values, y3_values, y4_values, y5_values, y1_legend, y2_legend, y3_legend,
                 y4_legend, y5_legend, x_label, y_label, title):
    plt.plot(x_values, y1_values, label=y1_legend)
    plt.plot(x_values, y2_values, label=y2_legend)
    plt.plot(x_values, y3_values, label=y3_legend)
    plt.plot(x_values, y4_values, label=y4_legend)
    plt.plot(x_values, y5_values, label=y5_legend)

    plt.title(title)
    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.legend()
    plt.show()

    return


def plot_opt_graph(lst_opt_pv_sizes, lst_opt_wt_sizes, lst_opt_st_sizes, lst_opt_SS, lst_opt_irr):
    y_lbl = "Values"
    x_lbl = "Itirations"

    x_values = range(1, len(lst_opt_pv_sizes) + 1)
    # plt.plot(lst_date_time,lst_pv_prod_ofmonth, label= "PV")
    # plt.plot(lst_date_time, lst_WT_prod_ofmonth, label= "WT")
    plt.plot(x_values, lst_opt_pv_sizes, label="PV sizes")
    plt.plot(x_values, lst_opt_wt_sizes, label="WT sizes")
    plt.plot(x_values, lst_opt_st_sizes, label="ST sizes")
    plt.plot(x_values, lst_opt_SS, label="Self sufficiency")
    plt.plot(x_values, lst_opt_irr, label="IRR")

    plt.title("Optimised results")
    plt.xlabel(x_lbl)
    plt.ylabel(y_lbl)
    plt.legend()
    plt.show()
    return


def cal_all_WT_prod():
    try:

        # progress bar start
        newWindow = Toplevel(root)
        newWindow.title("Progress Bar")
        newWindow.geometry('400x100+400+250')

        newWindow.protocol("WM_DELETE_WINDOW", disable_event)

        # progress_bar = ttk.Progressbar(newWindow, orient=HORIZONTAL, length=400, mode="determinate")
        # progress_bar.pack()
        lbl_progress_bar = Label(newWindow, text="Calculation in progress")
        lbl_progress_bar.pack()

        # progress_bar["value"] = 2
        # lbl_progress_bar.config(text="Downloading Data")
        root.update()
    except:
        messagebox.showerror("Error")
    try:
        if ent_start_year.get() > ent_end_year.get():
            messagebox.showerror("Date error", "Starting Year should be less than end year")
            return
        if len(ent_start_year.get()) != 4 or len(ent_end_year.get()) != 4:
            messagebox.showerror("Date error", "starting year or Ending year is not entered properly")
            return
        if int(ent_start_year.get()) < 2005 or int(ent_end_year.get()) > 2016:
            messagebox.showerror("Date error", "Incorrect value. Please, enter an integer between 2005 and 2016")
            return

        latitude = str(ent_latitude.get())
        longitude = str(ent_longitude.get())

        start_year = str(ent_start_year.get())
        end_year = str(ent_end_year.get())
        total_years = int(end_year) - int(start_year)

        url_first_part = "https://re.jrc.ec.europa.eu/api/seriescalc?"
        final_url = url_first_part + "lat=" + latitude + "&lon=" + longitude + "&startyear=" + start_year \
                    + "&endyear=" + end_year + "&optimalangles=1&outputformat=json&browser=1"
        res = requests.get(final_url)
        res_json_file = io.StringIO(res.text)
        src = json.load(res_json_file)
        output = src['outputs']
        output_hourly = output["hourly"]
        data = pd.DataFrame(output_hourly)
        if var_India.get() == 1:
            data = change_data_for_asia(data)
        data.index = pd.to_datetime(data["time"], format='%Y%m%d:%H%M', utc=True)
        data = data.drop("Int", axis=1)
        ################################################################################################################

        try:
            lst_wind_speed_hubHeight = []
            lst_wind_turbine_prod = []
            selected_turbine = var_turbine.get()
            selected_turbine_para = para_WT_df[selected_turbine]
            height_of_rotor = selected_turbine_para[0]
            terrain_rough = float(ent_terrain_roug.get())
            sel_WT_nom_pow = selected_turbine_para[1]
            col_wind_speed = data["WS10m"]
            nb_WT = int(ent_nb_turbines.get())
            ref_height = float(ent_mes_height.get())
            numerator = math.log(height_of_rotor / terrain_rough)
            denominator = math.log(ref_height / terrain_rough)

            int_col_wind_speed2 = power_curve_df.columns.get_loc(selected_turbine) - 1
            lst_power_PC = power_curve_df[selected_turbine].tolist()
            lst_wind_speed_PC = power_curve_df.iloc[:, int_col_wind_speed2].tolist()
            all_WT_df = pd.DataFrame()

            lst_time_data = []
            lst_wind_speed = []
            # for o in range (0, len(col_time)):
            #     lst_time_data.append(col_time[o])
            #     lst_wind_speed.append(col_wind_speed[o])
            # lst_time_data = data["time"].tolist
            # all_WT_df.insert(0,"Time",lst_time_data)
            all_WT_df = data[["time"]]

            # progress_bar["value"] = 10
            # lbl_progress_bar.config(text="Calculating all WT ptoduction")
            # root.update()

            int_WT = 1
            for j in range(0, int(len(power_curve_df.columns) / 2)):
                selected_turbine_name_int_WT = para_WT_df.columns[j]
                selected_turbine_para_int_WT = para_WT_df[selected_turbine_name_int_WT]
                height_of_rotor_int_WT = selected_turbine_para_int_WT[0]
                numerator_int_WT = math.log(height_of_rotor_int_WT / terrain_rough)
                denominator_int_WT = math.log(ref_height / terrain_rough)

                int_col_wind_speed_int_WT = power_curve_df.columns.get_loc(selected_turbine_name_int_WT) - 1
                lst_power_PC_int_WT = power_curve_df[selected_turbine_name_int_WT].tolist()
                lst_wind_speed_PC_int_WT = power_curve_df.iloc[:, int_col_wind_speed_int_WT].tolist()

                # progress_bar["value"] += 10
                # lbl_progress_bar.config(text="Calculating production of "+selected_turbine_name_int_WT)
                # root.update()

                lst_wind_turbine_prod_int_WT = []
                for t in range(0, len(col_wind_speed)):
                    corrected_wind_speed_int_WT = col_wind_speed[t] * (numerator_int_WT / denominator_int_WT)
                    wind_pow_interp_int_WT = scipy.interpolate.interp1d(lst_wind_speed_PC_int_WT, lst_power_PC_int_WT)
                    wind_turbine_prod_int_WT = wind_pow_interp_int_WT(corrected_wind_speed_int_WT)
                    wind_turbine_prod_int_WT = wind_turbine_prod_int_WT * nb_WT
                    lst_wind_turbine_prod_int_WT.append(wind_turbine_prod_int_WT)

                all_WT_df.insert(j + 1, para_WT_df.columns[j], lst_wind_turbine_prod_int_WT)
                int_WT += 2

            all_WT_df = all_WT_df.resample('Y').sum()
            # if var_SI_unit_nom_pow.get() == 1:
            #     all_WT_df = all_WT_df
            #
            # elif var_SI_unit_nom_pow.get() == 2:
            #     all_WT_df = all_WT_df.div(1000)
            #
            # elif var_SI_unit_nom_pow.get() == 3:
            #     all_WT_df = all_WT_df.div(1000000)
            #
            # else:
            #     messagebox.showerror("Error", "Please select the SI unit of nominal power")
            #     return
            all_WT_df = all_WT_df.round()

            WT_id = 0
            for each_col in all_WT_df.columns:
                selected_turbine_name_int_WT2 = para_WT_df.columns[WT_id]
                selected_turbine_nom_pow = para_WT_df[selected_turbine_name_int_WT2][1]
                all_WT_df[selected_turbine_name_int_WT2] = all_WT_df[selected_turbine_name_int_WT2].div(
                    selected_turbine_nom_pow)
                WT_id += 1

            all_WT_df = all_WT_df.div(nb_WT)
            float_WT_ageing = float(ent_ageing_WT.get()) / 100
            lst_25_yrs = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25]
            all_WT_25yrs_df = replicate_data_for_25yrs(total_years + 1, all_WT_df)
            production_with_ageing(all_WT_25yrs_df, float_WT_ageing)
            all_WT_25yrs_df.insert(0, "Years", lst_25_yrs)
            show_data_new_win_XY(all_WT_25yrs_df, tab_WT_results, 600,
                                 1200, 10, 30)  # inserting Wind turbine data in Wind turbine tab
            newWindow.destroy()
        except:
            messagebox.showerror("Error",
                                 "Python found error while calculating the Wind turbine production(Inside Try)")
    except:
        messagebox.showerror("Error", "Error while calculating WT Production")


def plot_heatmap(df, m):
    try:
        time.sleep(2)
        plt.pcolor(df)
        time.sleep(5)
        months = ["Jan", "Feb", "March", "April", "May", "June", "July", "August", "Sep", "Oct", "Nov", "Dec"]
        # plt.yticks(np.arange(0.5, len(df.index), 1), df.index)
        plt.yticks(np.arange(0.5, len(df.index), 1), months)
        # time.sleep(3)
        plt.xticks(np.arange(0.5, len(df.columns), 1), df.columns)
        # time.sleep(3)

        if m == 0:
            plt.xlabel("Total Yearly Global irradiance (kW/m2)")
        elif m == 1:
            plt.xlabel("Avg air temperature (degree Celsius)")
        elif m == 2:
            plt.xlabel("Avg wind speed (m/s) Measured at height 10m")

        for y in range(df.shape[0]):
            for x in range(df.shape[1]):
                plt.text(x + 0.5, y + 0.5, '%.2f' % df.iloc[y][x],
                         horizontalalignment='center',
                         verticalalignment='center',
                         )

        plt.show()
    except:
        messagebox.showerror("Error", "Python found a error while plotting heatmap ")
    return


def display_energy_flow(date, data):
    try:
        show_monthly_energy_flows(date, data)
        show_daily_energyflow_and_bal(date, data)
    except:
        messagebox.showerror("Error", "Error while calculating both energy flows")


def display_wind_heatmap(data, totalyears, startyear):
    list_of_yearly_wind_data = []
    c = 0

    for i in range(0, totalyears + 1):
        list_of_month_wind_data = []
        yearly_data = data.iloc[c: c + 12, 3:4]
        monthly_data = yearly_data["WS10m"]
        for eachitem in monthly_data:
            list_of_month_wind_data.append(eachitem)
        list_of_yearly_wind_data.append(list_of_month_wind_data)
        c += 12

    wind_df = pd.DataFrame()
    d = 0
    for each_yearly_datas in list_of_yearly_wind_data:
        wind_df.insert(d, str(startyear + d), each_yearly_datas)
        d += 1
    wind_df.round(2)

    time.sleep(1)
    plot_heatmap(wind_df, 2)
    return


def hide_lst_entry(lst_entries):
    try:
        for i in range(0, len(lst_entries)):
            lst_entries[i].configure(state="disabled")
    except:
        messagebox.showerror("Error", "Error while hiding list of entries")


def entry_set_and_disable(entry, text_var, value):
    entry.configure(state="normal")
    text_var.set(value)
    entry.configure(state="disabled")


def show_hide_peak_shaving():
    if var_peak_shaving.get() == "Limitation on Maximum generated power":
        ent_max_gen.configure(state="normal")
        ent_max_gen.delete(0, END)
        ent_max_inj.delete(0, END)
        ent_max_inj.config(state="disabled")
        return
    elif var_peak_shaving.get() == "Limitation on Maximum power injection":
        ent_max_inj.configure(state="normal")
        ent_max_inj.delete(0, END)
        ent_max_gen.delete(0, END)
        ent_max_gen.config(state="disabled")
        return
    elif var_peak_shaving.get() == "No limitations":
        ent_max_gen.delete(0, END)
        ent_max_inj.delete(0, END)
        ent_max_gen.config(state="disabled")
        ent_max_inj.config(state="disabled")
        return
    else:
        messagebox.showerror("Error", "Error in peak shaving typology selection")


def calculate_annual_load():
    try:
        load_wb = xl.load_workbook(ent_load_type.get())
        sheet_load = load_wb["Sheet1"]

        max_row = sheet_load.max_row
        lst_load_frm_excel = []
        for i in range(1, max_row):
            lst_load_frm_excel.append(sheet_load.cell(i + 1, 2).value)

        annual_load = sum(lst_load_frm_excel)
        if annual_load == None:
            return

        return annual_load
    except:
        messagebox.showerror("ERROR", "Error while calculating annual load")


def open_file():
    root.filename = filedialog.askopenfilename(initialdir="\filedialog",
                                               title="Select a excel file",
                                               filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))

    file_name = root.filename
    var_load_entry.set(file_name)

    ###############################################################################################################
    annual_load = calculate_annual_load()
    lbl_annual_load = Label(frm_load, text="Annual load = " + str(annual_load / 1000) + " (MWh)")
    lbl_annual_load.place(x=10, y=111)

    lbl_avg_daily_load = Label(frm_load, text="Avg daily load = " + str(round(annual_load / (1000 * 365))) + " (MWh)")
    lbl_avg_daily_load.place(x=10, y=135)

    global avg_daily_load
    avg_daily_load = round(annual_load / (1000 * 365))
    ###############################################################################################################


def show_nom_pow_Wind_farm():
    ent_nom_pow_WF.configure(state="normal")
    ent_nom_pow_WF.delete(0, END)

    selected_WT = var_turbine.get()
    selected_WT_nom_pow = para_WT_df[selected_WT][1]
    nb_WT = ent_nb_turbines.get()
    nom_pow_wf = int(nb_WT) * int(selected_WT_nom_pow)

    ent_nom_pow_WF.insert(END, str(nom_pow_wf))
    ent_nom_pow_WF.config(state="disabled")


def show_nominal_power():
    ent_WT_nom_pow.configure(state="normal")
    ent_WT_nom_pow.delete(0, END)
    ent_nom_pow_WF.configure(state="normal")
    ent_nom_pow_WF.delete(0, END)

    selected_WT = var_turbine.get()
    selected_WT_nom_pow = para_WT_df[selected_WT][1]

    nb_WT = ent_nb_turbines.get()
    nom_pow_wf = int(nb_WT) * int(selected_WT_nom_pow)

    ent_nom_pow_WF.insert(END, str(nom_pow_wf))

    ent_WT_nom_pow.insert(END, str(selected_WT_nom_pow))
    ent_WT_nom_pow.config(state="disabled")
    ent_nom_pow_WF.config(state="disabled")

    global nom_pow_wind_farm
    nom_pow_wind_farm = nom_pow_wf
    var_SI_wind_farm.set(lst_SI_units_wind_farm[0])
    return


def display_temp_heatmap(data, totalyears, startyear):
    list_of_yearly_temp_data = []
    t = 0
    for i in range(0, totalyears + 1):
        list_of_month_temp_data = []
        yearly_data = data.iloc[t: t + 12, 2:3]
        monthly_data = yearly_data["T2m"]
        for eachitem in monthly_data:
            list_of_month_temp_data.append(eachitem)
        list_of_yearly_temp_data.append(list_of_month_temp_data)
        t += 12
    temp_df = pd.DataFrame()
    y = 0
    for each_yearly_datas in list_of_yearly_temp_data:
        temp_df.insert(y, str(startyear + y), each_yearly_datas)
        y += 1

    time.sleep(1)
    plot_heatmap(temp_df, 1)
    return


def display_irradiation_heatmap(data, totalyears, startyear):
    list_of_yearly_irr_data = []
    list_of_month_irr_data = []
    j = 0

    for i in range(0, totalyears + 1):
        list_of_month_irr_data = []
        yearly_data = data.iloc[j: j + 12, 0:1]
        monthly_data = yearly_data["G(i)"]
        # yearly_data = yearly_data.rename({'G(i)': str(startyear+k)}, axis=1)
        for eachitem in monthly_data:
            list_of_month_irr_data.append(eachitem)

        list_of_yearly_irr_data.append(list_of_month_irr_data)
        j += 12
    irr_df = pd.DataFrame()
    k = 0
    for each_yearly_datas in list_of_yearly_irr_data:
        irr_df.insert(k, str(startyear + k), each_yearly_datas)
        k += 1

    time.sleep(1)
    irr_df = irr_df.div(1000)
    plot_heatmap(irr_df, 0)
    return


def plot_bar_and_line_graph(df):
    try:
        plt.rcParams["figure.figsize"] = [7.50, 3.50]
        plt.rcParams["figure.autolayout"] = True

        fig, ax = plt.subplots()
        lst_cash_flow = df['Total Cash Flows($/y)'].to_list()
        lst_NPV = df['NPV'].to_list()

        lst_cash_flow = [x / 1000000 for x in lst_cash_flow]
        lst_NPV = [x / 1000000 for x in lst_NPV]

        plot_df = pd.DataFrame()
        plot_df.insert(0,"1",lst_cash_flow)
        plot_df.insert(1, "2",lst_NPV)

        # df['Total Cash Flows($/y)'] = df['Total Cash Flows($/y)'].div(1000000)
        # df['NPV'] = df['NPV'].div(1000000)
        plot_df['1'].plot(kind='bar', color='red')
        plot_df['2'].plot(kind='line', marker='.', color='black', ms=10)
        # lst_cash_flow.plot(kind='bar', color='red')
        # lst_NPV.plot(kind='line', marker='.', color='black', ms=10)

        plt.xlabel("Years")
        plt.ylabel("Cash Flows (M$/y)")
        plt.title("NPV PV+WT")
        plt.show()
    except:
        messagebox.showerror("Error", "Error while plotting NPV graph")


def plot_graph(x_values, y_values, x_label, y_label, strTitle):
    try:

        mpl.rcParams['lines.linewidth'] = 2
        plt.plot(x_values, y_values)
        plt.xlabel(x_label)
        plt.ylabel(y_label)
        plt.title(strTitle)

        plt.show()

    except:
        messagebox.showerror("Error", "Python found a error while plotting graph ")
    return


def plot_2y_graph(x_values, y1_values, y2_values, x_label, y1_label, y2_label):
    try:

        fig, ax1 = plt.subplots(1, 2, sharex=True, sharey=True)

        ax1.plot(x_values, y1_values, 'g-')

        ax1.set_xlabel(x_label)
        ax1.set_ylabel(y1_label, color='g')
        ax2 = ax1.twinx()
        ax2.plot(x_values, y2_values, 'b-')
        ax2.set_ylabel(y2_label, color='b')



    except:
        messagebox.showerror("Error", "Python found a error while plotting 2Y-axis graph ")
    return


def show_data_new_win_XY(df, tabFrame, intHeight, intWidth, x, y):
    try:
        df = df.round(2)

        frame_data = LabelFrame(tabFrame, text="Data")
        frame_data.place(height=intHeight, width=intWidth, x=x, y=y)

        tv1 = ttk.Treeview(frame_data)
        tv1.place(relheight=1, relwidth=1)

        scroll_tv1_x = Scrollbar(frame_data, orient="horizontal", command=tv1.xview)
        scroll_tv1_y = Scrollbar(frame_data, orient="vertical", command=tv1.yview)
        tv1.configure(xscrollcommand=scroll_tv1_x.set, yscrollcommand=scroll_tv1_y.set)
        scroll_tv1_x.pack(side="bottom", fill="x")
        scroll_tv1_y.pack(side="right", fill="y")

        tv1["column"] = list(df.columns)
        tv1["show"] = "headings"
        for column in tv1["column"]:
            tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            tv1.insert("", "end", values=row)
    except:
        messagebox.showerror("Error", "Python found error while inserting battery data in Battery storage tab")
    return


def show_data_new_win(df, tabFrame, intHeight, intWidth):
    try:
        df = df.round(2)

        frame_data = LabelFrame(tabFrame, text="Data")
        frame_data.place(height=intHeight, width=intWidth, rely=0, relx=0)

        tv1 = ttk.Treeview(frame_data)
        tv1.place(relheight=1, relwidth=1)

        scroll_tv1_x = Scrollbar(frame_data, orient="horizontal", command=tv1.xview)
        scroll_tv1_y = Scrollbar(frame_data, orient="vertical", command=tv1.yview)
        tv1.configure(xscrollcommand=scroll_tv1_x.set, yscrollcommand=scroll_tv1_y.set)
        scroll_tv1_x.pack(side="bottom", fill="x")
        scroll_tv1_y.pack(side="right", fill="y")

        tv1["column"] = list(df.columns)
        tv1["show"] = "headings"
        for column in tv1["column"]:
            tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            tv1.insert("", "end", values=row)
    except:
        messagebox.showerror("Error", "Python found error while inserting battery data in Battery storage tab")
    return


def show_data(df):
    df = df.round(2)
    win_data = Toplevel()
    win_data.title("Data")
    win_data.geometry("800x500")
    frame_data = LabelFrame(win_data, text="Data")
    frame_data.place(height=500, width=800, rely=0, relx=0)

    tv1 = ttk.Treeview(frame_data)
    tv1.place(relheight=1, relwidth=1)

    scroll_tv1_x = Scrollbar(frame_data, orient="horizontal", command=tv1.xview)
    scroll_tv1_y = Scrollbar(frame_data, orient="vertical", command=tv1.yview)
    tv1.configure(xscrollcommand=scroll_tv1_x.set, yscrollcommand=scroll_tv1_y.set)
    scroll_tv1_x.pack(side="bottom", fill="x")
    scroll_tv1_y.pack(side="right", fill="y")

    tv1["column"] = list(df.columns)
    tv1["show"] = "headings"
    for column in tv1["column"]:
        tv1.heading(column, text=column)

    df_rows = df.to_numpy().tolist()
    for row in df_rows:
        tv1.insert("", "end", values=row)

    return


# def display_heat_map(data , totalyears , startyear, colName):
#     list_of_yearly_data = []
#     i = 0
#     for i in range(0 , totalyears+1):
#         list_of_month_data =[]
#         if colName == "G(i)":
#             yearly_data = data.iloc[i: i + 12, 0:1]
#         elif colName == "T2m":
#             yearly_data = data.iloc[i: i + 12, 2:3]
#         elif colName == "WS10m":
#             yearly_data = data.iloc[i: i + 12, 3:4]
#
#         monthly_data = yearly_data[colName]
#
#         for eachitem in monthly_data:
#             list_of_month_data.append(eachitem)
#
#         list_of_yearly_data.append(list_of_month_data)
#         i += 12
#     df = pd.DataFrame()
#     k=0
#     for each_yearly_datas in list_of_yearly_data:
#         df.insert(k,str(startyear+k),each_yearly_datas)
#         k +=1
#
#     time.sleep(1)
#     if colName=="G(i)":
#         plot_heatmap(df,0)
#     elif colName == "T2m":
#         plot_heatmap(df, 1)
#     elif colName == "WS10m":
#         plot_heatmap(df,2)
#     return


def show_daily_trend_ofmonth(strDate, df):
    try:
        sel_date = strDate.strftime("%d-%m-%Y")
        sel_year = strDate.year
        sel_month = strDate.month
        total_nb_day = monthrange(sel_year, sel_month)[1]

        start_day = str(sel_month) + "-01" + "-" + str(sel_year)

        end_day = str(sel_month) + "-" + str(total_nb_day) + "-" + str(sel_year)
        # load = float(ent_load.get())
        if sel_year < int(ent_start_year.get()) or sel_year > int(ent_end_year.get()):
            messagebox.showerror("Error", "Incorrect value. Please, enter an date between " + str(
                ent_start_year.get()) + " and " + str(ent_end_year.get()))
            return
        sel_month_df = df.loc[start_day:end_day]
        sel_month_AC_col = sel_month_df["Power in ALternate Current(kW)"]
        sel_month_WT_col = sel_month_df["Wind Turbine Active Power(kW)"]
        sel_month_overall_prod_col = sel_month_df["Overall production with Peak Shaving(kW)"]
        sel_month_time_col = sel_month_df["time"]
        sel_month_load_col = sel_month_df['Load(kW)']
        lst_pv_prod_ofmonth = []
        lst_WT_prod_ofmonth = []
        lst_overall_prod_after_PS = []
        lst_load = []
        lst_date_time = []
        for i in range(0, len(sel_month_AC_col)):
            lst_pv_prod_ofmonth.append(sel_month_AC_col[i])
            lst_WT_prod_ofmonth.append(sel_month_WT_col[i])
            day = datetime.strptime(sel_month_time_col[i], '%Y%m%d:%H%M')
            lst_date_time.append(day)
            lst_load.append(sel_month_load_col[i])
            lst_overall_prod_after_PS.append(sel_month_overall_prod_col[i])

        y_lbl = "PV Production (kW)"
        x_lbl = "Days"

        plt.plot(lst_date_time, lst_pv_prod_ofmonth, label="PV")
        plt.plot(lst_date_time, lst_WT_prod_ofmonth, label="WT")
        plt.plot(lst_date_time, lst_load, label="Load")
        plt.plot(lst_date_time, lst_overall_prod_after_PS, label="PV+WT")
        plt.xticks(rotation=30, ha='right')
        plt.title("Production for month " + strDate.strftime("%B"))
        plt.xlabel(x_lbl)
        plt.ylabel(y_lbl)
        plt.legend()
        plt.show()
        return

    except:
        messagebox.showerror("Error", "Error while plotting daily trend of the month")


def show_grid_exchange_ofmonth(strDate, df):
    try:
        sel_date = strDate.strftime("%d-%m-%Y")
        sel_year = strDate.year
        sel_month = strDate.month
        total_nb_day = monthrange(sel_year, sel_month)[1]

        start_day = str(sel_month) + "-01" + "-" + str(sel_year)

        end_day = str(sel_month) + "-" + str(total_nb_day) + "-" + str(sel_year)
        # load = float(ent_load.get())
        if sel_year < int(ent_start_year.get()) or sel_year > int(ent_end_year.get()):
            messagebox.showerror("Error", "Incorrect value. Please, enter an date between " + str(
                ent_start_year.get()) + " and " + str(ent_end_year.get()))
            return
        sel_month_df = df.loc[start_day:end_day]
        sel_month_AC_col = sel_month_df["Power in ALternate Current(kW)"]
        sel_month_WT_col = sel_month_df["Wind Turbine Active Power(kW)"]
        sel_month_overall_prod_col = sel_month_df["Overall production with Peak Shaving(kW)"]
        sel_month_time_col = sel_month_df["time"]
        sel_month_inj_grid_col = sel_month_df["Injection in grid with Peak shaving(kW)"]
        sel_month_abs_grid_col = sel_month_df["Absorption from grid(kW)"]
        sel_month_self_suff_col = sel_month_df["Self Suffiency"]
        sel_month_batt_discharge_col = sel_month_df["Battery Discharge(kW)"]
        sel_month_batt_charge_col = sel_month_df["Battery Charge(kW)"]
        sel_month_load_col = sel_month_df["Load(kW)"]

        lst_pv_prod_ofmonth = []
        lst_WT_prod_ofmonth = []
        lst_overall_prod_after_PS = []
        lst_load = []
        lst_date_time = []
        lst_month_inj = []
        lst_month_abs = []
        lst_self_suff_month = []
        lst_batt_dis_month = []
        lst_batt_charge_month = []

        for i in range(0, len(sel_month_AC_col)):
            lst_pv_prod_ofmonth.append(sel_month_AC_col[i])
            lst_WT_prod_ofmonth.append(sel_month_WT_col[i])
            day = datetime.strptime(sel_month_time_col[i], '%Y%m%d:%H%M')
            lst_date_time.append(day)
            lst_load.append(sel_month_load_col[i])
            lst_overall_prod_after_PS.append(sel_month_overall_prod_col[i])
            lst_month_inj.append(sel_month_inj_grid_col[i])
            lst_month_abs.append(sel_month_abs_grid_col[i])
            lst_self_suff_month.append(sel_month_self_suff_col[i])
            lst_batt_dis_month.append(sel_month_batt_discharge_col[i])
            lst_batt_charge_month.append(sel_month_batt_charge_col[i])

        y_lbl = "PV Production (kW)"
        x_lbl = "Days"

        # plt.plot(lst_date_time,lst_pv_prod_ofmonth, label= "PV")
        # plt.plot(lst_date_time, lst_WT_prod_ofmonth, label= "WT")
        plt.plot(lst_date_time, lst_month_inj, label="Injection into Grid")
        plt.plot(lst_date_time, lst_month_abs, label="Absorption from Grid")
        plt.plot(lst_date_time, lst_load, label="Load")
        plt.plot(lst_date_time, lst_overall_prod_after_PS, label="PV+WT")
        plt.plot(lst_date_time, lst_self_suff_month, label="Self Sufficiency")
        plt.plot(lst_date_time, lst_batt_dis_month, label="Battery Discharge")
        plt.plot(lst_date_time, lst_batt_charge_month, label="Battery Charge")
        plt.xticks(rotation=30, ha='right')
        plt.title("Grid Exchange for month " + strDate.strftime("%B"))
        plt.xlabel(x_lbl)
        plt.ylabel(y_lbl)
        plt.legend()
        plt.show()
        return

    except:
        messagebox.showerror("Error", "Error while plotting daily trend of the month")


def replicate_data_for_25yrs_with_age(total_years, df, ageing):
    try:
        remaining_years = 25 - total_years
        quotient = remaining_years // total_years
        i = 0
        new_df = df
        if quotient >= 1:
            while i < quotient:
                new_df = new_df.mul(1 - ageing * i)
                new_df = new_df.append(df)

                i += 1

        rem_yrs = 25 - total_years - (quotient * total_years)
        rem_yrs_df = df.iloc[0:rem_yrs]
        # new_df=new_df.append(rem_yrs_df)
        replicated_df = pd.concat([new_df, rem_yrs_df], axis=0)
        return replicated_df
    except:
        messagebox.showerror("Error", "Error while replicating data for 25 years")
    return


def production_with_ageing(df, ageing):
    try:
        for eachCol in df.columns:
            for i in range(0, 25):
                df[eachCol][i] = df[eachCol][i] * (1 - ageing * i)

        return df
    except:
        messagebox.showerror("Error", "Error while calculating Wind Turbine Production with ageing")
    return


def replicate_data_for_25yrs(total_years, df):
    try:
        remaining_years = 25 - total_years
        quotient = remaining_years // total_years
        i = 0
        new_df = df
        if quotient >= 1:
            while i < quotient:
                new_df = new_df.append(df)
                i += 1

        rem_yrs = 25 - total_years - (quotient * total_years)
        rem_yrs_df = df.iloc[0:rem_yrs]
        # new_df=new_df.append(rem_yrs_df)
        replicated_df = pd.concat([new_df, rem_yrs_df], axis=0)
        return replicated_df
    except:
        messagebox.showerror("Error", "Error while replicating data for 25 years")
    return


def show_data_in_win_rel(df, x, y, tab, frameHeight, frameWidth, frame_heading):
    try:

        frame_data = LabelFrame(tab, text=frame_heading)
        frame_data.place(height=frameHeight, width=frameWidth, rely=0, relx=0, y=y, x=x)

        tv1 = ttk.Treeview(frame_data)
        tv1.place(relheight=1, relwidth=1)

        scroll_tv1_x = Scrollbar(frame_data, orient="horizontal", command=tv1.xview)
        scroll_tv1_y = Scrollbar(frame_data, orient="vertical", command=tv1.yview)
        tv1.configure(xscrollcommand=scroll_tv1_x.set, yscrollcommand=scroll_tv1_y.set)
        scroll_tv1_x.pack(side="bottom", fill="x")
        scroll_tv1_y.pack(side="right", fill="y")

        tv1["column"] = list(df.columns)
        tv1["show"] = "headings"

        i = 0
        for column in tv1["column"]:
            if i == 0:
                tv1.column(column, width=125)
            else:
                tv1.column(column, width=30)
            tv1.heading(column, text=column)
            i += 1

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            tv1.insert("", "end", values=row)
    except:
        messagebox.showerror("Error", "Error while building treeview")


def show_daily_energyflow_and_bal(strDate, df):
    try:
        sel_date = strDate.strftime("%Y-%B-%d")
        sel_year = strDate.year
        sel_month = strDate.month
        sel_day = strDate.day
        start_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        load = float(ent_load.get())

        end_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        if sel_year < int(ent_start_year.get()) or sel_year > int(ent_end_year.get()):
            messagebox.showerror("Error", "Incorrect value. Please, enter an date between " + str(
                ent_start_year.get()) + " and " + str(ent_end_year.get()))
            return
        sel_date_df = df.loc[start_day: end_day]
        sel_date_df = sel_date_df.resample("D").sum()

        col_PV_prod = sel_date_df["Power in ALternate Current(kW)"]
        col_WT_prod = sel_date_df["Wind Turbine Active Power(kW)"]
        col_overall_prod_after_Both_PS = sel_date_df[
            "PV+WT production after peak shaving (any kind of peak shaving)"]
        col_load = sel_date_df["Load(kW)"]
        col_abs_from_grid = sel_date_df["Absorption from grid(kW)"]
        col_inj_into_grid = sel_date_df["Injection in grid with Peak shaving(kW)"]
        col_grid_exchange = sel_date_df["Grid Exchange(kW)"]
        col_self_sufficiency = sel_date_df["Self Suffiency"]
        col_batt_discharge = sel_date_df["Battery Discharge(kW)"]
        col_batt_charge = sel_date_df["Battery Charge(kW)"]

        lst_labels = ["PV Production", "WT Production", "PV+WT(With peak shaving)", "LOAD", "Battery Discharge",
                      "Battery Charge",
                      "Absoprtion from grid", "Injection into grid", "Grid Exchange", "Self sufficiency"]
        lst_values = [col_PV_prod[0], col_WT_prod[0], col_overall_prod_after_Both_PS[0], col_load[0],
                      col_batt_discharge[0], col_batt_charge[0],
                      col_abs_from_grid[0], col_inj_into_grid[0], col_grid_exchange[0], col_self_sufficiency[0]]
        lst_values = [round(num) for num in lst_values]
        lst_SI_units = ["kWh/day", "kWh/day", "kWh/day", "kWh/day", "kWh/day", 'kWh/day', "kWh/day", "kWh/day",
                        "kWh/day", "kWh/day"]

        daily_energy_flow_df = pd.DataFrame()
        daily_energy_flow_df.insert(0, "Energy Flows(AC)", lst_labels)
        daily_energy_flow_df.insert(1, "kWh/day", lst_values)
        # daily_energy_flow_df.insert(2, "", lst_SI_units)

        show_data_in_win_rel(daily_energy_flow_df, 625, 200, tab_PV_results, 300, 280, "Energy flow on " + sel_date)

        try:
            abs_from_grid_perc = str(round(col_abs_from_grid[0] * 100 / col_load[0])) + "%"
        except:
            abs_from_grid_perc = "--"
        try:
            inj_into_grid_perc = str(round(col_inj_into_grid[0] * 100 / col_load[0])) + "%"
        except:
            inj_into_grid_perc = "--"
        try:
            prod_from_renewables_perc = str(round(col_overall_prod_after_Both_PS[0] * 100 / col_load[0])) + "%"
        except:
            prod_from_renewables_perc = "--"
        try:
            self_suff_wrt_load = str(round(col_self_sufficiency[0] * 100 / col_load[0])) + "%"
        except:
            self_suff_wrt_load = "--"
        try:
            abs_from_grid_wrt_prod = str(round(col_abs_from_grid[0] * 100 / col_overall_prod_after_Both_PS[0])) + "%"
        except:
            abs_from_grid_wrt_prod = "--"
        try:
            inj_into_grid_wrt_prod = str(round(col_inj_into_grid[0] * 100 / col_overall_prod_after_Both_PS[0])) + "%"
        except:
            inj_into_grid_wrt_prod = "--"
        try:
            self_suff_wrt_prod = str(round(col_self_sufficiency[0] * 100 / col_overall_prod_after_Both_PS[0])) + "%"
        except:
            self_suff_wrt_prod = "--"

        lst_labels_2 = ["Absorption from the grid/load", "Injection in the grid/load",
                        "Production from renewables / load ",
                        "Self-sufficiency with respsect to load", "Absorption from the grid/generation",
                        "Injection in the grid/generation",
                        "Self-sufficiency with respect to generation"]
        lst_values_2 = [abs_from_grid_perc, inj_into_grid_perc, prod_from_renewables_perc, self_suff_wrt_load,
                        abs_from_grid_wrt_prod,
                        inj_into_grid_wrt_prod, self_suff_wrt_prod]

        daily_energy_bal_df = pd.DataFrame()
        daily_energy_bal_df.insert(0, "Energy Balance", lst_labels_2)
        daily_energy_bal_df.insert(1, " ", lst_values_2)

        show_data_in_win_rel(daily_energy_bal_df, 935, 200, tab_PV_results, 300, 280, "Energy Balance on " + sel_date)
    except:
        messagebox.showerror("Error", "Error while calculating daily energy flows")

    # try


def show_monthly_energy_flows(strDate, df):
    try:
        sel_date = strDate.strftime("%d-%m-%Y")
        sel_year = strDate.year
        sel_month = strDate.month
        total_nb_day = monthrange(sel_year, sel_month)[1]

        start_day = "01" + "-01" + "-" + str(sel_year)

        end_day = "12" + "-31" + "-" + str(sel_year)
        sel_year_df = df.loc[start_day:end_day]
        sel_year_df = sel_year_df.resample("M").sum()

        col_PV_prod = sel_year_df["Power in ALternate Current(kW)"].div(1000)
        col_WT_prod = sel_year_df["Wind Turbine Active Power(kW)"].div(1000)
        col_overall_prod_after_Both_PS = sel_year_df[
            "PV+WT production after peak shaving (any kind of peak shaving)"].div(1000)
        col_load = sel_year_df["Load(kW)"].div(1000)
        col_abs_from_grid = sel_year_df["Absorption from grid(kW)"].div(1000)
        col_inj_into_grid = sel_year_df["Injection in grid with Peak shaving(kW)"].div(1000)
        col_grid_exchange = sel_year_df["Grid Exchange(kW)"].div(1000)
        col_self_sufficiency = sel_year_df["Self Suffiency"].div(1000)
        lst_months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        energy_flow_df = pd.DataFrame()

        energy_flow_df.insert(0, "", lst_months)
        energy_flow_df.insert(1, "PV", col_PV_prod.to_list())
        energy_flow_df.insert(2, "WT", col_WT_prod.to_list())
        energy_flow_df.insert(3, "PV+WT", col_overall_prod_after_Both_PS.to_list())
        energy_flow_df.insert(4, "Load", col_load.to_list())
        energy_flow_df.insert(5, "Absorption from grid", col_abs_from_grid.to_list())
        energy_flow_df.insert(6, "Injection into grid", col_inj_into_grid.to_list())
        energy_flow_df.insert(7, "Grid Exchange", col_grid_exchange.to_list())
        energy_flow_df.insert(8, "Self Sufficiency", col_self_sufficiency.to_list())

        # energy_flow_df=energy_flow_df.div(1000)
        # show_data(energy_flow_df)
        ###########################################################################################################
        df = energy_flow_df.round(2)

        frame_data = LabelFrame(tab_PV_results,
                                text="Energy flow of each month of year " + str(sel_year) + "(MWh/month)")
        frame_data.place(height=300, width=600, rely=0, relx=0, y=200)

        tv1 = ttk.Treeview(frame_data)
        tv1.place(relheight=1, relwidth=1)

        scroll_tv1_x = Scrollbar(frame_data, orient="horizontal", command=tv1.xview)
        scroll_tv1_y = Scrollbar(frame_data, orient="vertical", command=tv1.yview)
        tv1.configure(xscrollcommand=scroll_tv1_x.set, yscrollcommand=scroll_tv1_y.set)
        scroll_tv1_x.pack(side="bottom", fill="x")
        scroll_tv1_y.pack(side="right", fill="y")

        tv1["column"] = list(df.columns)
        tv1["show"] = "headings"

        for column in tv1["column"]:
            tv1.column(column, width=75)
            tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            tv1.insert("", "end", values=row)
        ###########################################################################################################
        # frame_monthly_energy_data = LabelFrame(tab_PV_results, text="Monthly Energy Flow", padx=10, pady=10)
        # frame_monthly_energy_data.grid(row=9, column=0, columnspan=9, rowspan=14)
        # df = energy_flow_df.round(2)
        #
        # frame_data = LabelFrame(tab_PV_results, text="Energy flow")
        # frame_data.grid(row=9, column=0, columnspan=1, rowspan=14)
        # frame_data.configure(height=200, width=400)
        # tv1 = ttk.Treeview(frame_data)
        # tv1.grid(row=0, column=0)
        # # tv1.configure()
        #
        # scroll_tv1_x = Scrollbar(frame_data, orient="horizontal", command=tv1.xview)
        # scroll_tv1_y = Scrollbar(frame_data, orient="vertical", command=tv1.yview)
        # scroll_tv1_x.grid(row=0, column=0, sticky=tk.NS)
        # scroll_tv1_y.grid(row=0, column=0, sticky=tk.S)
        # tv1.configure(xscrollcommand=scroll_tv1_x.set, yscrollcommand=scroll_tv1_y.set)
        #
        #
        # tv1["column"] = list(df.columns)
        # tv1["show"] = "headings"
        # for column in tv1["column"]:
        #     tv1.heading(column, text=column)
        #
        # df_rows = df.to_numpy().tolist()
        # for row in df_rows:
        #     tv1.insert("", "end", values=row)
    ############################################################################################################

    # frame_monthly_energy_data = LabelFrame(tab_PV_results, text="Monthly Energy Flow", padx=10, pady=10)
    # frame_monthly_energy_data.grid(row=9, column=0, columnspan=9, rowspan=14)
    # Label(frame_monthly_energy_data, text="PV").grid(row=0, column=1)
    # Label(frame_monthly_energy_data, text="WT").grid(row=0, column=2)
    # Label(frame_monthly_energy_data, text="PV+WT production after peak shaving (any kind of peak shaving)").grid(row=0, column=3)
    # Label(frame_monthly_energy_data, text="Load").grid(row=0, column=4)
    # Label(frame_monthly_energy_data, text="Absorption from grid").grid(row=0, column=5)
    # Label(frame_monthly_energy_data, text="Injection into grid").grid(row=0, column=6)
    # Label(frame_monthly_energy_data, text="Grid Exchange").grid(row=0, column=7)
    # Label(frame_monthly_energy_data, text="Self Sufficiency").grid(row=0, column=8)
    # for i in range(0, len(sel_year_df)):
    #     str_pv_value =col_PV_prod[i]

    #     str_year = int(start_year)+i
    #     str_value = yearly_data_df[i]
    #
    #     row_pos = i +1
    #     Label(frame_yearly_data, text=str_year).grid(row=row_pos, column=0)
    #     Label(frame_yearly_data, text=str_value).grid(row=row_pos, column=1)
    except:
        messagebox.showerror("Error", "Error while creating Monthly energy flows dataa")
    return


def plot_pieChart(strDate, df, trend):
    try:
        sel_date = strDate.strftime("%Y-%B-%d")
        sel_year = strDate.year
        sel_month = strDate.month
        sel_day = strDate.day
        start_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)

        end_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        if sel_year < int(ent_start_year.get()) or sel_year > int(ent_end_year.get()):
            messagebox.showerror("Error", "Incorrect value. Please, enter an date between " + str(
                ent_start_year.get()) + " and " + str(ent_end_year.get()))
            return
        df = df.resample("D").sum()
        sel_date_df = df.loc[start_day: end_day]

        battery_discharge_col = sel_date_df["Battery Discharge(kW)"]
        abs_from_grid_col = sel_date_df["Absorption from grid(kW)"]
        self_sufficiency_col = sel_date_df["Self Suffiency"]
        pv_prod_col = sel_date_df["Power in ALternate Current(kW)"]
        wt_prod_col = sel_date_df["Wind Turbine Active Power(kW)"]

        if trend == "Load Supply":
            lst_labels = ["Battery Discharge", "Absorption from grid", "Self Sufficiency"]
            lst_values = [battery_discharge_col[0], abs_from_grid_col[0], self_sufficiency_col[0]]
            myexplode = [0, 0, 0.2]

            plt.pie(lst_values, labels=lst_labels, explode=myexplode, autopct='%1.1f%%')
            plt.legend(lst_labels, loc="upper right")
            plt.title("Load Supply on " + sel_date)
            plt.show()
        elif trend == "Energy Generation":
            lst_labels = ["PV production", "WT production"]
            lst_values = [pv_prod_col[0], wt_prod_col[0]]
            myexplode = [0, 0.2]

            plt.pie(lst_values, labels=lst_labels, explode=myexplode, autopct='%1.1f%%')
            plt.legend(lst_labels, loc="upper right")
            plt.title("Load Supply on " + sel_date)
            plt.show()
    except:
        messagebox.showerror("Error", "Error while creeating load supply pie chart")

    return


def export_daily_data(strDate, df):
    try:
        sel_date = strDate.strftime("%Y-%B-%d")
        sel_year = strDate.year
        sel_month = strDate.month
        sel_day = strDate.day
        start_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        load = float(ent_load.get())

        end_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        if sel_year < int(ent_start_year.get()) or sel_year > int(ent_end_year.get()):
            messagebox.showerror("Error", "Incorrect value. Please, enter an date between " + str(
                ent_start_year.get()) + " and " + str(ent_end_year.get()))
            return
        sel_date_df = df.loc[start_day: end_day]

        # export_to_excel_with_name(sel_date_df, sel_date+"Daily data")
        # export_to_excel(sel_date_df)

        sel_date_df.to_excel(
            r"D:\Thesis\dailydata" + sel_date + ".xlsx", sheet_name="Daily Data", index=False)

        messagebox.showinfo("Task Completed", "Export to excel completed")
    except:
        messagebox.showerror("Error", "Error while exporting data")


def show_daily_data(strDate, df):
    try:
        sel_date = strDate.strftime("%Y-%B-%d")
        sel_year = strDate.year
        sel_month = strDate.month
        sel_day = strDate.day
        start_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        load = float(ent_load.get())

        end_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        if sel_year < int(ent_start_year.get()) or sel_year > int(ent_end_year.get()):
            messagebox.showerror("Error", "Incorrect value. Please, enter an date between " + str(
                ent_start_year.get()) + " and " + str(ent_end_year.get()))
            return
        sel_date_df = df.loc[start_day: end_day]

        show_data(sel_date_df)
    except:
        messagebox.showerror("Error", "Error while extracting daily data")
    return


def show_daily_trend(strDate, df, trend):
    try:
        # sel_date = strDate.strftime("%d-%m-%Y")
        sel_date = strDate.strftime("%Y-%B-%d")
        sel_year = strDate.year
        sel_month = strDate.month
        sel_day = strDate.day
        start_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        load = float(ent_load.get())

        end_day = str(sel_month) + "-" + str(sel_day) + "-" + str(sel_year)
        if sel_year < int(ent_start_year.get()) or sel_year > int(ent_end_year.get()):
            messagebox.showerror("Error", "Incorrect value. Please, enter an date between " + str(
                ent_start_year.get()) + " and " + str(ent_end_year.get()))
            return

        sel_date_df = df.loc[start_day: end_day]
        # sel_date_AC_df=sel_date_df.filter(["Power in ALternate Current(kW)"])
        # sel_date_irr_df = sel_date_df.filter(["G(i)"])
        sel_date_AC_col = sel_date_df["Power in ALternate Current(kW)"]
        sel_date_irr_col = sel_date_df["G(i)"]
        sel_date_temp_col = sel_date_df["Cell Temp"]
        sel_date_WT_col = sel_date_df["Wind Turbine Active Power(kW)"]
        sel_date_batt_discharge_col = sel_date_df["Battery Discharge(kW)"]
        sel_date_batt_charge_col = sel_date_df["Battery Charge(kW)"]
        sel_date_overall_prod_with_PS_col = sel_date_df["Overall production with Peak Shaving(kW)"]
        sel_date_abs_grid_col = sel_date_df['Absorption from grid(kW)']
        sel_date_inj_grid_without_PS_col = sel_date_df["Injection in grid(kW)"]
        sel_date_inj_grid_with_PS_col = sel_date_df["Injection in grid with Peak shaving(kW)"]
        sel_date_load_col = sel_date_df["Load(kW)"]

        lst_pv_prod = []
        lst_irr_values = []
        lst_temp_values = []
        lst_wt_prod = []
        lst_load = []
        lst_batt_discharge = []
        lst_batt_charge = []
        lst_overall_prod_with_PS = []
        lst_abs_grid = []
        lst_inj_grid_without_PS = []
        lst_inj_grid_with_PS = []

        for i in range(0, len(sel_date_AC_col)):
            lst_pv_prod.append(sel_date_AC_col[i])
            lst_irr_values.append(sel_date_irr_col[i])
            lst_temp_values.append(sel_date_temp_col[i])
            lst_wt_prod.append(sel_date_WT_col[i])
            lst_load.append(sel_date_load_col[i])
            lst_batt_discharge.append(sel_date_batt_discharge_col[i])
            lst_batt_charge.append(sel_date_batt_charge_col[i])
            lst_overall_prod_with_PS.append(sel_date_overall_prod_with_PS_col[i])
            lst_abs_grid.append(sel_date_abs_grid_col[i])
            lst_inj_grid_without_PS.append(sel_date_inj_grid_without_PS_col[i])
            lst_inj_grid_with_PS.append(sel_date_inj_grid_with_PS_col[i])

        x_values = np.arange(1, len(sel_date_AC_col)+1, 1)
        x_label = "Hour"

        if trend == "PV production":
            y1_label = "Production (kW)"
            plt.plot(x_values, lst_wt_prod, label="WT")
            plt.plot(x_values, lst_pv_prod, label="PV")
            plt.plot(x_values, lst_load, label="Load")
            plt.plot(x_values, lst_batt_discharge, label="Battery discharge")
            plt.plot(x_values, lst_batt_charge, label="Battery charge")
            plt.xlabel(x_label)
            plt.ylabel(y1_label)
            plt.title("Production on " + sel_date)
            plt.legend()
            plt.show()
            # plot_graph(x_values, lst_pv_prod, x_label, y1_label, "PV production on " + sel_date)


        elif trend == "Irradiance":
            y1_label = "Irradiance (W/m2)"
            plot_graph(x_values, lst_irr_values, x_label, y1_label, "Irradiance on " + sel_date)
        elif trend == "Cell Temperature":
            y1_label = "Cell Temperature (C)"
            plot_graph(x_values, lst_temp_values, x_label, y1_label, "Cell Temperature on " + sel_date, )
        elif trend == "Grid Exchange":
            y1_label = "Grid Exchange (kW)"
            # plot_graph(x_values, lst_wt_prod, x_label, y1_label, "WT production on " + sel_date)
            mpl.rcParams['lines.linewidth'] = 2
            # plt.plot(x_values, lst_wt_prod, marker= "o")
            plt.plot(x_values, lst_overall_prod_with_PS, label="PV+WT with P.S")
            plt.plot(x_values, lst_load, label="Load")
            plt.plot(x_values, lst_abs_grid, label="Absorption from grid")
            plt.plot(x_values, lst_inj_grid_without_PS, label="Injection into grid without P.S")
            plt.plot(x_values, lst_inj_grid_with_PS, label="Injection into grid with P.S")
            plt.xlabel(x_label)
            plt.ylabel(y1_label)
            plt.title("Grid Exchange on " + sel_date)
            plt.legend()
            plt.show()
        # y2_label = "Irradiance"
        # plot_2y_graph(x_values,lst_pv_prod,lst_irr_values,x_label,y1_label,y2_label)
    except:
        messagebox.showerror("Error", "Python found a error while calculating daily trend")
    return


def show_grid_ex_barGraph_monthly(df, totalyears, startyear, rad_value):
    try:
        data = df.resample('M').sum()
        list_of_yearly_WT_data = []
        list_of_month_WT_data = []
        j = 0
        if rad_value == 1:
            str_SI = "kWh/month"
            div_df = df["Wind Turbine Active Power(kW)"]
        elif rad_value == 2:
            str_SI = "MWh/month"
            data = data.div(1000)
            div_df = df["Wind Turbine Active Power(kW)"].div(1000)
        elif rad_value == 3:
            str_SI = "GWh/month"
            div_df = df["Wind Turbine Active Power(kW)"].div(1000000)
            data = data.div(1000000)
            # df["Wind Turbine Active Power(kW)"] = df["Wind Turbine Active Power(kW)"].div(1000000)
        else:
            messagebox.showwarning("Warning", "Please select the SI unit")
            return

        # all_WT_df = df["Wind Turbine Active Power(kW)"]
        all_WT_df = div_df.resample('M').sum()
        for i in range(0, totalyears + 1):

            list_of_month_WT_data = []

            monthly_WT_data = all_WT_df.iloc[j: j + 12]

            for eachitem in monthly_WT_data:
                list_of_month_WT_data.append(eachitem)

            list_of_yearly_WT_data.append(list_of_month_WT_data)
            j += 12

        WT_df = pd.DataFrame()

        m = 0
        for each_yearly_datas in list_of_yearly_WT_data:
            WT_df.insert(m, str(int(startyear) + m), each_yearly_datas)
            m += 1

        WT_df = WT_df.mean(axis=1)

        ################################################################################################################
        data = data.iloc[0: 12]
        lst_inj_grid = []
        lst_abs_grid = []
        lst_self_sufficiency = []

        lst_inj_grid = data['Injection in grid with Peak shaving(kW)'].to_list()
        lst_abs_grid = data["Absorption from grid(kW)"].to_list()
        lst_self_sufficiency = data["Self Suffiency"].to_list()
        lst_y2Values = []

        for eachData in WT_df:
            lst_y2Values.append(eachData)

        x_value = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        x_values_int = np.arange(len(x_value))
        y2_value = lst_y2Values
        # plt.bar(x_values_int, y2_value, color='b', label="WT")
        plt.bar(x_values_int - 0.2, lst_inj_grid, color='r', label="Injection into Grid", width=0.2)
        plt.bar(x_values_int, lst_abs_grid, color='b', label="absorption from Grid", width=0.2)
        plt.bar(x_values_int + 0.2, lst_self_sufficiency, color='g', label="Self Sufficiency", width=0.2)
        plt.xlabel("Months")
        plt.ylabel("Exchange with grid (" + str_SI + ")")
        plt.xticks(x_values_int, x_value)
        plt.legend()
        plt.show()
        return
    except:
        messagebox.showerror("Error", "Python found a error while plotting monthly WT data")


def show_PV_barGraph_monthly(df, totalyears, startyear, rad_value):
    try:
        # load =float(ent_load.get())*24*30
        df = df.resample('M').sum()
        list_of_yearly_PV_data = []
        lst_of_yearly_load = []
        list_of_yearly_WT_data = []
        list_of_month_PV_data = []
        lst_yearly_overall_data = []
        j = 0
        if rad_value == 1:
            str_SI = "kWh/month"
        elif rad_value == 2:
            str_SI = "MWh/month"
            df["Power in ALternate Current(kW)"] = df["Power in ALternate Current(kW)"].div(1000).round(2)
            df["Overall production with Peak Shaving(kW)"] = df["Overall production with Peak Shaving(kW)"].div(
                1000).round(2)
            df["Load(kW)"] = df["Load(kW)"].div(1000).round(2)

        elif rad_value == 3:
            str_SI = "GWh/month"
            df["Power in ALternate Current(kW)"] = df["Power in ALternate Current(kW)"].div(1000000).round(2)
            df["Overall production with Peak Shaving(kW)"] = df["Overall production with Peak Shaving(kW)"].div(
                1000000).round(2)
            df["Load(kW)"] = df["Load(kW)"].div(1000000).round(2)

        else:
            messagebox.showwarning("Warning", "Please select the SI unit")
            return
        # all_PV_df = df["Power in ALternate Current(kW)"].resample('M').sum()
        # all_WT_df = df["Wind Turbine Active Power(kW)"].resample('M').sum()
        # df = df.resample('M').sum()
        for i in range(0, totalyears + 1):
            list_of_month_PV_data = []
            lst_month_overall_data = []
            lst_month_load_data = []
            # list_of_month_WT_data =[]
            yearly_PV_data = df.iloc[j: j + 12, 9:10]
            yearly_overall_data = df.iloc[j: j + 12, 13:14]
            yaerly_load_data = df.iloc[j: j + 12, 4:5]
            # monthly_WT_data = all_WT_df.iloc[j: j + 12]
            monthly_PV_data = yearly_PV_data["Power in ALternate Current(kW)"]
            monthly_overall_data = yearly_overall_data["Overall production with Peak Shaving(kW)"]
            monthly_load_data = yaerly_load_data["Load(kW)"]
            # monthly_WT_data = yearly_PV_data["Wind Turbine Active Power(kW)"]
            # yearly_data = yearly_data.rename({'G(i)': str(startyear+k)}, axis=1)
            for eachitem in monthly_PV_data:
                list_of_month_PV_data.append(eachitem)
            for eachitem in monthly_load_data:
                lst_month_load_data.append(eachitem)
            for eachitem in monthly_overall_data:
                lst_month_overall_data.append(eachitem)
            # for eachitem in monthly_WT_data:
            #     list_of_month_WT_data.append(eachitem)

            list_of_yearly_PV_data.append(list_of_month_PV_data)
            lst_yearly_overall_data.append(lst_month_overall_data)
            lst_of_yearly_load.append(lst_month_load_data)
            # list_of_yearly_WT_data.append(list_of_month_WT_data)
            j += 12
        PV_df = pd.DataFrame()
        overall_df = pd.DataFrame()
        load_df = pd.DataFrame()
        # WT_df = pd.DataFrame()
        k = 0
        for each_yearly_datas in list_of_yearly_PV_data:
            PV_df.insert(k, str(int(startyear) + k), each_yearly_datas)
            k += 1

        k = 0
        for each_yearly_datas in lst_of_yearly_load:
            load_df.insert(k, str(int(startyear) + k), each_yearly_datas)
            k += 1

        k = 0
        for each_yearly_datas in lst_yearly_overall_data:
            overall_df.insert(k, str(int(startyear) + k), each_yearly_datas)
            k += 1
        # m=0
        # for each_yearly_datas in list_of_yearly_WT_data:
        #     WT_df.insert(m,str(int(startyear) + m),each_yearly_datas)
        #     m +=1

        PV_df = PV_df.mean(axis=1)
        overall_df = overall_df.mean(axis=1)
        load_df = load_df.mean(axis=1)
        # WT_df =WT_df.mean(axis=1)
        lst_yValues = []
        lst_overll_values = []
        lst_load_values = []
        # lst_y2Values =[]
        lst_load = []
        for eachData in load_df:
            # lst_yValues.append(eachData)
            lst_yValues.append(eachData)

        for eachData in overall_df:
            lst_overll_values.append(eachData)
        # for eachData in WT_df:
        #     lst_y2Values.append(eachData)

        # x_values_int=[1,2,3,4,5,6,7,8,9,10,11,12]

        x_value = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        x_values_int = np.arange(len(x_value))
        y_value = lst_yValues
        y2_value = lst_overll_values
        plt.bar(x_values_int - 0.2, y_value, color='r', label="Load", width=0.4)
        plt.bar(x_values_int + 0.2, y2_value, color='b', label="PV+WT", width=0.4)
        # for i in range(0, len(y_value)):
        #     plt.bar(i+1, y_value[i],color = 'r')
        plt.xlabel("Months")
        plt.ylabel("PV+WT Production (" + str_SI + ")")
        plt.xticks(x_values_int, x_value)
        plt.legend()
        plt.show()



    except:
        messagebox.showerror("Error", "Python found a error while plotting monthly PV data")
    return


def convert_SI_units(lst_lbls, lst_values, lst_ent, var_results):
    try:
        lst_kWh = ["kWh/kW/year", "kWh/year", "kWh/kW/year", "kWh/year", "kWh/kW/year", "kWh/year", "kWh/kW/year",
                   "kWh/year", "kWh/year", "kWh/year", "kWh/year"]
        lst_MWh = ["MWh/MW/year", "MWh/year", "MWh/MW/year", "MWh/year", "MWh/MW/year", "MWh/year", "MWh/MW/year",
                   "MWh/year", "MWh/year", "MWh/year", "MWh/year"]
        lst_GWh = ["GWh/GW/year", "GWh/year", "GWh/GW/year", "GWh/year", "GWh/GW/year", "GWh/year", "GWh/GW/year",
                   "GWh/year", "GWh/year", "GWh/year", "GWh/year"]
        if var_results.get() == "kWh":
            for i in range(0, 11):
                lst_lbls[i].config(text=lst_kWh[i])

            for i in range(0, 7):
                value = int(lst_values[i]) * 1000
                lst_ent[i].delete(0, END)
                lst_ent[i].insert(END, value)

        elif var_results.get() == "MWh":
            for i in range(0, 11):
                lst_lbls[i].config(text=lst_MWh[i])

            for i in range(0, 7):
                value = int(lst_values[i])
                lst_ent[i].delete(0, END)
                lst_ent[i].insert(END, value)

        elif var_results.get() == "GWh":
            for i in range(0, 11):
                lst_lbls[i].config(text=lst_GWh[i])

            for i in range(0, 7):
                value = int(lst_values[i]) / 1000
                lst_ent[i].delete(0, END)
                lst_ent[i].insert(END, value)

    except:
        messagebox.showerror("Error", "Error while coverting SI units")


def export_to_excel_with_name(df, name):
    try:
        df.to_excel(r"D:\Thesis\_" + name + ".xlsx", sheet_name="NPV Data", index=False)

        messagebox.showinfo("Task Completed", "Export to excel completed")
    except:
        messagebox.showerror("Error", "Python found a error while exporting data to excel ")


def export_to_excel(df):
    try:
        df.to_excel(
            r"D:\Thesis\data_Lat_" + str(ent_latitude.get()) + "_Lon_" + str(ent_longitude.get()) + "_from_" + str(
                ent_start_year.get()) + "_to_" + str(ent_end_year.get()) + ".xlsx", sheet_name="All Data", index=False)

        messagebox.showinfo("Task Completed", "Export to excel completed")
    except:
        messagebox.showerror("Error", "Python found a error while exporting data to excel ")


def show_power_curve(df, strturbine):
    try:
        # col_wind_speed = df["Wind speed(m/s)"].tolist()
        int_col_wind_speed = df.columns.get_loc(strturbine) - 1
        col_pow_turbine = df[strturbine].tolist()
        col_wind_speed = df.iloc[:, int_col_wind_speed].tolist()
        plot_graph(col_wind_speed, col_pow_turbine, "Wind speed(m/s)", "Active Power(kW)",
                   "Power curve of " + strturbine)
        return
    except:
        messagebox.showerror("Error", "Python found a error while plotting power curve")


def disable_event():
    pass


def compare_power_curve(df):
    try:
        c = 0
        end_value = int(len(df.columns) / 2)
        for i in range(0, end_value):
            lgnd = df.columns[c + 1]
            xvalues = df.iloc[:, c].tolist()
            yvalues = df.iloc[:, c + 1].tolist()
            # plot_graph(xvalues,yvalues,"Wind speed(m/s)","Active Power(kW)","Comparison")
            mpl.rcParams['lines.linewidth'] = 2
            plt.plot(xvalues, yvalues, label=lgnd)
            c += 2

        plt.xlabel("Wind speed(m/s)")
        plt.ylabel("Active Power(kW)")
        plt.title("Comparison")
        plt.legend()
        plt.show()

    except:
        messagebox.showerror("Error", "Python found a error while plotting comparison power curve")


def processing():
    new = Toplevel(root)
    new.geometry("200x150")
    lbl = Label(new, text="--")
    lbl.grid()
    for i in range(50):
        time.sleep(0.1)
        # Un-comment the line below to fix
        # root.update()
        print(i)
        lbl['text'] = "{}".format(i)
    new.destroy()


def update_results_in_main_wind(data, results_df, nominal_pow, sel_WT_nom_pow, nb_WT):
    try:
        # udating values of results DF in main window
        all_data_sum = data.resample('Y').sum()
        # col_pv_prod_without_ps = all_data_sum[""]
        col_pv_prod_with_ps = all_data_sum['Production of PV after peak shaving(kW)'].div(1000)
        # col_WT_prod_without_ps= all_data_sum[""]
        col_WT_prod_with_ps = all_data_sum["Production of WT after peak shaving(kW)"].div(1000)
        results_col = results_df[str(start_year)]
        if nominal_pow != 0:
            productivity_pv_with_ps = (col_pv_prod_with_ps[0] / nominal_pow) * 1000
        else:
            productivity_pv_with_ps = 0
        productivity_WT_with_ps = (col_WT_prod_with_ps[0] / sel_WT_nom_pow) * 1000

        frm_results_main_win = LabelFrame(tab_inputs, text="Results of First year")
        frm_results_main_win.place(x=820, y=40, height=440, width=440)

        lbl_pv_prodcutivity = Label(frm_results_main_win, text="PV productivity")
        lbl_pv_prodcutivity.place(x=10, y=10)
        lbl_SI_pv_prodcutivity = Label(frm_results_main_win, text="MWh/MW/year")
        lbl_SI_pv_prodcutivity.place(x=240, y=10)
        ent_pv_prodcutivity = Entry(frm_results_main_win, width=10)
        ent_pv_prodcutivity.place(x=340, y=10)
        ent_pv_prodcutivity.insert(END, str(results_col[8]))

        lbl_pv_annual_prod = Label(frm_results_main_win, text="Annual PV production")
        lbl_pv_annual_prod.place(x=10, y=35)
        lbl_SI_pv_annual_prod = Label(frm_results_main_win, text="MWh/year")
        lbl_SI_pv_annual_prod.place(x=240, y=35)
        ent_pv_annual_prod = Entry(frm_results_main_win, width=10)
        ent_pv_annual_prod.place(x=340, y=35)
        ent_pv_annual_prod.insert(END, str(float(results_col[8]) * nominal_pow / 1000))

        lbl_pv_productivity_with_ps = Label(frm_results_main_win, text="Productivity of PV with Peak shaving")
        lbl_pv_productivity_with_ps.place(x=10, y=60)
        lbl_SI_pv_productivity_with_ps = Label(frm_results_main_win, text="MWh/MW/year")
        lbl_SI_pv_productivity_with_ps.place(x=240, y=60)
        ent_pv_productivity_with_ps = Entry(frm_results_main_win, width=10)
        ent_pv_productivity_with_ps.place(x=340, y=60)
        ent_pv_productivity_with_ps.insert(END, str(round(productivity_pv_with_ps)))

        lbl_Annual_pv_prod_with_ps = Label(frm_results_main_win, text="Annual PV production with Peak shaving")
        lbl_Annual_pv_prod_with_ps.place(x=10, y=85)
        lbl_SI_Annual_pv_prod_with_ps = Label(frm_results_main_win, text="MWh/year")
        lbl_SI_Annual_pv_prod_with_ps.place(x=240, y=85)
        ent_Annual_pv_prod_with_ps = Entry(frm_results_main_win, width=10)
        ent_Annual_pv_prod_with_ps.place(x=340, y=85)
        ent_Annual_pv_prod_with_ps.insert(END, str(round(col_pv_prod_with_ps[0])))

        lbl_WT_prodcutivity = Label(frm_results_main_win, text="WT productivity")
        lbl_WT_prodcutivity.place(x=10, y=110)
        lbl_SI_WT_prodcutivity = Label(frm_results_main_win, text="MWh/MW/year")
        lbl_SI_WT_prodcutivity.place(x=240, y=110)
        ent_WT_prodcutivity = Entry(frm_results_main_win, width=10)
        ent_WT_prodcutivity.place(x=340, y=110)
        ent_WT_prodcutivity.insert(END, str(results_col[10]))

        lbl_WT_annual_prod = Label(frm_results_main_win, text="Annual WT production")
        lbl_WT_annual_prod.place(x=10, y=135)
        lbl_SI_WT_annual_prod = Label(frm_results_main_win, text="MWh/year")
        lbl_SI_WT_annual_prod.place(x=240, y=135)
        ent_WT_annual_prod = Entry(frm_results_main_win, width=10)
        ent_WT_annual_prod.place(x=340, y=135)
        ent_WT_annual_prod.insert(END, str(float(results_col[10]) * nb_WT * sel_WT_nom_pow / 1000))

        lbl_WT_productivity_with_ps = Label(frm_results_main_win, text="Productivity of WT with Peak shaving")
        lbl_WT_productivity_with_ps.place(x=10, y=160)
        lbl_SI_WT_productivity_with_ps = Label(frm_results_main_win, text="MWh/MW/year")
        lbl_SI_WT_productivity_with_ps.place(x=240, y=160)
        ent_WT_productivity_with_ps = Entry(frm_results_main_win, width=10)
        ent_WT_productivity_with_ps.place(x=340, y=160)
        if nb_WT != 0:
            ent_WT_productivity_with_ps.insert(END, str(round(productivity_WT_with_ps / nb_WT)))
        else:
            ent_WT_productivity_with_ps.insert(END, str(0))
        lbl_Annual_WT_prod_with_ps = Label(frm_results_main_win, text="Annual WT production with peak shaving")
        lbl_Annual_WT_prod_with_ps.place(x=10, y=185)
        lbl_SI_Annual_WT_prod_with_ps = Label(frm_results_main_win, text="MWh/year")
        lbl_SI_Annual_WT_prod_with_ps.place(x=240, y=185)
        ent_Annual_WT_prod_with_ps = Entry(frm_results_main_win, width=10)
        ent_Annual_WT_prod_with_ps.place(x=340, y=185)
        ent_Annual_WT_prod_with_ps.insert(END, str(round(col_WT_prod_with_ps[0])))

        lbl_dischrged_energy = Label(frm_results_main_win, text="Discharged Energy")
        lbl_dischrged_energy.place(x=10, y=210)
        lbl_SI_dischrged_energy = Label(frm_results_main_win, text="MWh/year")
        lbl_SI_dischrged_energy.place(x=240, y=210)
        ent_dischrged_energy = Entry(frm_results_main_win, width=10)
        ent_dischrged_energy.place(x=340, y=210)
        ent_dischrged_energy.insert(END, str(results_col[20]))

        lbl_chrged_energy = Label(frm_results_main_win, text="Charged Energy")
        lbl_chrged_energy.place(x=10, y=235)
        lbl_SI_chrged_energy = Label(frm_results_main_win, text="MWh/year")
        lbl_SI_chrged_energy.place(x=240, y=235)
        ent_chrged_energy = Entry(frm_results_main_win, width=10)
        ent_chrged_energy.place(x=340, y=235)
        ent_chrged_energy.insert(END, str(results_col[19]))

        lbl_self_consumption = Label(frm_results_main_win, text="Self Consumption")
        lbl_self_consumption.place(x=10, y=260)
        lbl_SI_self_consumption = Label(frm_results_main_win, text="%")
        lbl_SI_self_consumption.place(x=240, y=260)
        ent_self_consumption = Entry(frm_results_main_win, width=10)
        ent_self_consumption.place(x=340, y=260)
        ent_self_consumption.insert(END, str(results_col[22]))

        lbl_self_sufficiency = Label(frm_results_main_win, text="Self Sufficiency")
        lbl_self_sufficiency.place(x=10, y=285)
        lbl_SI_self_sufficiency = Label(frm_results_main_win, text="%")
        lbl_SI_self_sufficiency.place(x=240, y=285)
        ent_self_sufficiency = Entry(frm_results_main_win, width=10)
        ent_self_sufficiency.place(x=340, y=285)
        ent_self_sufficiency.insert(END, str(results_col[23]))

        lbl_absorption_from_grid = Label(frm_results_main_win, text="Absorption from the grid / load")
        lbl_absorption_from_grid.place(x=10, y=310)
        lbl_SI_absorption_from_grid = Label(frm_results_main_win, text="%")
        lbl_SI_absorption_from_grid.place(x=240, y=310)
        ent_absorption_from_grid = Entry(frm_results_main_win, width=10)
        ent_absorption_from_grid.place(x=340, y=310)
        ent_absorption_from_grid.insert(END, str(results_col[24]))

        lbl_injection_to_grid = Label(frm_results_main_win, text="Injection in the grid / load")
        lbl_injection_to_grid.place(x=10, y=335)
        lbl_SI_injection_to_grid = Label(frm_results_main_win, text="%")
        lbl_SI_injection_to_grid.place(x=240, y=335)
        ent_injection_to_grid = Entry(frm_results_main_win, width=10)
        ent_injection_to_grid.place(x=340, y=335)
        ent_injection_to_grid.insert(END, str(results_col[25]))
    except:
        messagebox.showerror("Error", "Error while updating results in main window")


def produce_results_df(data, sel_WT_nom_pow):
    ######Creating results df##################################################################################################
    try:
        results_df = pd.DataFrame()
        lst_description = [
            "Site localization",
            "Longitude",
            "Latitude",
            "Peak shaving",
            "Limitation of the maximum generated power",
            "Limitation of the maximum injection",
            "Size of plants",
            "Total power of PV generators",
            "Productivity of PV",
            "Total power of WT generators",
            "Productivity of WT",
            "Capacity of storage",
            "Energy flows (AC)",
            "Production from PV+WT",
            "Annual load",
            "Annual injection in the grid",
            "Annual absorption from the grid",
            "Grid exchange (absorption+injection)",
            "Battery operation (electrochemical)",
            "Battery charge",
            "Battery discharge",
            "Energy balance",
            "Self-consumption",
            "Self-sufficiency",
            "Absorption from the grid / load",
            "Injection in the grid / load",
            "Production from renewables / load"

        ]
        lst_SI_units = [
            "",
            "",
            "",
            "",
            "MW",
            "MW",
            "",
            "MW",
            "kWh/kW/year",
            "MW",
            "kWh/kW/year",
            "MWh",
            "",
            "MWh/year",
            "MWh/year",
            "MWh/year",
            "MWh/year",
            "MWh/year",
            "",
            "MWh/year",
            "MWh/year",
            "",
            "",
            '',
            "",
            "",
            ""

        ]
        results_df.insert(0, "Description", lst_description)
        results_df.insert(1, "SI units", lst_SI_units)

        all_data_sum = data.resample('Y').sum()

        total_years = int(end_year) - int(start_year)
        # load = float(ent_load.get())
        col_load = all_data_sum["Load(kW)"].div(1000)
        nb_WT = int(ent_nb_turbines.get())
        col_absorption_frm_grid = all_data_sum["Absorption from grid(kW)"].div(1000)
        col_overall_prod_after_both_PS = all_data_sum[
            "PV+WT production after peak shaving (any kind of peak shaving)"].div(1000)
        col_inj_grid_with_PS = all_data_sum["Injection in grid with Peak shaving(kW)"].div(1000)
        longitude = str(ent_longitude.get())
        latitude = str(ent_latitude.get())
        if var_SI_unit_nom_pow.get() == 1:
            nominal_pow = float(ent_nominal_power.get())
        elif var_SI_unit_nom_pow.get() == 2:
            nominal_pow = float(ent_nominal_power.get()) * 1000
        elif var_SI_unit_nom_pow.get() == 3:
            nominal_pow = float(ent_nominal_power.get()) * 1000000
        else:
            messagebox.showerror("Error", "Please select the SI unit of nominal power")
            return

        col_PV_prod_yearly = all_data_sum["Power in ALternate Current(kW)"].div(1000)
        if nominal_pow != 0:
            col_PV_prod_yearly = col_PV_prod_yearly.div(nominal_pow / 1000)
        else:
            col_PV_prod_yearly = col_PV_prod_yearly
        total_pow_WT = (sel_WT_nom_pow / 1000) * nb_WT
        col_WT_prod_yearly = all_data_sum["Wind Turbine Active Power(kW)"].div(1000)
        if sel_WT_nom_pow != 0:
            col_WT_prod_yearly = col_WT_prod_yearly.div(sel_WT_nom_pow / 1000)
        else:
            col_WT_prod_yearly = col_WT_prod_yearly
        if nb_WT != 0:
            col_WT_prod_yearly = col_WT_prod_yearly / nb_WT
        else:
            col_WT_prod_yearly = col_WT_prod_yearly
        # annual_load = load * 24 * 365 / 1000

        col_grid_Exchange = all_data_sum["Grid Exchange(kW)"].div(1000)
        col_batt_charge = all_data_sum["Battery Charge(kW)"].div(1000)
        col_batt_discharge = all_data_sum["Battery Discharge(kW)"].div(1000)

        if var_SI_unit_storage_capacity.get() == 1:
            batt_capacity = float(ent_storage_capacity.get())
        if var_SI_unit_storage_capacity.get() == 2:
            batt_capacity = float(ent_storage_capacity.get()) * 1000
        if var_SI_unit_storage_capacity.get() == 3:
            batt_capacity = float(ent_storage_capacity.get()) * 1000000
        if var_peak_shaving.get() == "Limitation on Maximum power injection":
            if var_SI_unit_PS.get() == 1:
                lim_max_inj = float(ent_max_inj.get())
            elif var_SI_unit_PS.get() == 2:
                lim_max_inj = float(ent_max_inj.get()) * 1000
            elif var_SI_unit_PS.get() == 3:
                lim_max_inj = float(ent_max_inj.get()) * 1000000
            else:
                messagebox.showerror("Error", "Error in converting SI unit of peak shaving")
                return
        else:
            lim_max_inj = 0
        if var_peak_shaving.get() == "Limitation on Maximum generated power":
            if var_SI_unit_PS.get() == 1:
                lim_max_gen = float(ent_max_gen.get())
            elif var_SI_unit_PS.get() == 2:
                lim_max_gen = float(ent_max_gen.get()) * 1000
            elif var_SI_unit_PS.get() == 3:
                lim_max_gen = float(ent_max_gen.get()) * 1000000
            else:
                messagebox.showerror("Error", "Error in converting SI unit of peak shaving")
                return
        else:
            lim_max_gen = 0

        for i in range(0, total_years + 1):
            lst_of_eachYear_data = []
            year = int(start_year) + i
            annual_self_consumption = ((col_load[i] - col_absorption_frm_grid[i]) / col_overall_prod_after_both_PS[
                i]) * 100
            annual_self_sufficiency = ((col_load[i] - col_absorption_frm_grid[i]) / col_load[i]) * 100
            annual_absorption_perc = 100 - annual_self_sufficiency
            annual_inj_grid_perc = (col_inj_grid_with_PS[i] / col_load[i]) * 100
            annual_renewable_prod_perc = (col_overall_prod_after_both_PS[i] / col_load[i]) * 100
            lst_of_eachYear_data = [
                "",
                str(longitude),
                str(latitude),
                "",
                str(lim_max_gen / 1000),
                str(lim_max_inj / 1000),
                "",
                str(round(nominal_pow / 1000)),
                str(round(col_PV_prod_yearly[i])),
                str(total_pow_WT),
                str(round(col_WT_prod_yearly[i])),
                str(round(batt_capacity / 1000)),
                "",
                str(round(col_overall_prod_after_both_PS[i])),
                str(round(col_load[i])),
                str(round(col_inj_grid_with_PS[i])),
                str(round(col_absorption_frm_grid[i])),
                str(round(col_grid_Exchange[i])),
                "",
                str(round(col_batt_charge[i])),
                str(round(col_batt_discharge[i])),
                "",
                str(round(annual_self_consumption)) + "%",
                str(round(annual_self_sufficiency)) + "%",
                str(round(annual_absorption_perc)) + "%",
                str(round(annual_inj_grid_perc)) + "%",
                str(round(annual_renewable_prod_perc)) + "%"

            ]
            results_df.insert(i + 2, str(year), lst_of_eachYear_data)


    except:
        messagebox.showerror("Error", "Error while creating Results dataframe")
    return results_df


def produce_results_df_for_opt(data, PV_nominal_pow, sel_WT_nom_pow, opt_nb_WT, batt_capacity):
    try:
        results_df = pd.DataFrame()
        lst_description = [
            "Site localization",
            "Longitude",
            "Latitude",
            "Peak shaving",
            "Limitation of the maximum generated power",
            "Limitation of the maximum injection",
            "Size of plants",
            "Total power of PV generators",
            "Productivity of PV",
            "Total power of WT generators",
            "Productivity of WT",
            "Capacity of storage",
            "Energy flows (AC)",
            "Production from PV+WT",
            "Annual load",
            "Annual injection in the grid",
            "Annual absorption from the grid",
            "Grid exchange (absorption+injection)",
            "Battery operation (electrochemical)",
            "Battery charge",
            "Battery discharge",
            "Energy balance",
            "Self-consumption",
            "Self-sufficiency",
            "Absorption from the grid / load",
            "Injection in the grid / load",
            "Production from renewables / load"

        ]
        lst_SI_units = [
            "",
            "",
            "",
            "",
            "MW",
            "MW",
            "",
            "MW",
            "kWh/kW/year",
            "MW",
            "kWh/kW/year",
            "MWh",
            "",
            "MWh/year",
            "MWh/year",
            "MWh/year",
            "MWh/year",
            "MWh/year",
            "",
            "MWh/year",
            "MWh/year",
            "",
            "",
            '',
            "",
            "",
            ""

        ]
        results_df.insert(0, "Description", lst_description)
        results_df.insert(1, "SI units", lst_SI_units)

        all_data_sum = data.resample('Y').sum()

        total_years = int(end_year) - int(start_year)
        # load = float(ent_load.get())
        col_load = all_data_sum["Load(kW)"].div(1000)
        nb_WT = opt_nb_WT
        col_absorption_frm_grid = all_data_sum["Absorption from grid(kW)"].div(1000)
        col_overall_prod_after_both_PS = all_data_sum[
            "PV+WT production after peak shaving (any kind of peak shaving)"].div(1000)
        col_inj_grid_with_PS = all_data_sum["Injection in grid with Peak shaving(kW)"].div(1000)
        longitude = str(ent_longitude.get())
        latitude = str(ent_latitude.get())
        nominal_pow = PV_nominal_pow

        col_PV_prod_yearly = all_data_sum["Power in ALternate Current(kW)"].div(1000)
        if nominal_pow != 0:
            col_PV_prod_yearly = col_PV_prod_yearly.div(nominal_pow / 1000)
        else:
            col_PV_prod_yearly = col_PV_prod_yearly

        total_pow_WT = (sel_WT_nom_pow / 1000) * nb_WT
        col_WT_prod_yearly = all_data_sum["Wind Turbine Active Power(kW)"].div(1000)
        if sel_WT_nom_pow != 0:
            col_WT_prod_yearly = col_WT_prod_yearly.div(sel_WT_nom_pow / 1000)
        else:
            col_WT_prod_yearly = col_WT_prod_yearly

        if nb_WT != 0:
            col_WT_prod_yearly = col_WT_prod_yearly / nb_WT
        else:
            col_WT_prod_yearly = col_WT_prod_yearly
        # annual_load = load * 24 * 365 / 1000

        col_grid_Exchange = all_data_sum["Grid Exchange(kW)"].div(1000)
        col_batt_charge = all_data_sum["Battery Charge(kW)"].div(1000)
        col_batt_discharge = all_data_sum["Battery Discharge(kW)"].div(1000)

        if var_peak_shaving.get() == "Limitation on Maximum power injection":
            if var_SI_unit_PS.get() == 1:
                lim_max_inj = float(ent_max_inj.get())
            elif var_SI_unit_PS.get() == 2:
                lim_max_inj = float(ent_max_inj.get()) * 1000
            elif var_SI_unit_PS.get() == 3:
                lim_max_inj = float(ent_max_inj.get()) * 1000000
            else:
                messagebox.showerror("Error", "Error in converting SI unit of peak shaving")
                return
        else:
            lim_max_inj = 0
        if var_peak_shaving.get() == "Limitation on Maximum generated power":
            if var_SI_unit_PS.get() == 1:
                lim_max_gen = float(ent_max_gen.get())
            elif var_SI_unit_PS.get() == 2:
                lim_max_gen = float(ent_max_gen.get()) * 1000
            elif var_SI_unit_PS.get() == 3:
                lim_max_gen = float(ent_max_gen.get()) * 1000000
            else:
                messagebox.showerror("Error", "Error in converting SI unit of peak shaving")
                return
        else:
            lim_max_gen = 0

        for i in range(0, total_years + 1):
            lst_of_eachYear_data = []
            year = int(start_year) + i
            try:
                annual_self_consumption = round(
                    ((col_load[i] - col_absorption_frm_grid[i]) / col_overall_prod_after_both_PS[
                        i]) * 100)
            except:
                annual_self_consumption = "inf"
            annual_self_sufficiency = ((col_load[i] - col_absorption_frm_grid[i]) / col_load[i]) * 100
            annual_absorption_perc = 100 - annual_self_sufficiency
            annual_inj_grid_perc = (col_inj_grid_with_PS[i] / col_load[i]) * 100
            annual_renewable_prod_perc = (col_overall_prod_after_both_PS[i] / col_load[i]) * 100
            lst_of_eachYear_data = [
                "",
                str(longitude),
                str(latitude),
                "",
                str(lim_max_gen / 1000),
                str(lim_max_inj / 1000),
                "",
                str(nominal_pow / 1000),
                str(round(col_PV_prod_yearly[i])),
                str(total_pow_WT),
                str(round(col_WT_prod_yearly[i])),
                str(batt_capacity / 1000),
                "",
                str(round(col_overall_prod_after_both_PS[i])),
                str(round(col_load[i])),
                str(round(col_inj_grid_with_PS[i])),
                str(round(col_absorption_frm_grid[i])),
                str(round(col_grid_Exchange[i])),
                "",
                str(round(col_batt_charge[i])),
                str(round(col_batt_discharge[i])),
                "",
                str((annual_self_consumption)) + "%",
                str(round(annual_self_sufficiency)) + "%",
                str(round(annual_absorption_perc)) + "%",
                str(round(annual_inj_grid_perc)) + "%",
                str(round(annual_renewable_prod_perc)) + "%"

            ]
            results_df.insert(i + 2, str(year), lst_of_eachYear_data)


    except:
        messagebox.showerror("Error", "Error while creating Results dataframe")
    return results_df


def insert_list_into_list(mainList, smallList):
    for i in range(0, len(smallList)):
        mainList.append(smallList[i])
    return mainList


def create_load_list_frm_25years(total_years):
    try:
        load_wb = xl.load_workbook(ent_load_type.get())
        sheet_load = load_wb["Sheet1"]

        max_row = sheet_load.max_row
        max_col = sheet_load.max_column
        if max_col != 26:
            messagebox.showerror("Error", "Upload Excel file with Load of 25 years")
            return
        lst_load_25_years = []
        int_year = 0
        for j in range(2, total_years + 3):
            for i in range(1, max_row):
                lst_load_25_years.append(sheet_load.cell(i + 1, j).value)
            if (int(start_year) + int_year) % 4 == 0:
                lst_last_day = lst_load_25_years[-24:]
                lst_load_25_years = insert_list_into_list(lst_load_25_years, lst_last_day)

            int_year += 1
        return lst_load_25_years
    except:
        messagebox.showerror("Error", "Error while reading Load excel file from 25 years")


def create_load_list_frm_1st_year(total_years):
    # reading load excel file
    try:
        load_wb = xl.load_workbook(ent_load_type.get())
        sheet_load = load_wb["Sheet1"]
        if sheet_load.max_column != 3:
            messagebox.showerror("Error", "Load Proper excel file")
            return
        max_row = sheet_load.max_row
        col_load_index = sheet_load.cell(1, 2)
        lst_load_frm_excel = []
        for i in range(1, max_row):
            lst_load_frm_excel.append(sheet_load.cell(i + 1, 2).value)
        lst_percentage = []
        percentage = 100
        lst_percentage.append(percentage)
        for t in range(2, 26):
            percentage = sheet_load.cell(t + 1, 3).value + percentage
            lst_percentage.append(percentage)
            # perc_increase = perc_increase+percentage/100
        t = len(lst_load_frm_excel)

        lst_load_all_years = []
        lst_load_feb_29 = []
        for j in range(1392, 1416):
            lst_load_feb_29.append(lst_load_frm_excel[j])

        lst_load_with_feb29 = []
        lst_load_with_feb29 = insert_list_into_list(lst_load_with_feb29, lst_load_frm_excel)
        lst_load_with_feb29[1392:1392] = lst_load_feb_29

        lst_load_all_years = []
        for i in range(0, total_years + 1):
            if (int(start_year) + i) % 4 != 0:
                int_percentage = lst_percentage[i] / 100
                lst_load_frm_excel = [item * int_percentage for item in lst_load_frm_excel]
                lst_load_all_years = insert_list_into_list(lst_load_all_years, lst_load_frm_excel)
            else:
                int_percentage = lst_percentage[i] / 100
                lst_load_with_feb29 = [item * int_percentage for item in lst_load_with_feb29]
                lst_load_all_years = insert_list_into_list(lst_load_all_years, lst_load_with_feb29)

        return lst_load_all_years
    except:
        messagebox.showerror("Error", "Error while reading Load excel file from 1st year")
        return


def change_data_for_asia(data):
    lst_time = data["time"].tolist()
    lst_gi = data["G(i)"].tolist()
    lst_H_sun = data["H_sun"].tolist()
    lst_T2m = data["T2m"].tolist()
    lst_WS10m = data["WS10m"].tolist()
    lst_int= data["Int"].tolist()

#####################################################################################################################
    lst_time.pop(0)
    first_item = lst_time[0]
    lst_time.insert(0,first_item)
#####################################################################################################################

    last_gi = lst_gi[len(lst_gi) - 6:]
    last_H_sun = lst_H_sun[len(lst_gi) - 6:]
    last_T2m = lst_T2m[len(lst_gi) - 6:]
    last_WS10m = lst_WS10m[len(lst_gi) - 6:]
    last_int = lst_int[len(lst_gi) - 6:]

    lst_gi = lst_gi[:len(lst_gi) - 6]
    lst_H_sun = lst_H_sun[:len(lst_H_sun) - 6]
    lst_T2m = lst_T2m[:len(lst_T2m) - 6]
    lst_WS10m = lst_WS10m[:len(lst_WS10m) - 6]
    lst_int = lst_int[:len(lst_int) - 6]

    for i in range (0, len(last_gi)):
        lst_gi.insert(i,last_gi[i])
        lst_H_sun.insert(i, last_H_sun[i])
        lst_T2m.insert(i, last_T2m[i])
        lst_WS10m.insert(i, last_WS10m[i])
        lst_int.insert(i, last_int[i])

    df = pd.DataFrame()
    df.insert(0,"time",lst_time)
    df.insert(1, "G(i)", lst_gi)
    df.insert(2, "H_sun", lst_H_sun)
    df.insert(3, "T2m", lst_T2m)
    df.insert(4, "WS10m", lst_WS10m)
    df.insert(5, "Int", lst_int)


    return df

def download_data():
    time_start = tm.time()
    if ent_start_year.get() > ent_end_year.get():
        messagebox.showerror("Date error", "Starting Year should be less than end year")
        return
    if len(ent_start_year.get()) != 4 or len(ent_end_year.get()) != 4:
        messagebox.showerror("Date error", "starting year or Ending year is not entered properly")
        return
    if int(ent_start_year.get()) < 2005 or int(ent_end_year.get()) > 2016:
        messagebox.showerror("Date error", "Incorrect value. Please, enter an integer between 2005 and 2016")
        return
    if ent_load_type.get() == "":
        messagebox.showerror("Error", "Please select excel file with hourly load values")
        return

    latitude = str(ent_latitude.get())
    longitude = str(ent_longitude.get())
    global start_year
    global end_year
    start_year = str(ent_start_year.get())
    end_year = str(ent_end_year.get())
    total_years = int(end_year) - int(start_year)
    if var_load.get() == "Load of 1 year and increase %":
        lst_load_all_years = create_load_list_frm_1st_year(total_years)
        if lst_load_all_years == None:
            return
    elif var_load.get() == "Load of 25 years":
        lst_load_all_years = create_load_list_frm_25years(total_years)
        if lst_load_all_years == None:
            return
    ########################################################################################################################
    try:

        # progress bar start
        newWindow = Toplevel(root)
        newWindow.title("Progress Bar")
        newWindow.geometry('400x100+400+250')

        newWindow.protocol("WM_DELETE_WINDOW", disable_event)

        # progress_bar = ttk.Progressbar(newWindow, orient=HORIZONTAL, length=400, mode="determinate")
        # progress_bar.pack()
        lbl_progress_bar = Label(newWindow, text="Calculation in progress")
        lbl_progress_bar.pack()

        # progress_bar["value"] += 2
        #
        # lbl_progress_bar.config(text="Downloading Data")
        root.update()
    except:
        messagebox.showerror("Error")
    #######################################################################################################################
    time_start_download_data = tm.time()
    url_first_part = "https://re.jrc.ec.europa.eu/api/seriescalc?"
    final_url = url_first_part + "lat=" + latitude + "&lon=" + longitude + "&startyear=" + start_year \
                + "&endyear=" + end_year + "&optimalangles=1&outputformat=json&browser=1"
    res = requests.get(final_url)
    res_json_file = io.StringIO(res.text)
    src = json.load(res_json_file)
    output = src['outputs']
    output_hourly = output["hourly"]
    data = pd.DataFrame(output_hourly)
    ########################################################################################################################

    if var_India.get() == 1:
        data = change_data_for_asia(data)


    ########################################################################################################################
    data.index = pd.to_datetime(data["time"], format='%Y%m%d:%H%M', utc=True)
    data = data.drop("Int", axis=1)
    data_raw = data.drop("time", axis=1)

    time_end_download_data = tm.time()
    total_time_to_download_data = round(time_end_download_data - time_start_download_data)

    global data_for_opt
    data_for_opt = data
    ########################################################################################################################
    try:
        data.insert(5, "Load(kW)", lst_load_all_years)
    except:
        messagebox.showerror("Error", "Error while appending load data")

    ##############################################################o###########################################################
    data_mean = data_raw.resample('M').mean()
    data_sum = data_raw.resample('M').sum()

    time_start_PV_Production = tm.time()
    # inserting delat G in the dataframe
    g0 = ent_G0.get()
    lst_g0 = []
    col_go = data["G(i)"]

    for each_row in col_go:
        lst_g0.append(each_row - float(g0))

    data.insert(6, "Delta G", lst_g0)

    # inserting cell temp in the dataframe
    lst_cellTemp = []
    noct = ent_noct.get()
    noct_Tref = ent_Trif_noct.get()
    delta_NOCT = float(noct) - float(noct_Tref)
    col_Ta = data["T2m"]
    col_deltaG0 = data["Delta G"]

    for i in range(0, len(col_Ta)):
        cell_temp = col_Ta[i] + delta_NOCT / 800 * col_deltaG0[i]
        lst_cellTemp.append(cell_temp)

    data.insert(7, "Cell Temp", lst_cellTemp)

    # inserting DC power column in DF

    if var_SI_unit_nom_pow.get() == 1:
        nominal_pow = float(ent_nominal_power.get())
    elif var_SI_unit_nom_pow.get() == 2:
        nominal_pow = float(ent_nominal_power.get()) * 1000
    elif var_SI_unit_nom_pow.get() == 3:
        nominal_pow = float(ent_nominal_power.get()) * 1000000
    else:
        messagebox.showerror("Error", "Please select the SI unit of nominal power")
        return
    lst_DC_power = []
    eta_dirt = float(ent_etaDirt.get())
    eta_mismatch = float(ent_etaMM.get())
    eta_ref = float(ent_etaRef.get())
    eta_cable = float(ent_etaCable.get())

    gamma_th = float(ent_Gamma.get())
    temp_STC = float(ent_T_STC.get())
    col_cell_temp = data["Cell Temp"]

    for i in range(0, len(col_deltaG0)):
        DC_pow = nominal_pow * col_deltaG0[i] / 1000 * (
                    1 + gamma_th * (col_cell_temp[i] - temp_STC)) * eta_dirt * eta_ref * eta_mismatch * eta_cable
        if DC_pow < 0.001:
            DC_pow = 0
        lst_DC_power.append(DC_pow)

    data.insert(8, "DC Power", lst_DC_power)
    ########################################################################################################################
    # calculation of Pac by formula
    if nominal_pow != 0:
        p_0 = 0.002 * nominal_pow
        k_linear = 0.005
        k_quad = (0.02 / nominal_pow)
        lst_AC_power = [None] * len(lst_DC_power)
        lst_AC_DC_eff = [None] * len(lst_DC_power)
        for i in range(0, len(lst_DC_power)):
            if lst_DC_power[i] != 0:
                b = 1 + k_linear
                a = k_quad
                c = lst_DC_power[i] - p_0
                num_AC_power = -b + ((b ** 2) + 4 * a * c) ** (0.5)
                de_AC_power = 2 * a
                lst_AC_power[i] = num_AC_power / de_AC_power
            else:
                lst_AC_power[i] = 0
        for i in range(0, len(lst_DC_power)):
            if lst_DC_power[i] != 0:
                lst_AC_DC_eff[i] = lst_AC_power[i] / lst_DC_power[i]
            else:
                lst_AC_DC_eff[i] = 0
    else:
        lst_AC_power = [0] * len(lst_DC_power)
        lst_AC_DC_eff = [0] * len(lst_DC_power)

    data.insert(9, "DC/AC effficiency", lst_AC_DC_eff)

    data.insert(10, "Power in ALternate Current(kW)", lst_AC_power)

    time_end_PV_Production = tm.time()
    total_time_PV_Production = round(time_end_PV_Production - time_start_PV_Production, 2)
    ########################################################################################################################
    #     # progress_bar["value"] = 5
    #     # lbl_progress_bar.config(text="Calculating Invertor Efficiency")
    #     # root.update()
    #     time_end_PV_Production = tm.time()
    #     total_time_PV_Production = round(time_end_PV_Production - time_start_PV_Production, 2)
    # #creating dataframe to find invertor efficiency
    #
    # #insertinf Pac% col in invertor DF
    #     time_start_Inv_eff = tm.time()
    #     inv_df =pd.DataFrame()
    #     inv_perc = 0.0
    #     lst_Pac_percentage = []
    #     while inv_perc <120:
    #         lst_Pac_percentage.append(inv_perc)
    #         inv_perc += 0.25
    #
    #     inv_df.insert(0,"Pac%",lst_Pac_percentage)
    #
    # # inserting AC power col in inverter DF
    #     lst_Pac = []
    #     col_Pac_per = inv_df["Pac%"]
    #     for i in range (0, len(col_Pac_per)):
    #         powerAC = nominal_pow*col_Pac_per[i]/100
    #         lst_Pac.append(powerAC)
    #
    #     inv_df.insert(1,"Pac(kW)",lst_Pac)
    #
    #
    #
    # # inserting inverter efficiency in invertor DF
    #
    #     lst_eta_inv = []
    #     col_Pac = inv_df["Pac(kW)"]
    #
    #     if nominal_pow != 0:
    #         noload_loss = 0.7 / 100 * nominal_pow
    #         # noload_loss = float(ent_noloadloss.get())
    #         linear_loss=noload_loss/nominal_pow
    #         # linear_loss = float(ent_linearloss.get())
    #         quadratic_loss = linear_loss/nominal_pow
    #         # quadratic_loss = float(ent_Q_loss.get())
    #     else:
    #         noload_loss = float(ent_noloadloss.get())
    #         linear_loss = float(ent_linearloss.get())
    #         quadratic_loss= float(ent_Q_loss.get())
    #     for i in range(0, len(col_Pac)):
    #         eta_inv = col_Pac[i]/(col_Pac[i]+noload_loss+(linear_loss *col_Pac[i])+(quadratic_loss*(col_Pac[i])**2))*100
    #         lst_eta_inv.append(eta_inv)
    #
    #     inv_df.insert(2,"eta_inv",lst_eta_inv)
    #
    # # inserting DC power in invertor DF
    #     try:
    #         lst_dc_pow = []
    #         col_eta_inv = inv_df["eta_inv"]
    #         for i in range(0, len(col_eta_inv)):
    #             pow_dc_inv = col_Pac[i] / col_eta_inv[i]
    #             lst_dc_pow.append(pow_dc_inv)
    #
    #         inv_df.insert(3, "Pdc", lst_dc_pow)
    #     except:
    #         messagebox.showerror("Error", "Python found a error while Calculating DC power")
    #
    # # inserting Load factor in invertor DF
    #     lst_load_factor =[]
    #     for i in range(0,len(col_Pac)):
    #         load_factor = col_Pac[i]/nominal_pow*100
    #         lst_load_factor.append(load_factor)
    #
    #     inv_df.insert(4,"Load Factor (%)", lst_load_factor)
    #
    # # inserting DC/AC inv efficiency in DF
    #     try:
    #         lst_AC_DC_eff = []
    #         col_DC_pow = data["DC Power"]
    #         # col_loadFactor = inv_df["Load Factor (%)"]
    #         # lst_load_factor2 =[]
    #         # for i in range (0, len(col_loadFactor)):
    #         #     lst_load_factor2.append(col_loadFactor[i])
    #         for i in range(0, len(col_DC_pow)):
    #             if nominal_pow !=0:
    #                 dc_nompow_ratio = (col_DC_pow[i] / nominal_pow)*100
    #             else:
    #                 dc_nompow_ratio=0
    #             y_interp = scipy.interpolate.interp1d(lst_load_factor, lst_eta_inv)
    #             DCAC_eff = y_interp(dc_nompow_ratio)
    #
    #             lst_AC_DC_eff.append((DCAC_eff))
    #
    #         data.insert(9,"DC/AC effficiency", lst_AC_DC_eff)
    #         time_end_Inv_eff = tm.time()
    #         time_inv_eff = round(time_end_Inv_eff -time_start_Inv_eff, 2 )
    #     except:
    #         messagebox.showerror("Error", "Python found a error while Calculating inverter efficiency ")
    #
    # # inserting DC/AC inv efficiency in DF
    # #     progress_bar["value"] = 10
    # #     lbl_progress_bar.config(text="Calculating AC power produced by PV")
    # #     root.update()
    #
    #     lst_AC_pow = []
    #     col_DCAC_eff = data["DC/AC effficiency"]
    #     for i in range(0, len(col_DC_pow)):
    #         AC_pow = col_DC_pow[i]*col_DCAC_eff[i]/100
    #         lst_AC_pow.append(AC_pow)
    #
    #     data.insert(10,"Power in ALternate Current(kW)",lst_AC_pow)
    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    yearly_data_df = data.resample('Y').sum()
    yearly_data_df = yearly_data_df["Power in ALternate Current(kW)"]
    yearly_data_df = yearly_data_df.div(nominal_pow)
    lbl_SI_PV_prod = "(kWh/kW/year)"

    # if var_SI_unit_nom_pow.get()==1:
    #     yearly_data_df=yearly_data_df
    #     lbl_SI_PV_prod = "(kWh)"
    # elif var_SI_unit_nom_pow.get()==2:
    #     yearly_data_df=yearly_data_df.div(1000)
    #     lbl_SI_PV_prod = "(MWh)"
    # elif var_SI_unit_nom_pow.get()==3:
    #     yearly_data_df=yearly_data_df.div(1000000)
    #     lbl_SI_PV_prod = "(GWh)"
    # else:
    #     messagebox.showerror("Error","Please select the SI unit of nominal power")
    #     return
    yearly_data_df = yearly_data_df.round()

    # Calcualtion of wind turbine production
    time_start_WT_production = tm.time()
    try:
        lst_wind_speed_hubHeight = []
        lst_wind_turbine_prod = []
        selected_turbine = var_turbine.get()
        selected_turbine_para = para_WT_df[selected_turbine]
        height_of_rotor = selected_turbine_para[0]
        terrain_rough = float(ent_terrain_roug.get())
        sel_WT_nom_pow = selected_turbine_para[1]
        col_wind_speed = data["WS10m"]
        nb_WT = int(ent_nb_turbines.get())
        ref_height = float(ent_mes_height.get())
        numerator = math.log(height_of_rotor / terrain_rough)
        denominator = math.log(ref_height / terrain_rough)

        int_col_wind_speed2 = power_curve_df.columns.get_loc(selected_turbine) - 1
        lst_power_PC = power_curve_df[selected_turbine].tolist()
        lst_wind_speed_PC = power_curve_df.iloc[:, int_col_wind_speed2].tolist()
        all_WT_df = pd.DataFrame()
        col_time = data["time"]

        lst_time_data = []
        lst_wind_speed = []
        # for o in range (0, len(col_time)):
        #     lst_time_data.append(col_time[o])
        #     lst_wind_speed.append(col_wind_speed[o])
        # lst_time_data = data["time"].tolist
        # all_WT_df.insert(0,"Time",lst_time_data)
        all_WT_df = data[["time"]]
        # all_WT_df.insert(1, "Wind Speed(m/s)", lst_wind_speed)
        for i in range(0, len(col_wind_speed)):
            corrected_wind_speed = col_wind_speed[i] * (numerator / denominator)
            wind_pow_interp = scipy.interpolate.interp1d(lst_wind_speed_PC, lst_power_PC)
            wind_turbine_prod = wind_pow_interp(corrected_wind_speed)
            wind_turbine_prod = wind_turbine_prod * nb_WT
            lst_wind_speed_hubHeight.append(corrected_wind_speed)
            lst_wind_turbine_prod.append(wind_turbine_prod)
        data.insert(11, "Wind speed at Hub height(m/s)", lst_wind_speed_hubHeight)
        data.insert(12, "Wind Turbine Active Power(kW)", lst_wind_turbine_prod)

        # progress_bar["value"] = 15
        # lbl_progress_bar.config(text="Caculating AC power produced by Wind Turbine")
        # root.update()
        time_end_WT_production = tm.time()
        time_WT_production = round(time_end_WT_production - time_start_WT_production, 2)
        # int_WT= 1
        # for j in range (0, int(len(power_curve_df.columns)/2)):
        #     selected_turbine_name_int_WT = para_WT_df.columns[j]
        #     selected_turbine_para_int_WT = para_WT_df[selected_turbine_name_int_WT]
        #     height_of_rotor_int_WT = selected_turbine_para_int_WT[0]
        #     numerator_int_WT = math.log(height_of_rotor_int_WT / terrain_rough)
        #     denominator_int_WT = math.log(ref_height / terrain_rough)
        #     lbl_progress_bar.config(text="Calculating AC power produced by "+str(selected_turbine_name_int_WT))
        #     root.update()
        #     int_col_wind_speed_int_WT = power_curve_df.columns.get_loc(selected_turbine_name_int_WT) - 1
        #     lst_power_PC_int_WT = power_curve_df[selected_turbine_name_int_WT].tolist()
        #     lst_wind_speed_PC_int_WT = power_curve_df.iloc[:, int_col_wind_speed_int_WT].tolist()
        #
        #     lst_wind_turbine_prod_int_WT = []
        #     for t in range (0, len(col_wind_speed)):
        #         corrected_wind_speed_int_WT = col_wind_speed[t] * (numerator_int_WT / denominator_int_WT)
        #         wind_pow_interp_int_WT = scipy.interpolate.interp1d(lst_wind_speed_PC_int_WT, lst_power_PC_int_WT)
        #         wind_turbine_prod_int_WT = wind_pow_interp_int_WT(corrected_wind_speed_int_WT)
        #         wind_turbine_prod_int_WT = wind_turbine_prod_int_WT * nb_WT
        #         lst_wind_turbine_prod_int_WT.append(wind_turbine_prod_int_WT)
        #
        #     all_WT_df.insert(j+1,para_WT_df.columns[j],lst_wind_turbine_prod_int_WT)
        #     int_WT +=2

        # all_WT_df = all_WT_df.resample('Y').sum()

        # all_WT_df = all_WT_df.round()
    except:
        messagebox.showerror("Error", "Python found error while calculating the Wind turbine production")
    ########################################################################################################Commenting on 21-07 ends
    # Battery storage calculations#####################################################################################
    # progress_bar["value"] = 20
    # lbl_progress_bar.config(text="Limitation on Maximum generated power")
    # root.update()
    time_start_batt_bal = tm.time()
    try:
        lst_overall_prod = []
        col_PV_prod = data["Power in ALternate Current(kW)"]
        col_WT_prod = data["Wind Turbine Active Power(kW)"]
        for i in range(0, len(col_PV_prod)):
            overall_prod = col_PV_prod[i] + col_WT_prod[i]
            lst_overall_prod.append(overall_prod)

        data.insert(13, "Overall Production(kW)", lst_overall_prod)

        lst_over_prod_with_PS = []
        col_ovr_prod = data["Overall Production(kW)"]

        if var_peak_shaving.get() == "Limitation on Maximum generated power":
            if var_SI_unit_PS.get() == 1:
                lim_max_gen = float(ent_max_gen.get())
            elif var_SI_unit_PS.get() == 2:
                lim_max_gen = float(ent_max_gen.get()) * 1000
            elif var_SI_unit_PS.get() == 3:
                lim_max_gen = float(ent_max_gen.get()) * 1000000
            else:
                messagebox.showerror("Error", "Error in converting SI unit of peak shaving")
                return

        for i in range(0, len(col_ovr_prod)):

            if var_peak_shaving.get() == "Limitation on Maximum generated power":
                if col_ovr_prod[i] > lim_max_gen:
                    overall_prod_with_PS = lim_max_gen
                else:
                    overall_prod_with_PS = col_ovr_prod[i]
            else:
                overall_prod_with_PS = col_ovr_prod[i]
            # elif var_peak_shaving == "":
            #     overall_prod_with_PS = col_ovr_prod[i]

            lst_over_prod_with_PS.append(overall_prod_with_PS)

        data.insert(14, "Overall production with Peak Shaving(kW)", lst_over_prod_with_PS)

        lst_wastage_PS_gen = []
        col_over_prod_with_PS = data["Overall production with Peak Shaving(kW)"]
        # progress_bar["value"] = 25
        # lbl_progress_bar.config(text="Calculating Overall(PV+WT) production")
        # root.update()
        for i in range(0, len(col_over_prod_with_PS)):
            wastage_PS_gen = col_ovr_prod[i] - col_over_prod_with_PS[i]
            lst_wastage_PS_gen.append(wastage_PS_gen)

        data.insert(15, "Not produced energy due to peak shaving(kW)", lst_wastage_PS_gen)

        lst_balance = []
        # load = float(ent_load.get())
        col_load = data["Load(kW)"]
        col_overall_prod = data["Overall Production(kW)"]
        for i in range(0, len(col_over_prod_with_PS)):
            balance = col_over_prod_with_PS[i] - col_load[i]
            lst_balance.append(balance)

        data.insert(16, "Balance(kW)", lst_balance)
    except:
        messagebox.showerror("Error", "Python found error while calculating the battery storage")
    # Battery df#####################################################################################

    try:
        col_balance = data["Balance(kW)"]
        lst_SOC = []
        lst_pow_discharge = []
        lst_pow_charge = []
        lst_batt_usage_DC_side = []
        lst_batt_usage_AC_side = []
        lst_batt_discharge = []
        lst_batt_charge = []
        lst_balance_with_storage = []
        lst_absorption_frm_grid = []
        lst_injection_in_grid = []
        discharge_eff = float(ent_discharge_eff.get()) / 100
        charge_eff = float(ent_charge_eff.get()) / 100
        min_SOC = float(ent_min_soc.get())
        max_SOC = float(ent_max_soc.get())

        if var_SI_unit_storage_capacity.get() == 1:
            batt_capacity = float(ent_storage_capacity.get())
        if var_SI_unit_storage_capacity.get() == 2:
            batt_capacity = float(ent_storage_capacity.get()) * 1000
        if var_SI_unit_storage_capacity.get() == 3:
            batt_capacity = float(ent_storage_capacity.get()) * 1000000

        for i in range(0, len(col_balance)):
            if i == 0:
                SOC = min_SOC
            else:
                SOC = ((lst_SOC[i - 1] / 100) - (lst_batt_usage_DC_side[i - 1] / batt_capacity)) * 100

            lst_SOC.append(SOC)

            avlb_pow_discharge = ((lst_SOC[i] - min_SOC) / 100) * batt_capacity
            avlb_pow_charge = ((max_SOC - lst_SOC[i]) / 100) * batt_capacity

            if col_balance[i] < 0:
                lst1 = [avlb_pow_discharge, -col_balance[i] / discharge_eff, batt_capacity]
                batt_usage_DC_side = min(lst1)
                batt_usage_AC_side = batt_usage_DC_side * discharge_eff
            else:
                lst1 = [avlb_pow_charge, col_balance[i] * charge_eff, batt_capacity]
                batt_usage_DC_side = - min(lst1)
                batt_usage_AC_side = batt_usage_DC_side / charge_eff
            # progress_bar["value"] = 30
            # lbl_progress_bar.config(text="Calculating SOC% of battery")
            # root.update()
            lst_pow_discharge.append(avlb_pow_discharge)
            lst_pow_charge.append(avlb_pow_charge)
            lst_batt_usage_DC_side.append(batt_usage_DC_side)
            lst_batt_usage_AC_side.append(batt_usage_AC_side)
            # lbl_progress_bar.config(text="Calculating Battery Discharge")
            # root.update()
            if lst_batt_usage_AC_side[i] >= 0:
                batt_discharge = lst_batt_usage_AC_side[i]
            else:
                batt_discharge = 0
            # lbl_progress_bar.config(text="Calculating Battery Charge")
            # root.update()
            if lst_batt_usage_AC_side[i] <= 0:
                batt_charge = -lst_batt_usage_AC_side[i]
            else:
                batt_charge = 0
            # lbl_progress_bar.config(text="Calculating Battery Balance")
            # root.update()
            balance_with_storage = col_balance[i] + lst_batt_usage_AC_side[i]
            if balance_with_storage < 0:
                absortionfrom_grid = - balance_with_storage
            else:
                absortionfrom_grid = 0

            if balance_with_storage >= 0:
                injectioin_grid = balance_with_storage
            else:
                injectioin_grid = 0

            lst_batt_discharge.append(batt_discharge)
            lst_batt_charge.append(batt_charge)
            lst_balance_with_storage.append(balance_with_storage)
            lst_absorption_frm_grid.append(absortionfrom_grid)
            lst_injection_in_grid.append(injectioin_grid)

        # progress_bar["value"] = 35
        # lbl_progress_bar.config(text="Calculation of Grid exchange")
        # root.update()
        batt_df = data[["Power in ALternate Current(kW)", "Wind Turbine Active Power(kW)", "Overall Production(kW)",
                        "Overall production with Peak Shaving(kW)", "Not produced energy due to peak shaving(kW)",
                        "Balance(kW)"]]
        batt_df.insert(6, "State of charge (%)", lst_SOC)
        batt_df.insert(7, "Available power to charge(kW)", lst_pow_charge)
        batt_df.insert(8, "Available power to discharge(kW)", lst_pow_discharge)
        batt_df.insert(9, "Actual battery usage inside battery (DC side) ", lst_batt_usage_DC_side)
        batt_df.insert(10, "Actual battery usage outside battery (AC side) ", lst_batt_usage_AC_side)
        batt_df.insert(11, "Battery Discharge(kW)", lst_batt_discharge)
        batt_df.insert(12, "Battery Charge(kW)", lst_batt_charge)
        batt_df.insert(13, "Balance with Storage (kW)", lst_balance_with_storage)
        # batt_df.insert(14, "Absorption from grid(kW)", lst_absorption_frm_grid)
        # batt_df.insert(15, "Injection in grid(kW)", lst_injection_in_grid)
        ###Inserting Battery df into data df#################################################################################
        data.insert(17, "State of charge (%)", lst_SOC)
        data.insert(18, "Available power to charge(kW)", lst_pow_charge)
        data.insert(19, "Available power to discharge(kW)", lst_pow_discharge)
        data.insert(20, "Actual battery usage inside battery (DC side) ", lst_batt_usage_DC_side)
        data.insert(21, "Actual battery usage outside battery (AC side) ", lst_batt_usage_AC_side)
        data.insert(22, "Battery Discharge(kW)", lst_batt_discharge)
        data.insert(23, "Battery Charge(kW)", lst_batt_charge)
        data.insert(24, "Balance with Storage (kW)", lst_balance_with_storage)
        data.insert(25, "Absorption from grid(kW)", lst_absorption_frm_grid)
        data.insert(26, "Injection in grid(kW)", lst_injection_in_grid)
    except:
        messagebox.showerror("Error", "Python found error while calculating the battery soc")

    time_end_batt_bal = tm.time()
    total_time_batt_bal = round(time_end_batt_bal - time_start_batt_bal, 2)
    ####################################################################################
    try:
        time_start_grid_exchange = tm.time()
        # Limitation on injection into grid
        if var_peak_shaving.get() == "Limitation on Maximum power injection":
            if var_SI_unit_PS.get() == 1:
                lim_max_inj = float(ent_max_inj.get())
            elif var_SI_unit_PS.get() == 2:
                lim_max_inj = float(ent_max_inj.get()) * 1000
            elif var_SI_unit_PS.get() == 3:
                lim_max_inj = float(ent_max_inj.get()) * 1000000
            else:
                messagebox.showerror("Error", "Error in converting SI unit of peak shaving")
                return

        col_inj_grid = data["Injection in grid(kW)"]
        lst_inj_with_PS = []
        for i in range(0, len(col_inj_grid)):
            if var_peak_shaving.get() == "Limitation on Maximum power injection":
                if col_inj_grid[i] <= lim_max_inj:
                    inj_with_PS = col_inj_grid[i]
                else:
                    inj_with_PS = lim_max_inj
            else:
                inj_with_PS = col_inj_grid[i]

            lst_inj_with_PS.append(inj_with_PS)

        data.insert(27, "Injection in grid with Peak shaving(kW)", lst_inj_with_PS)

        col_inj_grid_with_PS = data["Injection in grid with Peak shaving(kW)"]
        lst_wastage_PS_inj = []
        for i in range(0, len(col_inj_grid_with_PS)):
            wastage_PS_inj = col_inj_grid[i] - col_inj_grid_with_PS[i]
            lst_wastage_PS_inj.append(wastage_PS_inj)

        data.insert(28, "Not produced energy due to peak shaving of injection(kW)", lst_wastage_PS_inj)

        lst_overall_prod_after_both_PS = []
        col_waste_PS_gen = data["Not produced energy due to peak shaving(kW)"]
        col_waste_PS_inj = data["Not produced energy due to peak shaving of injection(kW)"]
        for i in range(0, len(col_inj_grid_with_PS)):
            overall_prod_after_both_PS = col_ovr_prod[i] - col_waste_PS_gen[i] - col_waste_PS_inj[i]
            lst_overall_prod_after_both_PS.append(overall_prod_after_both_PS)

        data.insert(29, "PV+WT production after peak shaving (any kind of peak shaving)",
                    lst_overall_prod_after_both_PS)
        col_absorption_frm_grid = data["Absorption from grid(kW)"]
        lst_grid_exchange = []
        for i in range(0, len(col_absorption_frm_grid)):
            if col_absorption_frm_grid[i] == 0:
                grid_exchange = col_inj_grid_with_PS[i]
            else:
                grid_exchange = col_absorption_frm_grid[i]

            lst_grid_exchange.append(grid_exchange)

        data.insert(30, "Grid Exchange(kW)", lst_grid_exchange)
        # progress_bar["value"] = 40
        # lbl_progress_bar.config(text="Calculating Self Sufficiency")
        # root.update()
        time_end_grid_exchange = tm.time()
        total_time_grid_exchange = round(time_end_grid_exchange - time_start_grid_exchange, 2)
        try:
            time_start_SS = tm.time()
            lst_self_sufficiency = []

            for i in range(0, len(col_absorption_frm_grid)):
                if col_absorption_frm_grid[i] == 0:
                    self_sufficiency = col_load[i]
                else:
                    self_sufficiency = col_load[i] - col_absorption_frm_grid[i]

                lst_self_sufficiency.append(self_sufficiency)

            data.insert(31, "Self Suffiency", lst_self_sufficiency)
            time_end_SS = tm.time()
            total_time_SS = round(time_end_SS - time_start_SS, 2)
        except:
            messagebox.showerror("Error", "Error while calculating self sufficiency")
    except:
        messagebox.showerror("Error", "Error while calculating injection into grid with peak shaving")

    ##Ripartition of peak shaving between PV and WT##################################################################################
    try:
        time_start_peak_shaving = tm.time()
        lst_quota_PV = []
        for i in range(0, len(col_ovr_prod)):
            if col_ovr_prod[i] == 0:
                quota_PV = 0
            else:
                quota_PV = (col_PV_prod[i] / col_ovr_prod[i]) * 100

            lst_quota_PV.append(quota_PV)

        data.insert(32, "Quota of production of PV with respect to (PV+WT) [%]", lst_quota_PV)

        lst_quota_WT = []
        col_overall_prod_after_both_PS = data["PV+WT production after peak shaving (any kind of peak shaving)"]

        for i in range(0, len(col_overall_prod_after_both_PS)):
            if col_overall_prod_after_both_PS[i] == 0:
                quota_WT = 0
            else:
                quota_WT = (col_WT_prod[i] / col_ovr_prod[i]) * 100

            lst_quota_WT.append(quota_WT)

        data.insert(33, "Quota of production of WT with respect to (PV+WT) [%]", lst_quota_WT)

        load = float(ent_load.get())
        lst_PV_prod_after_PS = []
        lst_WT_prod_after_PS = []
        lst_inj_grid_PV = []
        lst_inj_grid_WT = []
        lst_self_consumption_PV = []
        lst_self_consumption_WT = []
        lst_load = []

        col_quota_PV = data["Quota of production of PV with respect to (PV+WT) [%]"]
        col_quota_WT = data["Quota of production of WT with respect to (PV+WT) [%]"]
        col_self_suffiency = data["Self Suffiency"]

        for i in range(0, len(col_quota_PV)):
            PV_prod_after_PS = (col_quota_PV[i] / 100) * col_overall_prod_after_both_PS[i]
            PV_prod_after_WT = (col_quota_WT[i] / 100) * col_overall_prod_after_both_PS[i]
            inj_grid_PV = (col_quota_PV[i] / 100) * col_inj_grid_with_PS[i]
            inj_grid_WT = (col_quota_WT[i] / 100) * col_inj_grid_with_PS[i]
            self_consumption_PV = (col_quota_PV[i] / 100) * col_self_suffiency[i]
            self_consumption_WT = (col_quota_WT[i] / 100) * col_self_suffiency[i]

            lst_PV_prod_after_PS.append(PV_prod_after_PS)
            lst_WT_prod_after_PS.append(PV_prod_after_WT)
            lst_inj_grid_PV.append(inj_grid_PV)
            lst_inj_grid_WT.append(inj_grid_WT)
            lst_self_consumption_PV.append(self_consumption_PV)
            lst_self_consumption_WT.append(self_consumption_WT)
            lst_load.append(load)

        data.insert(34, "Production of PV after peak shaving(kW)", lst_PV_prod_after_PS)
        data.insert(35, "Production of WT after peak shaving(kW)", lst_WT_prod_after_PS)
        data.insert(36, "Injection in the grid from PV(kW)", lst_inj_grid_PV)
        data.insert(37, "Injection in the grid from WT(kW)", lst_inj_grid_WT)
        data.insert(38, "Self-consumtion for PV(kW)", lst_self_consumption_PV)
        data.insert(39, "Self-consumtion for WT(kW)", lst_self_consumption_WT)
        # data.insert(39,"Load(kW)", lst_load)
        # progress_bar["value"] = 45
        # lbl_progress_bar.config(text="Calculating Self Consumption")
        # root.update()
        time_end_peak_shaving = tm.time()
        total_time_peak_shaving = round(time_end_peak_shaving - time_start_peak_shaving, 2)
    except:
        messagebox.showerror("Error", "Error while calculating Ripartition of peak shaving between PV and WT")

    ####################################################################################
    # Display Montly data
    frame_MD = LabelFrame(tab_PV_results, text="Show Monthly Data", padx=10, pady=10)
    frame_MD.place(x=10, y=10, height=150, width=400)
    # frame_MD.place(row=7, column=0, pady=10, columnspan=2)
    btn_irrdata = Button(frame_MD, text="Irr Data", width=10,
                         command=lambda: display_irradiation_heatmap(data_sum, total_years, int(start_year)))
    # btn_irrdata = Button(frame_MD, text="Show Irr Data", command=lambda : display_heat_map(data_sum, total_years, int(start_year), "G(i)"))
    btn_irrdata.place(x=10, y=10)
    btn_tempdata = Button(frame_MD, text="Temp Data", width=10,
                          command=lambda: display_temp_heatmap(data_mean, total_years, int(start_year)))
    # btn_tempdata = Button(frame_MD, text="Show Temp Data",command=lambda: display_heat_map(data_mean, total_years, int(start_year), "T2m"))
    btn_tempdata.place(x=130, y=10)
    btn_winddata = Button(frame_MD, text="Wind Data", width=10,
                          command=lambda: display_wind_heatmap(data_mean, total_years, int(start_year)))
    # btn_winddata = Button(frame_MD, text="Show Wind Data",command=lambda: display_heat_map(data_mean, total_years, int(start_year), "WS10m"))
    btn_winddata.place(x=250, y=10)

    var_SI_unit = IntVar()
    btn_PV_data = Button(frame_MD, text="PV+WT production", width=15,
                         command=lambda: show_PV_barGraph_monthly(data, total_years, start_year, var_SI_unit.get()))
    btn_PV_data.place(x=30, y=50)
    btn_WT_data = Button(frame_MD, text="Exchange with Grid", width=15,
                         command=lambda: show_grid_ex_barGraph_monthly(data, total_years, start_year,
                                                                       var_SI_unit.get()))
    btn_WT_data.place(x=160, y=50)
    rad_SI_unit_kw = Radiobutton(frame_MD, text="kWh/month", variable=var_SI_unit, value=1)
    rad_SI_unit_kw.place(x=10, y=90)
    rad_SI_unit_Mw = Radiobutton(frame_MD, text="MWh/month", variable=var_SI_unit, value=2)
    rad_SI_unit_Mw.place(x=120, y=90)
    rad_SI_unit_Gw = Radiobutton(frame_MD, text="GWh/month", variable=var_SI_unit, value=3)
    rad_SI_unit_Gw.place(x=230, y=90)

    # btn_graph_eff = Button(frame_para, text= "Graph", command= lambda: plot_graph(lst_load_factor,lst_eta_inv, "Load Factor (%) = Pac/Pac-nom","Efficiency of the DC/AC converter (%)","") )
    # # btn_graph_eff.grid(row=5, column=0,pady=10, columnspan= 4)
    # btn_graph_eff.place(x=50, y=111)

    frame_all_data = LabelFrame(tab_PV_results, text="All Data", padx=10, pady=10)
    # frame_all_data.grid(row=6, column=1, pady=10, columnspan=2, rowspan=1)
    # frame_all_data.grid(row=0, column=6, columnspan=2,  sticky=tk.NW)
    # frame_all_data.place(x=500, y=500, height=75, width= 200)
    btn_show_data = Button(frame_all_data, text="Show Data", command=lambda: show_data(data))
    # btn_show_data.place(x=10, y=5)

    btn_export_excel = Button(tab_Allresults, text="Export", command=lambda: export_to_excel(data))
    btn_export_excel.place(x=0, y=0)

    frame_batt_data = LabelFrame(tab_PV_results, text="Battery Data", padx=10, pady=10)
    frame_batt_data.grid(row=1, column=6, sticky=tk.NW)
    # progress_bar["value"] = 50
    # lbl_progress_bar.config(text="Creating Daily graphs")
    # root.update()
    # btn_show_batt_data = Button(frame_batt_data, text="Show Battery Data", command=lambda: show_data(batt_df))
    # btn_show_batt_data = Button(frame_batt_data, text="Show Battery Data", command=lambda: show_data_new_win(batt_df,tab_storage))
    # btn_show_batt_data.grid(row=0, column=0)

    frame_daily_data = LabelFrame(tab_PV_results, text="Daily Trend", padx=5, pady=5)
    # frame_daily_data.grid(row=0, column=3, pady=10, columnspan=2, rowspan=6, sticky=tk.NW)
    frame_daily_data.place(x=600, y=10, height=150, width=620)
    startingday = date(int(start_year), 1, 1)
    endingday = date(int(end_year), 12, 31)
    cal = DateEntry(frame_daily_data, selectmode="day", year=int(start_year), month=1, day=1)
    cal.place(x=10, y=30)

    btn_daily_PV_prod = Button(frame_daily_data, text="Production", width=15,
                               command=lambda: show_daily_trend(cal.get_date(), data, 'PV production'))
    btn_daily_PV_prod.place(x=360, y=10)

    btn_daily_irr_values = Button(frame_daily_data, text="Irradiance", width=15,
                                  command=lambda: show_daily_trend(cal.get_date(), data, 'Irradiance'))
    btn_daily_irr_values.place(x=120, y=10)

    btn_daily_temp_values = Button(frame_daily_data, text="Cell Temperature", width=15,
                                   command=lambda: show_daily_trend(cal.get_date(), data, 'Cell Temperature'))
    btn_daily_temp_values.place(x=240, y=10)

    btn_daily_grid_exchange = Button(frame_daily_data, text="Grid Exchange", width=15,
                                     command=lambda: show_daily_trend(cal.get_date(), data, 'Grid Exchange'))
    btn_daily_grid_exchange.place(x=480, y=10)

    btn_monthly_trend = Button(frame_daily_data, text="Daily Production", width=15,
                               command=lambda: show_daily_trend_ofmonth(cal.get_date(), data))
    btn_monthly_trend.place(x=120, y=40)

    btn_monthly_grid_ex = Button(frame_daily_data, text="Daily Grid Exchange",
                                 command=lambda: show_grid_exchange_ofmonth(cal.get_date(), data))
    btn_monthly_grid_ex.place(x=360, y=40)

    btn_supply_load = Button(frame_daily_data, text="Load Supply", width=15,
                             command=lambda: plot_pieChart(cal.get_date(), data, "Load Supply"))
    btn_supply_load.place(x=240, y=40)

    btn_energy_gen = Button(frame_daily_data, text="Energy Generation", width=15,
                            command=lambda: plot_pieChart(cal.get_date(), data, "Energy Generation"))
    btn_energy_gen.place(x=480, y=40)

    # appending yearly PV production in PV results tab ####################################################
    # lbl_year =Label(frame_yearly_data, text="Year",font=("Helvatical bold", 8))
    # lbl_year.grid(row=0, column=0)
    # lbl_Value = Label(frame_yearly_data, text="Specific PV Production "+ lbl_SI_PV_prod,font=("Helvatical bold", 8))
    # lbl_Value.grid(row=0, column=1)
    # for i in range(0, len(yearly_data_df)):
    #     str_year = int(start_year)+i
    #     str_value = yearly_data_df[i]
    #
    #     row_pos = i +1
    #     Label(frame_yearly_data, text=str_year).grid(row=row_pos, column=0)
    #     Label(frame_yearly_data, text=str_value).grid(row=row_pos, column=1)
    ########################################################################################################
    show_data_new_win_XY(data, tab_Allresults, 600, 1200, 10, 30)  # inserting battery data in battery storage tab

    ########################################################################################################Commenting on 21-07
    # WT_id = 0
    # for each_col in all_WT_df.columns:
    #     selected_turbine_name_int_WT2 = para_WT_df.columns[WT_id]
    #     selected_turbine_nom_pow = para_WT_df[selected_turbine_name_int_WT2][1]
    #     all_WT_df[selected_turbine_name_int_WT2] = all_WT_df[selected_turbine_name_int_WT2].div(selected_turbine_nom_pow)
    #     WT_id +=1
    #
    # all_WT_df = all_WT_df.div(nb_WT)
    # float_WT_ageing = float(ent_ageing_WT.get())/100
    # lst_25_yrs = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25]
    # all_WT_25yrs_df = replicate_data_for_25yrs(total_years+1,all_WT_df )
    # production_with_ageing(all_WT_25yrs_df,float_WT_ageing)
    # all_WT_25yrs_df.insert(0,"Years", lst_25_yrs)
    # show_data_new_win(all_WT_25yrs_df,tab_WT_results,500,800)#inserting Wind turbine data in Wind turbine tab
    ########################################################################################################Commenting on 21-07 ends
    # appending yearly WT production in WT results tab#########################################################
    #     frame_all_wt = LabelFrame(tab_WT_results, text="All WT results")
    #     frame_all_wt.grid(row=0, column=0, sticky=tk.NW)
    #
    #     lbl_year_WT = Label(frame_all_wt, text="Year", font=("Helvatical bold", 8))
    #     lbl_year_WT.grid(row=0, column=0)
    #
    #
    #     for i in range(0, len(all_WT_df)):
    #         row_pos = i + 1
    #         str_year = int(start_year) + i
    #         Label(frame_all_wt, text=str_year).grid(row=row_pos, column=0)
    #
    #     try:
    #         for int_WT in range(0, len(all_WT_df.columns)-1):
    #             Wt_prod = all_WT_df.iloc[:, int_WT ]
    #             WT_name = all_WT_df.columns[int_WT ]
    #             col_pos = int_WT + 1
    #             lbl_Value_WT = Label(frame_all_wt, text=WT_name + "" + lbl_SI_PV_prod, font=("Helvatical bold", 8))
    #             lbl_Value_WT.grid(row=0, column=col_pos)
    #             selected_turbine_name_int_WT2 = para_WT_df.columns[int_WT]
    #             selected_turbine_para_int_WT2 = para_WT_df[selected_turbine_name_int_WT2]
    #             nominal_pow_WT = selected_turbine_para_int_WT2[1]
    #
    #             for i in range(0, len(Wt_prod)):
    #                 # str_year = int(start_year)+i
    #                 str_value = Wt_prod[i]/nominal_pow_WT
    #                 row_pos = i + 1
    #                 # Label(frame_all_wt, text=str_year).grid(row=row_pos, column=0)
    #                 Label(frame_all_wt, text=str_value).grid(row=row_pos, column=col_pos)
    #     except:
    #         messagebox.showerror("Error","Python found error while appending the Wind turbine production")
    #     progress_bar["value"] = 55
    #     lbl_progress_bar.config(text="Calculating Results")
    #     root.update()
    try:
        results_df = produce_results_df(data, sel_WT_nom_pow)
        show_data_new_win(results_df, tab_results, 600, 1200)

        # frm_SI_results = LabelFrame(tab_results, text="SI Units")
        # frm_SI_results.place(row=9, column=9)
        # lbl_SI_pow = Label(frm_SI_results,text="Unit of measurement for power")
        # lbl_SI_pow.place(row=0, column=0)
    except:
        messagebox.showerror("error", "error while producing results df")
    #
    # update_results_in_main_wind(data, results_df,nominal_pow,)
    ###########################################################################################################

    # updating values of results DF in main window
    global lst_result_values
    lst_result_values = []

    all_data_sum = data.resample('Y').sum()
    # col_pv_prod_without_ps = all_data_sum[""]
    col_pv_prod_with_ps = all_data_sum['Production of PV after peak shaving(kW)'].div(1000)
    # col_WT_prod_without_ps= all_data_sum[""]
    col_WT_prod_without_ps = all_data_sum["Power in ALternate Current(kW)"].div(1000)
    col_WT_prod_with_ps = all_data_sum["Production of WT after peak shaving(kW)"].div(1000)
    results_col = results_df[str(start_year)]

    if nominal_pow != 0:
        productivity_pv_with_ps = (col_pv_prod_with_ps[0] / nominal_pow) * 1000
        productivity_pv_without_ps = (col_WT_prod_without_ps[0] / nominal_pow) * 1000
    else:
        productivity_pv_without_ps = 0
        productivity_pv_with_ps = 0

    if sel_WT_nom_pow != 0:
        productivity_WT_with_ps = (col_WT_prod_with_ps[0] / sel_WT_nom_pow) * 1000
    else:
        productivity_WT_with_ps = 0
    ent_pv_prodcutivity.configure(state="normal")
    ent_pv_prodcutivity.delete(0, END)
    ent_pv_prodcutivity.insert(END, str(round(productivity_pv_without_ps)))
    # lst_result_values.append(results_col[8])

    ent_pv_annual_prod.configure(state="normal")
    ent_pv_annual_prod.delete(0, END)
    ent_pv_annual_prod.insert(END, str(float(results_col[8]) * nominal_pow / 1000))
    lst_result_values.append(float(results_col[8]) * nominal_pow / 1000)

    ent_pv_productivity_with_ps.configure(state="normal")
    ent_pv_productivity_with_ps.delete(0, END)
    ent_pv_productivity_with_ps.insert(END, str(round(productivity_pv_with_ps)))
    # lst_result_values.append(round(productivity_pv_with_ps))

    ent_Annual_pv_prod_with_ps.configure(state="normal")
    ent_Annual_pv_prod_with_ps.delete(0, END)
    ent_Annual_pv_prod_with_ps.insert(END, str(round(col_pv_prod_with_ps[0])))
    lst_result_values.append(round(col_pv_prod_with_ps[0]))

    ent_WT_prodcutivity.configure(state="normal")
    ent_WT_prodcutivity.delete(0, END)
    ent_WT_prodcutivity.insert(END, str(results_col[10]))
    # lst_result_values.append(results_col[10])

    ent_WT_annual_prod.configure(state="normal")
    ent_WT_annual_prod.delete(0, END)
    ent_WT_annual_prod.insert(END, str(float(results_col[10]) * nb_WT * sel_WT_nom_pow / 1000))
    lst_result_values.append(float(results_col[10]) * nb_WT * sel_WT_nom_pow / 1000)

    ent_WT_productivity_with_ps.configure(state="normal")
    ent_WT_productivity_with_ps.delete(0, END)
    if nb_WT != 0:
        ent_WT_productivity_with_ps.insert(END, str(round(productivity_WT_with_ps / nb_WT)))
        # lst_result_values.append(round(productivity_WT_with_ps/nb_WT))
    else:
        ent_WT_productivity_with_ps.insert(END, str(0))
        # lst_result_values.append(0)

    ent_Annual_WT_prod_with_ps.configure(state="normal")
    ent_Annual_WT_prod_with_ps.delete(0, END)
    ent_Annual_WT_prod_with_ps.insert(END, str(round(col_WT_prod_with_ps[0])))
    lst_result_values.append(round(col_WT_prod_with_ps[0]))

    ent_dischrged_energy.configure(state="normal")
    ent_dischrged_energy.delete(0, END)
    ent_dischrged_energy.insert(END, str(results_col[20]))
    lst_result_values.append(results_col[20])

    ent_chrged_energy.configure(state="normal")
    ent_chrged_energy.delete(0, END)
    ent_chrged_energy.insert(END, str(results_col[19]))
    lst_result_values.append(results_col[19])

    ent_self_consumption.configure(state="normal")
    ent_self_consumption.delete(0, END)
    ent_self_consumption.insert(END, str(results_col[22]))
    # lst_result_values.append(results_col[22])

    ent_self_sufficiency.configure(state="normal")
    ent_self_sufficiency.delete(0, END)
    ent_self_sufficiency.insert(END, str(results_col[23]))
    # lst_result_values.append(results_col[23])

    ent_absorption_from_grid.configure(state="normal")
    ent_absorption_from_grid.delete(0, END)
    ent_absorption_from_grid.insert(END, str(results_col[24]))
    # lst_result_values.append(results_col[24])

    ent_injection_to_grid.configure(state="normal")
    ent_injection_to_grid.delete(0, END)
    ent_injection_to_grid.insert(END, str(results_col[25]))
    # lst_result_values.append(results_col[25])

    ent_annual_load.configure(state="normal")
    ent_annual_load.delete(0, END)
    ent_annual_load.insert(END, str(results_col[14]))
    lst_result_values.append(results_col[14])

    var_results_SI_units.set(lst_SI_units[1])
    #######################################################################################################
    show_monthly_energy_flows(cal.get_date(), data)
    show_daily_energyflow_and_bal(cal.get_date(), data)

    btn_energy_flow = Button(frame_daily_data, text="Update Table",
                             command=lambda: display_energy_flow(cal.get_date(), data))
    btn_energy_flow.place(x=330, y=90)
    # progress_bar["value"] = 60
    # lbl_progress_bar.config(text="Replicating data for 25 years")
    # root.update()

    frm_show_daily_data = LabelFrame(tab_PV_results, height=70, width=200, text="Daily data")
    frm_show_daily_data.place(x=935, y=500)
    btn_show_daily_data = Button(frm_show_daily_data, text="Show data",
                                 command=lambda: show_daily_data(cal.get_date(), data))
    btn_show_daily_data.place(x=10, y=10)
    btn_export_daily_data = Button(frm_show_daily_data, text="Export data",
                                   command=lambda: export_daily_data(cal.get_date(), data))
    btn_export_daily_data.place(x=100, y=10)
    ########################################################################################################
    # lbl_avg_power_needed = Label(tab_inputs, text=  )
    ########################################################################################################
    time_start_replicate_data = tm.time()
    data_sum_fr_25 = data.resample("Y").sum()
    yrs_df = replicate_data_for_25yrs(total_years + 1, data_sum_fr_25)
    time_end_replicate_data = tm.time()
    total_time_replicate_data = round(time_end_replicate_data - time_start_replicate_data, 2)
    ###########################################################################################################
    time_start_NPV_calculation = tm.time()
    # creation of NPV table
    npv_df = pd.DataFrame()
    # elec_sp_1 = float(ent_ele_sp.get())
    # elec_sp_2 = float(ent_ele_sp2.get())
    ageing = float(ent_ageing.get()) / 100
    ageing_WT = float(ent_ageing_WT.get()) / 100
    # elec_sp_1_WT = float(ent_ele_sp_WT.get())
    # elec_sp_2_WT = float(ent_ele_sp2_WT.get())
    price_self_consumed_energy = float(ent_value_SC_energy.get())
    lst_NPV_PV_prod = yrs_df["Production of PV after peak shaving(kW)"].to_list()
    lst_NPV_PV_annual_inj = yrs_df["Injection in the grid from PV(kW)"].to_list()
    lst_NPV_self_consumption_PV = yrs_df["Self-consumtion for PV(kW)"].to_list()
    lst_NPV_self_consumption_WT = yrs_df["Self-consumtion for WT(kW)"].to_list()
    lst_NPV_PV_prod_age = []
    lst_NPV_PV_annual_inj_age = []
    lst_NPV_self_consumption_PV_age = []
    lst_price_elec_nt_purchased = []
    lst_grid_inj_price = []
    lst_selling_price = []
    lst_selling_price_WT = []
    lst_grid_inj_price_WT = []
    lst_price_self_consumption_PV = []
    lst_price_self_consumption_WT = []
    lst_NPV_years = []
    lst_grid_inj_price.append(0)
    lst_NPV_PV_prod_age.append(0)
    lst_NPV_PV_annual_inj_age.append(0)
    lst_selling_price.append(0)
    lst_price_self_consumption_PV.append(0)
    lst_NPV_self_consumption_PV_age.append(0)
    lst_price_elec_nt_purchased.append(0)
    lst_grid_inj_price_WT.append(0)
    lst_selling_price_WT.append(0)
    lst_price_self_consumption_WT.append(0)

    lst_NPV_WT_prod = yrs_df["Production of WT after peak shaving(kW)"].to_list()
    lst_NPV_WT_annual_inj = yrs_df["Injection in the grid from WT(kW)"].to_list()

    lst_NPV_self_consumption_WT = yrs_df["Self-consumtion for WT(kW)"].to_list()
    lst_NPV_WT_prod_age = []
    lst_NPV_WT_annual_inj_age = []
    lst_NPV_self_consumption_WT_age = []

    lst_NPV_ann_inj = yrs_df["Injection in grid with Peak shaving(kW)"].to_list()
    lst_NPV_ann_abs = yrs_df["Absorption from grid(kW)"].to_list()

    lst_NPV_WT_prod_age.append(0)
    lst_NPV_WT_annual_inj_age.append(0)
    lst_NPV_self_consumption_WT_age.append(0)

    if (lst_NPV_ann_inj[0] - lst_NPV_ann_abs[0]) > 0:
        annual_surplus = (lst_NPV_ann_inj[0] - lst_NPV_ann_abs[0]) / 1000
    else:
        annual_surplus = 0

    # To find battery lifetime
    lst_negetive_cash_flow_batt = []
    lst_tax_red_st = []

    batt_discharge_NPV = yrs_df["Battery Discharge(kW)"][0]
    nb_of_cycles_by_batt = batt_discharge_NPV / batt_capacity
    if nb_of_cycles_by_batt != 0:
        theoritical_lifetime = 10000 / nb_of_cycles_by_batt
    else:
        theoritical_lifetime = 1000000

    max_lifetime = min(10, theoritical_lifetime)
    validity_of_batt = int(ent_val_ST.get())
    investment_cost_batt = float(ent_ins_cost_ST.get())

    ##########################################################################################################################
    entry_set_and_disable(ent_size, var_ent_size, nominal_pow / 1000)
    entry_set_and_disable(ent_prod, var_ent_prod, round(lst_NPV_PV_prod[0] / 1000))
    entry_set_and_disable(ent_grid_inj, var_ent_grid_inj, lst_NPV_PV_annual_inj[0] / 1000)
    entry_set_and_disable(ent_life, var_ent_life, 25)
    entry_set_and_disable(ent_spec_prod, var_ent_spec_prod, results_col[8])

    entry_set_and_disable(ent_size_WT, var_ent_size_WT, sel_WT_nom_pow * nb_WT / 1000)
    entry_set_and_disable(ent_prod_WT, var_ent_prod_WT, round(lst_NPV_WT_prod[0] / 1000))
    entry_set_and_disable(ent_grid_inj_WT, var_ent_grid_inj_WT, lst_NPV_WT_annual_inj[0] / 1000)
    entry_set_and_disable(ent_life_WT, var_ent_life_WT, 25)
    entry_set_and_disable(ent_spec_prod_WT, var_ent_spec_prod_WT, ent_WT_productivity_with_ps.get())

    entry_set_and_disable(ent_size_ST, var_ent_size_ST, batt_capacity / 1000)
    entry_set_and_disable(ent_SC_energy_PV, var_ent_SC_energy_PV, round(lst_NPV_self_consumption_PV[0] / 1000))
    entry_set_and_disable(ent_SC_energy_WT, var_ent_SC_energy_WT, round(lst_NPV_self_consumption_WT[0] / 1000))

    entry_set_and_disable(ent_annual_surplus, var_ent_annual_surplus, annual_surplus)
    entry_set_and_disable(ent_NM_energy, var_ent_NM_energy, min(lst_NPV_ann_inj[0], lst_NPV_ann_abs[0]) / 1000)
    #######################################################################################################################
    for i in range(0, 26):
        # lst_NPV_years.append(int(start_year)+i)
        lst_NPV_years.append(i)
    for i in range(0, 25):
        pv_prod_after_ageing = lst_NPV_PV_prod[i] * (1 - ageing * i) / 1000
        lst_NPV_PV_prod_age.append(pv_prod_after_ageing)
        NPV_PV_annual_inj_age = lst_NPV_PV_annual_inj[i] * (1 - ageing * i) / 1000
        lst_NPV_PV_annual_inj_age.append(NPV_PV_annual_inj_age)
        Wt_prod_after_ageing = lst_NPV_WT_prod[i] * (1 - ageing_WT * i) / 1000
        lst_NPV_WT_prod_age.append(Wt_prod_after_ageing)
        NPV_WT_annual_inj_age = lst_NPV_WT_annual_inj[i] * (1 - ageing_WT * i) / 1000
        lst_NPV_WT_annual_inj_age.append(NPV_WT_annual_inj_age)

    # for i in range(0, 25):
    #     if i < 5:
    #         lst_selling_price.append(elec_sp_1)
    #         lst_selling_price_WT.append(elec_sp_1_WT)
    #     else:
    #         lst_selling_price.append(elec_sp_2)
    #         lst_selling_price_WT.append(elec_sp_2_WT)

    for i in range(0, 25):
        lst_selling_price.append(float(lst_var_sp_pv[i]))
        lst_selling_price_WT.append(float(lst_var_sp_wt[i]))
    # lst_selling_price = lst_var_sp_pv
    # lst_selling_price_WT = lst_var_sp_wt

    for i in range(0, 25):
        if var_dd_SC.get() == "Yes":
            grid_inj_price = lst_NPV_PV_annual_inj_age[i] * lst_selling_price[i] * 1000
            grid_inj_price_WT = lst_NPV_WT_annual_inj_age[i] * lst_selling_price_WT[i] * 1000
        else:
            grid_inj_price = lst_NPV_PV_prod_age[i] * lst_selling_price[i] * 1000
            grid_inj_price_WT = lst_NPV_WT_prod_age[i] * lst_selling_price_WT[i] * 1000

        lst_grid_inj_price.append(grid_inj_price)
        lst_grid_inj_price_WT.append(grid_inj_price_WT)

    npv_df.insert(0, "Years", lst_NPV_years)
    npv_df.insert(1, "Grid injection of PV($/y)", lst_grid_inj_price)
    npv_df.insert(2, "PV production (MWh/y)", lst_NPV_PV_prod_age)
    npv_df.insert(3, "PV annual injection (MWh/y)", lst_NPV_PV_annual_inj_age)
    npv_df.insert(4, "Price of sold  electricity from PV ($/kWh)", lst_selling_price)

    for i in range(0, 25):
        lst_price_elec_nt_purchased.append(price_self_consumed_energy)
        if var_dd_SC.get() == "Yes":
            NPV_self_consumption_PV_age = lst_NPV_self_consumption_PV[i] * (1 - ageing * i) / 1000
            lst_NPV_self_consumption_PV_age.append(NPV_self_consumption_PV_age)
            NPV_self_consumption_WT_age = lst_NPV_self_consumption_WT[i] * (1 - ageing_WT * i) / 1000
            lst_NPV_self_consumption_WT_age.append(NPV_self_consumption_WT_age)
        else:
            lst_NPV_self_consumption_PV_age.append(0)
            lst_NPV_self_consumption_WT_age.append(0)

        total_price_self_consumption_PV = lst_price_elec_nt_purchased[i + 1] * lst_NPV_self_consumption_PV_age[
            i + 1] * 1000
        lst_price_self_consumption_PV.append(total_price_self_consumption_PV)
        total_price_self_consumption_WT = lst_price_elec_nt_purchased[i + 1] * lst_NPV_self_consumption_WT_age[
            i + 1] * 1000
        lst_price_self_consumption_WT.append(total_price_self_consumption_WT)

    npv_df.insert(5, "Self Consumption of PV($/y)", lst_price_self_consumption_PV)
    npv_df.insert(6, "PV annual self consumption (MWh/y)", lst_NPV_self_consumption_PV_age)
    npv_df.insert(7, "Price of  electricity not purchased from the grid due to PV self consumption ($/kWh)",
                  lst_price_elec_nt_purchased)

    no_of_years_tax_PV = float(ent_val_PV.get())
    no_of_years_tax_WT = float(ent_val_WT.get())

    pv_size = float(ent_size.get()) * 1000
    wt_size = float(ent_size_WT.get()) * 1000

    tax_reduction_pv = float(ent_tax_red_PV.get())
    tax_reduction_wt = float(ent_tax_red_WT.get())

    investment_cost_PV = float(ent_ins_cost.get())
    investment_cost_WT = float(ent_ins_cost_WT.get())
    investment_cost_ST = float(ent_ins_cost_ST.get())

    OandM_cost_PV = float(ent_om.get())
    OandM_cost_WT = float(ent_om_WT.get())

    total_inv_cost_PV = investment_cost_PV * pv_size
    total_inv_cost_WT = investment_cost_WT * wt_size
    total_inv_cost_ST = investment_cost_ST * batt_capacity

    lbl_tax_red_PV_cost.configure(text="$" + str(total_inv_cost_PV))
    lbl_tax_red_WT_cost.configure(text="$" + str(total_inv_cost_WT))
    lbl_tax_red_ST_cost.configure(text="$" + str(total_inv_cost_ST))

    lst_total_tax_reduction_PV = []
    lst_total_positive_cash_flow_PV = []
    lst_total_negetive_cash_flow_PV = []
    lst_total_tax_reduction_WT = []
    lst_total_positive_cash_flow_WT = []
    lst_total_negetive_cash_flow_WT = []

    # total_negetive_cash_flow_PV = -total_inv_cost_PV
    # total_negetive_cash_flow_WT = - total_inv_cost_WT

    lst_total_tax_reduction_PV.append(0)
    lst_total_positive_cash_flow_PV.append(0)
    lst_total_tax_reduction_WT.append(0)
    lst_total_positive_cash_flow_WT.append(0)
    lst_total_negetive_cash_flow_PV.append(-total_inv_cost_PV)
    lst_total_negetive_cash_flow_WT.append(- total_inv_cost_WT)

    for i in range(0, 25):
        if i < no_of_years_tax_PV:
            total_tax_reduction_PV = tax_reduction_pv * total_inv_cost_PV / 100
        else:
            total_tax_reduction_PV = 0

        if i < no_of_years_tax_WT:
            total_tax_reduction_WT = tax_reduction_wt * total_inv_cost_WT / 100
        else:
            total_tax_reduction_WT = 0

        lst_total_tax_reduction_PV.append(total_tax_reduction_PV)
        total_positive_cash_flow_PV = lst_total_tax_reduction_PV[i + 1] + lst_price_self_consumption_PV[i + 1] + \
                                      lst_grid_inj_price[i + 1]
        lst_total_positive_cash_flow_PV.append(total_positive_cash_flow_PV)

        lst_total_tax_reduction_WT.append(total_tax_reduction_WT)
        total_positive_cash_flow_WT = lst_total_tax_reduction_WT[i + 1] + lst_price_self_consumption_WT[i + 1] + \
                                      lst_grid_inj_price_WT[i + 1]
        lst_total_positive_cash_flow_WT.append(total_positive_cash_flow_WT)

        total_negetive_cash_flow_PV = -OandM_cost_PV * pv_size
        # total_negetive_cash_flow_WT= - OandM_cost_WT*wt_size
        total_negetive_cash_flow_WT = - OandM_cost_WT * lst_NPV_WT_prod_age[1] * 1000

        lst_total_negetive_cash_flow_PV.append(total_negetive_cash_flow_PV)
        lst_total_negetive_cash_flow_WT.append(total_negetive_cash_flow_WT)

    npv_df.insert(8, "Tax reduction for PV ($/y)", lst_total_tax_reduction_PV)
    npv_df.insert(9, "Positive Cash Flows of PV  ($/y)", lst_total_positive_cash_flow_PV)
    npv_df.insert(10, "Negative Cash Flows of PV ($/y)", lst_total_negetive_cash_flow_PV)

    npv_df.insert(11, "Grid injection of WT ($/y)", lst_grid_inj_price_WT)
    npv_df.insert(12, "WT production (MWh/y)", lst_NPV_WT_prod_age)
    npv_df.insert(13, "WT annual injection (MWh/y)", lst_NPV_WT_annual_inj_age)
    npv_df.insert(14, "Price of sold  electricity from WT ($/kWh)", lst_selling_price_WT)

    npv_df.insert(15, "Self Consumption of WT($/y)", lst_price_self_consumption_WT)
    npv_df.insert(16, "WT annual self consumption (MWh/y)", lst_NPV_self_consumption_WT_age)
    npv_df.insert(17, "Price of  electricity not purchased from the grid due to WT self consumption ($/kWh)",
                  lst_price_elec_nt_purchased)

    npv_df.insert(18, "Tax reduction for WT ($/y)", lst_total_tax_reduction_WT)
    npv_df.insert(19, "Positive Cash Flows of WT ($/y)", lst_total_positive_cash_flow_WT)
    npv_df.insert(20, "Negative Cash Flows of WT ($/y)", lst_total_negetive_cash_flow_WT)

    for i in range(0, 26):
        if i / max_lifetime == int(i / int(ent_val_ST.get())):
            batt_expense = -batt_capacity * investment_cost_batt
        else:
            batt_expense = 0

        lst_negetive_cash_flow_batt.append(batt_expense)

    lst_tax_red_st.append(0)
    for i in range(0, 25):
        if i < int(ent_val_ST.get()):
            tax_reduction_batt = float(ent_tax_red_ST.get()) * total_inv_cost_ST / 100
        else:
            tax_reduction_batt = 0
        lst_tax_red_st.append(tax_reduction_batt)

    npv_df.insert(21, "Negative cash flow related to battery investment/replacement ($/y)", lst_negetive_cash_flow_batt)
    npv_df.insert(22, "Tax reduction for storage", lst_tax_red_st)

    lst_total_positive_cash_flow = []
    lst_total_negetive_cash_flow = []
    lst_not_actualised_cash_flow = []
    lst_total_actual_cash_flow = []
    lst_NPV = []
    discount_rate = float(ent_rate.get()) / 100

    for i in range(0, 26):
        total_positive_cash_flow = lst_total_positive_cash_flow_PV[i] + lst_total_positive_cash_flow_WT[i] + \
                                   lst_tax_red_st[i]
        total_negetive_cash_flow = lst_negetive_cash_flow_batt[i] + lst_total_negetive_cash_flow_WT[i] + \
                                   lst_total_negetive_cash_flow_PV[i]
        not_actualised_cash_flow = total_positive_cash_flow + total_negetive_cash_flow
        total_actual_cash_flow = not_actualised_cash_flow / ((1 + discount_rate) ** i)

        if i == 0:
            NPV = total_actual_cash_flow
        else:
            NPV = lst_NPV[i - 1] + total_actual_cash_flow

        lst_total_positive_cash_flow.append(total_positive_cash_flow)
        lst_total_negetive_cash_flow.append(total_negetive_cash_flow)
        lst_not_actualised_cash_flow.append(not_actualised_cash_flow)
        lst_total_actual_cash_flow.append(total_actual_cash_flow)
        lst_NPV.append(NPV)

    npv_df.insert(23, "Positive Cash Flows ($/y)", lst_total_positive_cash_flow)
    npv_df.insert(24, "Negative Cash Flows ($/y)", lst_total_negetive_cash_flow)
    npv_df.insert(25, "Cash flow ( not actualized)", lst_not_actualised_cash_flow)
    npv_df.insert(26, "Total Cash Flows($/y)", lst_total_actual_cash_flow)
    npv_df.insert(27, "NPV", lst_NPV)

    IRR = npf.irr(lst_not_actualised_cash_flow) * 100
    NPV_future = lst_NPV[25] / 1000000
    PBT = sum(npv_df.NPV < 0)

    entry_set_and_disable(ent_irr, var_ent_irr, round(IRR, 2))
    entry_set_and_disable(ent_NPV, var_ent_NPV, round(NPV_future))
    entry_set_and_disable(ent_PBT, var_ent_PBT, PBT)

    lst_million_dollor_NPV = [x / 1000000 for x in lst_NPV]

    frame_NPV_graph = LabelFrame(tab_fin_ana, text="NPV data")
    frame_NPV_graph.place(x=10, y=550, height=65, width=420)

    # btn_NPV_graph = Button(frame_NPV_graph, text="NPV Graph", command=lambda : plot_graph(lst_NPV_years,lst_million_dollor_NPV,"Years","NPV[M$]","NPV Graph" ))
    btn_NPV_graph = Button(frame_NPV_graph, text="NPV Graph",
                           command=lambda: plot_bar_and_line_graph(npv_df))

    btn_NPV_graph.place(x=10, y=10)

    btn_show_NPV_data = Button(frame_NPV_graph, text="Show NPV data", command=lambda: show_data(npv_df))
    btn_show_NPV_data.place(x=90, y=10)

    btn_export_NPV_data = Button(frame_NPV_graph, text="Export NPV data",
                                 command=lambda: export_to_excel_with_name(npv_df, "NPV_data"))
    btn_export_NPV_data.place(x=190, y=10)
    ########################################################################################################
    # creation of LCOE df
    lcoe_df = pd.DataFrame()
    lst_discounted_pv_prod = []
    lst_discounted_wt_prod = []
    lst_yearly_cost_PV = []
    lst_yearly_cost_WT = []
    lst_NPV_PV_farm = []
    lst_NPV_WT_farm = []

    pv_rate = float(ent_rate.get()) / 100
    wt_rate = float(ent_rate_WT.get()) / 100
    wt_prod = float(ent_prod_WT.get()) * 1000
    # lst_PV_productivity = lst_NPV_PV_prod_age / pv_size
    if pv_size != 0:

        lst_PV_productivity = [(x * 1000) / pv_size for x in lst_NPV_PV_prod_age]
    else:
        lst_PV_productivity = [x * 0 for x in lst_NPV_PV_prod_age]
    # lst_WT_productivity = lst_NPV_WT_prod_age/ wt_size
    if wt_size != 0:
        lst_WT_productivity = [(x * 1000) / wt_size for x in lst_NPV_WT_prod_age]
    else:
        lst_WT_productivity = [x * 0 for x in lst_NPV_WT_prod_age]

    for i in range(0, 26):
        discounted_pv_prod = lst_NPV_PV_prod_age[i] / ((1 + pv_rate) ** i)
        discounted_wt_prod = lst_NPV_WT_prod_age[i] / ((1 + wt_rate) ** i)
        lst_discounted_pv_prod.append(discounted_pv_prod)
        lst_discounted_wt_prod.append(discounted_wt_prod)

        if i == 0:
            yearly_cost_PV = investment_cost_PV * pv_size
            NPV_PV_farm = yearly_cost_PV
            yearly_cost_WT = investment_cost_WT * wt_size
            NPV_WT_farm = yearly_cost_WT
        else:
            yearly_cost_PV = OandM_cost_PV * pv_size
            NPV_PV_farm = yearly_cost_PV / ((1 + pv_rate) ** i)
            yearly_cost_WT = OandM_cost_WT * wt_prod
            NPV_WT_farm = yearly_cost_WT / ((1 + wt_rate) ** i)
        lst_yearly_cost_PV.append(yearly_cost_PV)
        lst_NPV_PV_farm.append(NPV_PV_farm)
        lst_yearly_cost_WT.append(yearly_cost_WT)
        lst_NPV_WT_farm.append(NPV_WT_farm)

    lcoe_df.insert(0, "Years", lst_NPV_years)
    lcoe_df.insert(1, "NPV of Total Cost of PV plant($/y)", lst_NPV_PV_farm)
    lcoe_df.insert(2, "Yearly Cost of PV plant($/y)", lst_yearly_cost_PV)
    lcoe_df.insert(3, "Discounted PV Production of PV plant (MWh/y)", lst_discounted_pv_prod)
    lcoe_df.insert(4, "PV annual production (MWh/y)", lst_NPV_PV_prod_age)
    lcoe_df.insert(5, "PV Specific Production (kWh/kW/y)", lst_PV_productivity)
    lcoe_df.insert(6, "NPV of Total Cost of WT farm($/y)", lst_NPV_WT_farm)
    lcoe_df.insert(7, "Yearly Cost of WT farm($/y)", lst_yearly_cost_WT)
    lcoe_df.insert(8, "Discounted PV Production of WT farm(MWh/y)", lst_discounted_wt_prod)
    lcoe_df.insert(9, "WT annual production (MWh/y)", lst_NPV_WT_prod_age)
    lcoe_df.insert(10, "WT Specific Production (kWh/kW/y)", lst_WT_productivity)

    sum_NPV_PV_farm = lcoe_df["NPV of Total Cost of PV plant($/y)"].sum()
    sum_discounted_pv_prod = lcoe_df["Discounted PV Production of PV plant (MWh/y)"].sum()

    sum_NPV_WT_farm = lcoe_df["NPV of Total Cost of WT farm($/y)"].sum()
    sum_discounted_wt_prod = lcoe_df["Discounted PV Production of WT farm(MWh/y)"].sum()

    lcoe_PV = sum_NPV_PV_farm / (sum_discounted_pv_prod * 1000)
    lcoe_WT = sum_NPV_WT_farm / (sum_discounted_wt_prod * 1000)

    loce_total_plant = (sum_NPV_PV_farm + sum_NPV_WT_farm) / ((sum_discounted_wt_prod + sum_discounted_pv_prod) * 1000)
    rounded_loce_total_plant = round(loce_total_plant, 2)

    entry_set_and_disable(ent_LCOE, var_ent_LCOE, rounded_loce_total_plant)
    time_end_NPV_calculation = tm.time()
    total_time_financial_data = round(time_end_NPV_calculation - time_start_NPV_calculation, 2)
    ########################################################################################################
    newWindow.destroy()
    lbl_irr_main_tab = Label(tab_inputs, text="IRR = "+str(round(IRR,1)))
    lbl_irr_main_tab.place(x=850, y=500)
    time_end = tm.time()

    sorted_WT_productivity = lst_WT_productivity
    sorted_WT_productivity.sort()
    min_wt_productivity = round(sorted_WT_productivity[1],1)
    max_wt_productivity = round(sorted_WT_productivity[-1],1)
    avg_wt_productivity = round(mean(sorted_WT_productivity[1:]),1)

    lbl_min_Wt_productivity = Label(tab_inputs, text="WT Productivity(kWh/kW/year)\n Min = "+str(min_wt_productivity)+"\n"
    "Max = "+str (max_wt_productivity)+"\n Avg = "+str(avg_wt_productivity)
    )
    lbl_min_Wt_productivity.place(x=950, y=500)

    total_time_taken = round(time_end - time_start, 2)
    messagebox.showinfo("Completed", "Download and calculation of data completed \n"
                                     "time to download data = " + str(total_time_to_download_data) + " seconds\n"
                                                                                                     "time to calculate PV Production = " + str(
        total_time_PV_Production) + " seconds\n"

                                    "time to calculte WT production = " + str(time_WT_production) + " seconds\n"
                                                                                                    "time for calculating battery balance = " + str(
        total_time_batt_bal) + " seconds\n"
                               "time for calculating grid exchange = " + str(total_time_grid_exchange) + " seconds\n"
                                                                                                         "time taken for calculating self sufficiency = " + str(
        total_time_SS) + " seconds\n"
                         "time taken to calculate peak shaving = " + str(total_time_peak_shaving) + " seconds\n"
                                                                                                    "time taken to replicate 25 yrs of data = " + str(
        total_time_replicate_data) + " seconds\n"
                                     "time taken for financial analysis = " + str(
        total_time_financial_data) + " seconds\n"
                                     "Total time taken = " + str(total_time_taken) + " seconds")


    return IRR


#########################################################################################################
# FrontEND
tab_inputs = Frame(my_notebook)
tab_advanced_inputs = Frame(my_notebook)
tab_fin_ana = Frame(my_notebook)
tab_PV_results = Frame(my_notebook)
tab_WT_results = Frame(my_notebook)
tab_Allresults = Frame(my_notebook)
tab_results = Frame(my_notebook)
tab_opt = Frame(my_notebook)

tab_inputs.pack(fill="both", expand=1)
tab_advanced_inputs.pack(fill="both", expand=1)
tab_PV_results.pack(fill="both", expand=1)
tab_WT_results.pack(fill="both", expand=1)
tab_Allresults.pack(fill="both", expand=1)
tab_fin_ana.pack(fill="both", expand=1)
tab_results.pack(fill="both", expand=1)
tab_opt.pack(fill="both", expand=1)

my_notebook.add(tab_inputs, text="Inputs")
my_notebook.add(tab_advanced_inputs, text="Advanced Inputs")
my_notebook.add(tab_PV_results, text="PV Results")
my_notebook.add(tab_WT_results, text="WT Results")
my_notebook.add(tab_Allresults, text="All Data")
my_notebook.add(tab_fin_ana, text="Financial Analysis")
my_notebook.add(tab_results, text="Result Summary")
my_notebook.add(tab_opt, text="Optimization")

# Defining of Label Button and Entry
lbl_title = Label(tab_inputs, text="RES Tool V11", font=("Helvatical bold", 10))
lbl_title.place(x=500, y=10)
# lbl_title.grid(row=0, column=0, columnspan=2, pady=10)

frame_inputs = LabelFrame(tab_inputs, text="Installation site selection and weather data download", pady=10)
frame_inputs.place(x=40, y=40, height=250, width=310)
# frame_inputs.grid(row=1, column=0, columnspan=2, rowspan=5,pady=10)

lbl_latitude = Label(frame_inputs, text="Lattitude")
lbl_latitude.place(x=70, y=10)
# lbl_latitude.grid(row=1, column=0)
ent_latitude = Entry(frame_inputs, width=10)
ent_latitude.place(x=135, y=10)
# ent_latitude.grid(row=1, column=1,padx=5)
ent_latitude.insert(END, str(45))

lbl_longitude = Label(frame_inputs, text="Longitude")
lbl_longitude.place(x=70, y=35)
# lbl_longitude.grid(row=2, column=0)
ent_longitude = Entry(frame_inputs, width=10)
ent_longitude.place(x=135, y=35)
# ent_longitude.grid(row=2, column=1,padx=5)
ent_longitude.insert(END, str(7.5))

lbl_start_year = Label(frame_inputs, text="Start Year")
lbl_start_year.place(x=70, y=60)
# lbl_start_year.grid(row=3, column=0)
ent_start_year = Entry(frame_inputs, width=10)
ent_start_year.place(x=135, y=60)
# ent_start_year.grid(row=3, column=1,padx=5)
ent_start_year.insert(END, str(2010))

lbl_end_year = Label(frame_inputs, text="End Year")
lbl_end_year.place(x=70, y=85)
# lbl_end_year.grid(row=4, column=0)
ent_end_year = Entry(frame_inputs, width=10)
ent_end_year.place(x=135, y=85)
# ent_end_year.grid(row=4, column=1,padx=5)
ent_end_year.insert(END, str(2012))

btn_dwnld_data = Button(frame_inputs, text="Download & Calculate", command=download_data)
btn_quit = Button(frame_inputs, text="Quit", command=root.quit)
btn_dwnld_data.place(x=70, y=130)
btn_quit.place(x=125, y=165)
# btn_dwnld_data.grid(row=5, column=0,padx=10, pady=20)
# btn_quit.grid(row=5, column=1)(

lbl_India = Label(tab_inputs, text="For India")
lbl_India.place(x=50, y=500)

var_India = IntVar()
var_India.set(2)
rad_India_yes = Radiobutton(tab_inputs, text="Yes", variable=var_India, value=1)
rad_India_no = Radiobutton(tab_inputs, text= "No", variable= var_India, value=2)

rad_India_yes.place(x=105, y= 500)
rad_India_no.place(x=160, y=500)
###########################################################################################################################
frame_inputs_2 = LabelFrame(tab_inputs, text="Nominal Powers of Generators and Storage")
frame_inputs_2.place(x=370, y=40, height=250, width=430)

lbl_nominal_power = Label(frame_inputs_2, text="Nominal Power PV", fg="red")
ent_nominal_power = Entry(frame_inputs_2, width=10)
ent_nominal_power.insert(END, str(10000))
var_SI_unit_nom_pow = IntVar()
var_SI_unit_nom_pow.set(1)
rad_SI_unit_nom_pow_kw = Radiobutton(frame_inputs_2, text="kW", variable=var_SI_unit_nom_pow, value=1)
rad_SI_unit_nom_pow_Mw = Radiobutton(frame_inputs_2, text="MW", variable=var_SI_unit_nom_pow, value=2)
rad_SI_unit_nom_pow_Gw = Radiobutton(frame_inputs_2, text="GW", variable=var_SI_unit_nom_pow, value=3)

# lbl_nominal_power.grid(row=0,column=0)
# ent_nominal_power.grid(row=0,column=1)
lbl_nominal_power.place(x=10, y=10)
ent_nominal_power.place(x=120, y=10)

# rad_SI_unit_nom_pow_kw.grid(row=0, column=2)
# rad_SI_unit_nom_pow_Mw.grid(row=0, column=3)
# rad_SI_unit_nom_pow_Gw.grid(row=0, column=4)
rad_SI_unit_nom_pow_kw.place(x=200, y=10)
rad_SI_unit_nom_pow_Mw.place(x=250, y=10)
rad_SI_unit_nom_pow_Gw.place(x=300, y=10)
###########################################################################################################################
lbl_nb_turbines = Label(frame_inputs_2, text="No of Turbines", fg="red")
ent_nb_turbines = Entry(frame_inputs_2, width=10)
ent_nb_turbines.insert(END, str(10))

# lbl_nb_turbines.grid(row=0, column=0)
# ent_nb_turbines.grid(row=0, column=1)
lbl_nb_turbines.place(x=10, y=35)
ent_nb_turbines.place(x=120, y=35)

lbl_turbineID = Label(frame_inputs_2, text="Select a turbine")
lbl_turbineID.place(x=10, y=60)

lst_of_turbine = [
    "Gamesa G114-2000kW IIA/IIIA",
    "VESTAS V 150 - 4.0",
    "VESTAS V 90 3 MW",
    "LTW101 2500_kW IIIA+",
    "LTW101 3000_kW IIIA+",
    "Hitachi HTW2.0-80",
    "LTW77 800_kW IIA/IIIA",
    "Vestas V90-2000_kW IIA",
    "ENERCON E-82 E-2 IIA",
    "Nordex N117 Gamma IIIA"
]
var_turbine = StringVar()
var_turbine.set(lst_of_turbine[0])
drpdwn_turbine = OptionMenu(frame_inputs_2, var_turbine, *lst_of_turbine)

var_turbine.trace("w", lambda *args: show_nominal_power())
drpdwn_turbine.place(x=120, y=60)

# Creating power curve df###################################################################################
lst_wind_speed_pcGamesa = [0, 1, 2, 2.9, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
                           24, 25, 25.1, 26, 27, 28, 29, 30]
lst_pow_Gamesa = [0, 0, 0, 0, 32, 146, 342, 621, 1008, 1486, 1836, 1965, 1994, 1999, 2000, 2000, 2000, 2000, 2000, 2000,
                  2000, 2000, 2000, 1906, 1681, 1455, 1230, 0, 0, 0, 0, 0, 0]
lst_wind_speed_pcVestas150 = [0, 1, 2, 2.9, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
                              24, 24.1, 25, 26, 27, 28, 29, 30]
lst_pow_vestas150 = [0, 0, 0, 0, 81, 285, 597, 1062, 1709, 2545, 3458, 3934, 3999, 4000, 4000, 4000, 4000, 4000, 4000,
                     4000, 4000, 4000, 3742, 2730, 1805, 1284, 0, 0, 0, 0, 0, 0, 0]
lst_wind_speed_pcVestas90 = [0, 1, 2, 3, 3.9, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
                             24, 25, 25.1, 26, 27, 28, 29, 30]
lst_pow_vestas90 = [0, 0, 0, 0, 0, 75, 200, 350, 600, 900, 1300, 1700, 2100, 2500, 2900, 2950, 3000, 3000, 3000, 3000,
                    3000, 3000, 3000, 3000, 3000, 3000, 3000, 0, 0, 0, 0, 0, 0]
lst_wind_speed_pcLTW2500 = [0, 1, 2, 2.9, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
                            24, 25, 25.1, 26, 27, 28, 29, 30]
lst_pow_LTW2500 = [0, 0, 0, 0, 41, 118, 258, 470, 769, 1154, 1634, 2125, 2402, 2500, 2500, 2500, 2500, 2500, 2500, 2500,
                   2500, 2500, 2500, 2500, 2500, 2500, 2500, 0, 0, 0, 0, 0, 0]
lst_wind_speed_pcLTW3000 = [0, 1, 2, 2.9, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
                            24, 25, 25.1, 26, 27, 28, 29, 30]
lst_pow_LTW3000 = [0, 0, 0, 0, 29, 111, 252, 461, 739, 1079, 1474, 1914, 2354, 2701, 2897, 2974, 3000, 3000, 3000, 3000,
                   3000, 3000, 3000, 3000, 3000, 3000, 3000, 0, 0, 0, 0, 0, 0]
lst_wind_speed_pcHitachi = [0, 1, 2, 3, 3.9, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
                            24, 25, 25.1, 26, 27, 28, 29, 30]
lst_pow_Hitachi = [0, 0, 0, 0, 0, 62, 168, 320, 516, 771, 1083, 1417, 1727, 1913, 2000, 2000, 2000, 2000, 2000, 2000,
                   2000, 2000, 2000, 2000, 2000, 2000, 2000, 0, 0, 0, 0, 0, 0]
lst_wind_speed_pcLTW800 = [0, 1, 2, 2.9, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
                           24, 25, 25.1, 26, 27, 28, 29, 30]
lst_pow_LTW800 = [0, 0, 0, 0, 24, 69, 150, 269, 433, 605, 736, 788, 800, 800, 800, 800, 800, 800, 800, 800, 800, 800,
                  800, 800, 800, 800, 800, 0, 0, 0, 0, 0, 0]
lst_wind_speed_pcVestas90_2000 = [0, 1, 2, 3, 3.9, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22,
                                  23, 24, 25, 25.1, 26, 27, 28, 29, 30]
lst_pow_vestas90_2000 = [0, 0, 0, 0, 0, 75, 190, 354, 582, 883, 1240, 1604, 1893, 2005, 2027, 2029, 2030, 2030, 2030,
                         2030, 2030, 2030, 2030, 2030, 2030, 2030, 2030, 0, 0, 0, 0, 0, 0]
lst_wind_speed_pcEnercon = [0, 1, 1.9, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
                            24, 25, 26, 27, 28, 28.1, 29, 30]
lst_pow_Enercon = [0, 0, 0, 3, 25, 82, 174, 321, 532, 815, 1180, 1580, 1810, 1980, 2050, 2050, 2050, 2050, 2050, 2050,
                   2050, 2050, 2050, 2050, 2050, 2050, 2050, 2050, 2050, 2050, 0, 0, 0]
lst_wind_speed_pcNordex = [0, 1, 2, 2.9, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 20.1, 21, 22,
                           23, 24, 25, 26, 27, 28, 29, 30]
lst_pow_Nordex = [0, 0, 0, 0, 23, 154, 356, 644, 1037, 1528, 2039, 2325, 2385, 2400, 2400, 2400, 2400, 2400, 2400, 2400,
                  2400, 2400, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

power_curve_df = pd.DataFrame()
power_curve_df.insert(0, "Wind speed(m/s) of Gamesa", lst_wind_speed_pcGamesa)
power_curve_df.insert(1, "Gamesa G114-2000kW IIA/IIIA", lst_pow_Gamesa)
power_curve_df.insert(2, "Wind speed(m/s) of Vestas 150", lst_wind_speed_pcVestas150)
power_curve_df.insert(3, "VESTAS V 150 - 4.0", lst_pow_vestas150)
power_curve_df.insert(4, "Wind speed(m/s) of Vestas 90", lst_wind_speed_pcVestas90)
power_curve_df.insert(5, "VESTAS V 90 3 MW", lst_pow_vestas90)
power_curve_df.insert(6, "Wind speed(m/s) of LTW 2500", lst_wind_speed_pcLTW2500)
power_curve_df.insert(7, "LTW101 2500_kW IIIA+", lst_pow_LTW2500)
power_curve_df.insert(8, "Wind speed(m/s) of LTW 3000", lst_wind_speed_pcLTW3000)
power_curve_df.insert(9, "LTW101 3000_kW IIIA+", lst_pow_LTW3000)
power_curve_df.insert(10, "Wind speed(m/s) of Hitachi", lst_wind_speed_pcHitachi)
power_curve_df.insert(11, "Hitachi HTW2.0-80", lst_pow_Hitachi)
power_curve_df.insert(12, "Wind speed(m/s) of LTW 800", lst_wind_speed_pcLTW800)
power_curve_df.insert(13, "LTW77 800_kW IIA/IIIA", lst_pow_LTW800)
power_curve_df.insert(14, "Wind speed(m/s) of Vestas 90-2000", lst_wind_speed_pcVestas90_2000)
power_curve_df.insert(15, "Vestas V90-2000_kW IIA", lst_pow_vestas90_2000)
power_curve_df.insert(16, "Wind speed(m/s) of Enercon", lst_wind_speed_pcEnercon)
power_curve_df.insert(17, "ENERCON E-82 E-2 IIA", lst_pow_Enercon)
power_curve_df.insert(18, "Wind speed(m/s) of Nordex", lst_wind_speed_pcNordex)
power_curve_df.insert(19, "Nordex N117 Gamma IIIA", lst_pow_Nordex)

# creating parameters df of wind turbines################################################
para_WT_df = pd.DataFrame(index=["Height of rotor", "Nominal Power", "Specific anual production", "Length of blade"])
para_WT_df.insert(0, "Gamesa G114-2000kW IIA/IIIA", [93, 2000, 2593, 57])
para_WT_df.insert(1, "VESTAS V 150 - 4.0", [105, 4000, 0, 77.5])
para_WT_df.insert(2, "VESTAS V 90 3 MW", [80, 3000, 0, 44])
para_WT_df.insert(3, "LTW101 2500_kW IIIA+", [80, 2500, 0, 50.5])
para_WT_df.insert(4, "LTW101 3000_kW IIIA+", [93.5, 3000, 0, 50.5])
para_WT_df.insert(5, "Hitachi HTW2.0-80", [78, 2000, 0, 39])
para_WT_df.insert(6, "LTW77 800_kW IIA/IIIA", [61.5, 800, 0, 38.5])
para_WT_df.insert(7, "Vestas V90-2000_kW IIA", [80, 2030, 0, 45])
para_WT_df.insert(8, "ENERCON E-82 E-2 IIA", [78, 2050, 0, 41])
para_WT_df.insert(9, "Nordex N117 Gamma IIIA", [91, 2400, 0, 58.4])

#########################################################################################################
lbl_WT_nom_pow = Label(frame_inputs_2, text="Nominal power of each WTG")
ent_WT_nom_pow = Entry(frame_inputs_2, width=10)
wt_nom_pow = para_WT_df["Gamesa G114-2000kW IIA/IIIA"][1]
ent_WT_nom_pow.insert(END, str(wt_nom_pow))
ent_WT_nom_pow.config(state="disabled")
lbl_SI_nom_pow = Label(frame_inputs_2, text="kW")

sv_nom_pow_wf = StringVar()
lbl_nom_pow_WF = Label(frame_inputs_2, text="Nominal power of Wind farm")
# ent_nom_pow_WF =  Entry(frame_inputs_2, width= 10, validate="focusout", validatecommand=show_nom_pow_Wind_farm)
# sv_nom_pow_wf.trace("w", lambda name, index, mode, sv= sv_nom_pow_wf: show_nom_pow_Wind_farm())

ent_nom_pow_WF = Entry(frame_inputs_2, width=10, textvariable=sv_nom_pow_wf)
# sv_nom_pow_wf.trace_add("write",show_nom_pow_Wind_farm)
lbl_SI_nom_pow_WF = Label(frame_inputs_2, text="kW")

lst_SI_units_wind_farm = [
    "kWh",
    "MWh",
    "GWh"
]

var_SI_wind_farm = StringVar()
var_SI_wind_farm.set(lst_SI_units_wind_farm[0])

drpdwn_wind_farm_SI = OptionMenu(frame_inputs_2, var_SI_wind_farm, *lst_SI_units_wind_farm)
var_SI_wind_farm.trace("w", lambda *args: change_SI_wind_farm())

nb_WT = ent_nb_turbines.get()
nom_pow_wf = int(nb_WT) * int(wt_nom_pow)
ent_nom_pow_WF.insert(END, str(nom_pow_wf))
ent_nom_pow_WF.config(state="disabled")
global nom_pow_wind_farm
nom_pow_wind_farm = float(ent_nom_pow_WF.get())

lbl_WT_nom_pow.place(x=10, y=95)
ent_WT_nom_pow.place(x=230, y=95)
lbl_SI_nom_pow.place(x=300, y=95)

lbl_nom_pow_WF.place(x=10, y=120)
ent_nom_pow_WF.place(x=230, y=120)
# lbl_SI_nom_pow_WF.place(x=300, y=120)
drpdwn_wind_farm_SI.place(x=300, y=120)

btn_power_curve = Button(frame_inputs_2, text="Power Curve",
                         command=lambda: show_power_curve(power_curve_df, var_turbine.get()))
btn_compare_PC = Button(frame_inputs_2, text="Compare Power curves",
                        command=lambda: compare_power_curve(power_curve_df))
btn_power_curve.place(x=10, y=150)
btn_compare_PC.place(x=100, y=150)
###########################################################################################################################
lbl_storage_capacity = Label(frame_inputs_2, text="Storage Capacity", fg="red")
ent_storage_capacity = Entry(frame_inputs_2, width=10)
ent_storage_capacity.insert(END, str(1))
lbl_storage_capacity.place(x=10, y=185)
ent_storage_capacity.place(x=120, y=185)
# lbl_SI_storage_capacity = Label(frame_batt_para,text="MWh")

var_SI_unit_storage_capacity = IntVar()
var_SI_unit_storage_capacity.set(2)
rad_SI_unit_storage_capacity_kW = Radiobutton(frame_inputs_2, text="kWh", variable=var_SI_unit_storage_capacity,
                                              value=1)
rad_SI_unit_storage_capacity_MW = Radiobutton(frame_inputs_2, text="MWh", variable=var_SI_unit_storage_capacity,
                                              value=2)
rad_SI_unit_storage_capacity_GW = Radiobutton(frame_inputs_2, text="GWh", variable=var_SI_unit_storage_capacity,
                                              value=3)

rad_SI_unit_storage_capacity_kW.place(x=200, y=185)
rad_SI_unit_storage_capacity_MW.place(x=260, y=185)
rad_SI_unit_storage_capacity_GW.place(x=320, y=185)

###########################################################################################################################
# Inside frame abbrevation
frame_abb = LabelFrame(tab_advanced_inputs, text="Inputs and Parameters of PV conversion model", padx=10, pady=10)

lbl_etaDirt = Label(frame_abb, text="ETA_Dirt")
ent_etaDirt = Entry(frame_abb, width=10)
ent_etaDirt.insert(END, str(0.976))
lbl_SI_etaDirt = Label(frame_abb, text="()")
lbl_D_etaDirt = Label(frame_abb, text="Losses due to dirt", font=("Helvatical bold", 7))
# lbl_D_etaDirt.config(font=5)

lbl_etaRef = Label(frame_abb, text="ETA_Reflection")
ent_etaRef = Entry(frame_abb, width=10)
ent_etaRef.insert(END, str(0.973))
lbl_SI_etaRef = Label(frame_abb, text="()")
lbl_D_etaRef = Label(frame_abb, text="Losses due to reflection", font=("Helvatical bold", 7))

lbl_etaMM = Label(frame_abb, text="ETA_Mismatch")
ent_etaMM = Entry(frame_abb, width=10)
ent_etaMM.insert(END, str(0.97))
lbl_SI_etaMM = Label(frame_abb, text="()")
lbl_D_etaMM = Label(frame_abb, text="Losses due to mismatch", font=("Helvatical bold", 7))

lbl_etaCable = Label(frame_abb, text="ETA_Cable")
ent_etaCable = Entry(frame_abb, width=10)
ent_etaCable.insert(END, str(0.99))
lbl_SI_etaCable = Label(frame_abb, text="()")
lbl_D_etaCable = Label(frame_abb, text="Losses in cable", font=("Helvatical bold", 7))

lbl_Gamma = Label(frame_abb, text="gamma_th")
ent_Gamma = Entry(frame_abb, width=10)
ent_Gamma.insert(END, str(-0.005))
lbl_SI_Gamma = Label(frame_abb, text="1/C")
lbl_D_Gamma = Label(frame_abb, text="Power reduction Co-efficient", font=("Helvatical bold", 7))

lbl_G0 = Label(frame_abb, text="G0")
ent_G0 = Entry(frame_abb, width=10)
ent_G0.insert(END, str(17.7))
lbl_SI_G0 = Label(frame_abb, text="W/m2", padx=5)
lbl_D_G0 = Label(frame_abb, text="Minimum radiation to switch on the system", font=("Helvatical bold", 7))

lbl_T_STC = Label(frame_abb, text="T_stc")
ent_T_STC = Entry(frame_abb, width=10)
ent_T_STC.insert(END, str(25))
lbl_SI_T_STC = Label(frame_abb, text="C")
lbl_D_T_STC = Label(frame_abb, text="Temperature test condition", font=("Helvatical bold", 7))

lbl_noct = Label(frame_abb, text="NOCT", fg="blue")
ent_noct = Entry(frame_abb, width=10)
ent_noct.insert(END, str(47))
lbl_SI_noct = Label(frame_abb, text="C")
lbl_D_noct = Label(frame_abb, text="Normal operating cell temperature", font=("Helvatical bold", 7))

lbl_Trif_noct = Label(frame_abb, text="T_rif NOCT")
ent_Trif_noct = Entry(frame_abb, width=10)
ent_Trif_noct.insert(END, str(20))
lbl_SI_Trif_noct = Label(frame_abb, text="C")
lbl_D_Trif_noct = Label(frame_abb, text="Reference temperature for NOCT", font=("Helvatical bold", 7))

lbl_rad_noct = Label(frame_abb, text="rad_NOCT")
ent_rad_noct = Entry(frame_abb, width=10)
ent_rad_noct.insert(END, str(0.8))
lbl_SI_rad_noct = Label(frame_abb, text="kW/m2")
lbl_D_rad_noct = Label(frame_abb, text="Reference radiation for NOCT", font=("Helvatical bold", 7))

# Inside frame abbrevation
frame_abb.place(x=10, y=10, height=300, width=500)

lbl_etaDirt.place(x=10, y=10)
ent_etaDirt.place(x=150, y=10)
lbl_SI_etaDirt.place(x=200, y=10)
lbl_D_etaDirt.place(x=250, y=10)

lbl_etaRef.place(x=10, y=35)
ent_etaRef.place(x=150, y=35)
lbl_SI_etaRef.place(x=200, y=35)
lbl_D_etaRef.place(x=250, y=35)

lbl_etaMM.place(x=10, y=60)
ent_etaMM.place(x=150, y=60)
lbl_SI_etaMM.place(x=200, y=60)
lbl_D_etaMM.place(x=250, y=60)

lbl_etaCable.place(x=10, y=85)
ent_etaCable.place(x=150, y=85)
lbl_SI_etaCable.place(x=200, y=85)
lbl_D_etaCable.place(x=250, y=85)

lbl_Gamma.place(x=10, y=110)
ent_Gamma.place(x=150, y=110)
lbl_SI_Gamma.place(x=200, y=110)
lbl_D_Gamma.place(x=250, y=110)

lbl_G0.place(x=10, y=135)
ent_G0.place(x=150, y=135)
lbl_SI_G0.place(x=200, y=135)
lbl_D_G0.place(x=250, y=135)

lbl_T_STC.place(x=10, y=160)
ent_T_STC.place(x=150, y=160)
lbl_SI_T_STC.place(x=200, y=160)
lbl_D_T_STC.place(x=250, y=160)

lbl_noct.place(x=10, y=185)
ent_noct.place(x=150, y=185)
lbl_SI_noct.place(x=200, y=185)
lbl_D_noct.place(x=250, y=185)

lbl_Trif_noct.place(x=10, y=210)
ent_Trif_noct.place(x=150, y=210)
lbl_SI_Trif_noct.place(x=200, y=210)
lbl_D_Trif_noct.place(x=250, y=210)

lbl_rad_noct.place(x=10, y=235)
ent_rad_noct.place(x=150, y=235)
lbl_SI_rad_noct.place(x=200, y=235)
lbl_D_rad_noct.place(x=250, y=235)
######################################################################

# Inside frame Parameters for model
frame_para = LabelFrame(tab_advanced_inputs, text="Parameters of DC/AC conversion model", padx=10, pady=10)

lbl_noloadloss = Label(frame_para, text="no-load losses")
ent_noloadloss = Entry(frame_para, width=10)
ent_noloadloss.insert(END, str(70))
lbl_SI_noloadloss = Label(frame_para, text="()")
lbl_D_noloadloss = Label(frame_para, text="Corresponding to 0.7% losses at nominal power", font=("Helvatical bold", 7))

lbl_linearloss = Label(frame_para, text="Linear losses")
ent_linearloss = Entry(frame_para, width=10)
ent_linearloss.insert(END, str(0.007))
lbl_SI_linearloss = Label(frame_para, text="()")
lbl_D_linearloss = Label(frame_para, text="Corresponding to 0.7% losses at nominal power", font=("Helvatical bold", 7))

lbl_Q_loss = Label(frame_para, text="Quadratic losses")
ent_Q_loss = Entry(frame_para, width=10)
ent_Q_loss.insert(END, str(0.0000007))
lbl_SI_Q_loss = Label(frame_para, text="()")
lbl_D_Q_loss = Label(frame_para, text="Corresponding to 0.7% losses at nominal power", font=("Helvatical bold", 7))

lbl_n_point = Label(frame_para, text="n_point_interpolation")
ent_n_point = Entry(frame_para, width=10)
ent_n_point.insert(END, str(100))
lbl_SI_n_point = Label(frame_para, text="()")
lbl_D_n_point = Label(frame_para, text="Points to interpolate efficiency curve of the convertor",
                      font=("Helvatical bold", 7))

frame_para.place(x=520, y=10, height=200, width=500)

lbl_noloadloss.place(x=10, y=10)
ent_noloadloss.place(x=150, y=10)
lbl_SI_noloadloss.place(x=200, y=10)
lbl_D_noloadloss.place(x=250, y=10)

lbl_linearloss.place(x=10, y=35)
ent_linearloss.place(x=150, y=35)
lbl_SI_linearloss.place(x=200, y=35)
lbl_D_linearloss.place(x=250, y=35)

lbl_Q_loss.place(x=10, y=60)
ent_Q_loss.place(x=150, y=60)
lbl_SI_Q_loss.place(x=200, y=60)
lbl_D_Q_loss.place(x=250, y=60)

lbl_n_point.place(x=10, y=85)
ent_n_point.place(x=150, y=85)
lbl_SI_n_point.place(x=200, y=85)
lbl_D_n_point.place(x=250, y=85)
######################################################################
# peak shaving parameters###################################################################

frame_peak_shaving = LabelFrame(tab_inputs, text="Peak shaving", padx=10, pady=10)
lst_peak_shaving_options = [
    "No limitations",
    "Limitation on Maximum generated power",
    "Limitation on Maximum power injection"

]
var_peak_shaving = StringVar()
var_peak_shaving.set(lst_peak_shaving_options[0])
lbl_peak_shaving = Label(frame_peak_shaving, text="Peak shaving typology")
drpdwn_peak_shaving = OptionMenu(frame_peak_shaving, var_peak_shaving, *lst_peak_shaving_options)
var_peak_shaving.trace("w", lambda *args: show_hide_peak_shaving())
lbl_max_gen = Label(frame_peak_shaving, text="Maximum generation from PV+WT")
ent_max_gen = Entry(frame_peak_shaving, width=10)
ent_max_gen.config(state="disabled")
lbl_max_inj = Label(frame_peak_shaving, text="Maximum injection in the grid")
ent_max_inj = Entry(frame_peak_shaving, width=10)
ent_max_inj.config(state="disabled")

var_SI_unit_PS = IntVar()
var_SI_unit_PS.set(2)
rad_SI_unit_PS_kW = Radiobutton(frame_peak_shaving, text="kW", variable=var_SI_unit_PS, value=1)
rad_SI_unit_PS_MW = Radiobutton(frame_peak_shaving, text="MW", variable=var_SI_unit_PS, value=2)
rad_SI_unit_PS_GW = Radiobutton(frame_peak_shaving, text="GW", variable=var_SI_unit_PS, value=3)

frame_peak_shaving.place(x=370, y=300, height=180, width=430)

lbl_peak_shaving.place(x=10, y=10)
drpdwn_peak_shaving.place(x=140, y=5)

lbl_max_gen.place(x=10, y=45)
ent_max_gen.place(x=220, y=45)

lbl_max_inj.place(x=10, y=70)
ent_max_inj.place(x=220, y=70)

rad_SI_unit_PS_kW.place(x=10, y=95)
rad_SI_unit_PS_MW.place(x=60, y=95)
rad_SI_unit_PS_GW.place(x=110, y=95)
# wind data parameters###################################################################
frame_wind_para = LabelFrame(tab_advanced_inputs, text="Parameters of Wind Turbine", padx=10, pady=10)

lbl_mes_height = Label(frame_wind_para, text="Measurement Height")
ent_mes_height = Entry(frame_wind_para, width=10)
ent_mes_height.insert(END, str(10))
lbl_SI_mes_height = Label(frame_wind_para, text="(m)")
lbl_D_mes_height = Label(frame_wind_para, text="The height of wind station to measure wind speed",
                         font=("Helvatical bold", 7))

lbl_air_density = Label(frame_wind_para, text="Air Density")
ent_air_density = Entry(frame_wind_para, width=10)
ent_air_density.insert(END, str(1.225))
lbl_SI_air_density = Label(frame_wind_para, text="[Kg/m3]")
lbl_D_air_density = Label(frame_wind_para, text="Air Density", font=("Helvatical bold", 7))

lbl_terrain_roug = Label(frame_wind_para, text="Roughness of terrain")
ent_terrain_roug = Entry(frame_wind_para, width=10)
ent_terrain_roug.insert(END, str(0.15))
lbl_SI_terrain_roug = Label(frame_wind_para, text="()")
lbl_D_terrain_roug = Label(frame_wind_para, text="It is a factor witch represents the irregularity\nof the ground",
                           font=("Helvatical bold", 7))

# Inside wind parameters frame###################################
# frame_wind_para.grid(row=11, column=0, columnspan=4, rowspan=7, sticky=tk.NW)
frame_wind_para.place(x=10, y=320, height=130, width=500)

lbl_mes_height.place(x=10, y=10)
ent_mes_height.place(x=150, y=10)
lbl_SI_mes_height.place(x=200, y=10)
lbl_D_mes_height.place(x=250, y=10)

lbl_air_density.place(x=10, y=35)
ent_air_density.place(x=150, y=35)
lbl_SI_air_density.place(x=200, y=35)
lbl_D_air_density.place(x=250, y=38)

lbl_terrain_roug.place(x=10, y=60)
ent_terrain_roug.place(x=150, y=60)
lbl_SI_terrain_roug.place(x=200, y=60)
lbl_D_terrain_roug.place(x=250, y=60)

######################################################################


#########################################################################################################
# Battery storage parameters##############################################################################
frame_batt_para = LabelFrame(tab_advanced_inputs, text="Parameters of Battery storage", padx=10, pady=15)

lbl_max_soc = Label(frame_batt_para, text="Max State Of Charge")
ent_max_soc = Entry(frame_batt_para, width=10)
ent_max_soc.insert(END, str(100))
lbl_SI_max_soc = Label(frame_batt_para, text="%")

lbl_min_soc = Label(frame_batt_para, text="Min State Of Charge")
ent_min_soc = Entry(frame_batt_para, width=10)
ent_min_soc.insert(END, str(20))
lbl_SI_min_soc = Label(frame_batt_para, text="%")

lbl_load = Label(frame_batt_para, text="Constant Load")
ent_load = Entry(frame_batt_para, width=10)
ent_load.insert(END, str(1000))
lbl_SI_load = Label(frame_batt_para, text="kW")

lbl_discharge_eff = Label(frame_batt_para, text="Discharge efficiency")
ent_discharge_eff = Entry(frame_batt_para, width=10)
ent_discharge_eff.insert(END, str(90))
lbl_SI_discharge_eff = Label(frame_batt_para, text="%")

lbl_charge_eff = Label(frame_batt_para, text="Charge efficiency")
ent_charge_eff = Entry(frame_batt_para, width=10)
ent_charge_eff.insert(END, str(90))
lbl_SI_charge_eff = Label(frame_batt_para, text="%")
#########################################################################################################

# frame_batt_para.grid(row=18, column=0, columnspan=3, rowspan=6, sticky=tk.NW)
frame_batt_para.place(x=520, y=220, height=150, width=250)

lbl_max_soc.place(x=10, y=10)
ent_max_soc.place(x=150, y=10)
lbl_SI_max_soc.place(x=200, y=10)

lbl_min_soc.place(x=10, y=35)
ent_min_soc.place(x=150, y=35)
lbl_SI_min_soc.place(x=200, y=35)

lbl_discharge_eff.place(x=10, y=60)
ent_discharge_eff.place(x=150, y=60)
lbl_SI_discharge_eff.place(x=200, y=60)

lbl_charge_eff.place(x=10, y=85)
ent_charge_eff.place(x=150, y=85)
lbl_SI_charge_eff.place(x=200, y=85)

# lbl_load .grid(row=5, column=0)
# ent_load .grid(row=5, column=1)
# lbl_SI_load .grid(row=5, column=2)
######################################################################
# load frame
frm_load = LabelFrame(tab_inputs, text="Import of Electrical Load Profile")

lst_load_type = [
    "Load of 1 year and increase %",
    "Load of 25 years"
]
var_load_entry = tk.StringVar()
lbl_load_type = Label(frm_load, text="Load Input type")
ent_load_type = Entry(frm_load, width=30, textvariable=var_load_entry)
var_load = StringVar()
var_load.set(lst_load_type[0])
drpdwn_load = OptionMenu(frm_load, var_load, *lst_load_type)
btn_load_excel = Button(frm_load, text="Load Excel", command=open_file)

frm_load.place(x=40, y=300, height=180, width=310)
lbl_load_type.place(x=10, y=30)
drpdwn_load.place(x=100, y=28)
btn_load_excel.place(x=10, y=70)
ent_load_type.place(x=100, y=75)
#########################################################################################################
# positioning of Label Button and Entry

##Peak shaving frame###################################################

######################################################################
# frm_load.grid(row=18, column=3, columnspan=3, rowspan=6, sticky=tk.NW)
# lbl_load_type.grid(row=0, column=0)
# drpdwn_load.grid(row=0, column=1)
# btn_load_excel.grid(row=1, column=0)
# ent_load_type.grid(row=1, column=1)

###Front end of Financial analysis###################################################################
########Input Parameters for PhotoVoltaic (PV) Financial Assessment####################################################
frame_PV_FA = LabelFrame(tab_fin_ana, text="Input Parameters for PhotoVoltaic (PV) Financial Assessment")
frame_PV_FA.place(x=10, y=10, height=310, width=420)

lbl_ins_cost = Label(frame_PV_FA, text="Ins Cost")
ent_ins_cost = Entry(frame_PV_FA, width=10)
lbl_SI_ins_cost = Label(frame_PV_FA, text="$/kW")
lbl_des_ins_cost = Label(frame_PV_FA, text="(Total Installment cost)")

ent_ins_cost.insert(END, str(700))

lbl_ins_cost.place(x=10, y=10)
ent_ins_cost.place(x=100, y=10)
lbl_SI_ins_cost.place(x=170, y=10)
lbl_des_ins_cost.place(x=220, y=10)

lbl_om = Label(frame_PV_FA, text="O & M")
ent_om = Entry(frame_PV_FA, width=10)
lbl_SI_om = Label(frame_PV_FA, text="$/kW/y")
lbl_des_om = Label(frame_PV_FA, text="(Operation and Maintenece cost)")

ent_om.insert(END, str(10))

lbl_om.place(x=10, y=35)
ent_om.place(x=100, y=35)
lbl_SI_om.place(x=170, y=35)
lbl_des_om.place(x=220, y=35)

var_ent_size = tk.StringVar()
lbl_size = Label(frame_PV_FA, text="Size")
ent_size = Entry(frame_PV_FA, width=10, textvariable=var_ent_size)
lbl_SI_size = Label(frame_PV_FA, text="MWp")
lbl_des_size = Label(frame_PV_FA, text="(Size of the Plant)")
ent_size.config(state="disabled")

lbl_size.place(x=10, y=60)
ent_size.place(x=100, y=60)
lbl_SI_size.place(x=170, y=60)
lbl_des_size.place(x=220, y=60)

var_ent_prod = tk.StringVar()
lbl_prod = Label(frame_PV_FA, text="Production")
ent_prod = Entry(frame_PV_FA, width=10, textvariable=var_ent_prod)
lbl_SI_prod = Label(frame_PV_FA, text="MWh/y")
lbl_des_prod = Label(frame_PV_FA, text="(Energy Production of the Plant of first year)", font=("Helvatical bold", 7))
ent_prod.config(state="disabled")

lbl_prod.place(x=10, y=85)
ent_prod.place(x=100, y=85)
lbl_SI_prod.place(x=170, y=85)
lbl_des_prod.place(x=220, y=85)

var_ent_grid_inj = tk.StringVar()
lbl_grid_inj = Label(frame_PV_FA, text="Grid Injection")
ent_grid_inj = Entry(frame_PV_FA, width=10, textvariable=var_ent_grid_inj)
lbl_SI_grid_inj = Label(frame_PV_FA, text="MWh/y")
lbl_des_grid_inj = Label(frame_PV_FA, text=" Starting Specific Production (first year)", font=("Helvatical bold", 7))
ent_grid_inj.config(state="disabled")

lbl_grid_inj.place(x=10, y=110)
ent_grid_inj.place(x=100, y=110)
lbl_SI_grid_inj.place(x=170, y=110)
lbl_des_grid_inj.place(x=220, y=110)

lbl_rate = Label(frame_PV_FA, text="Rate")
ent_rate = Entry(frame_PV_FA, width=10)
lbl_SI_rate = Label(frame_PV_FA, text="%")
lbl_des_rate = Label(frame_PV_FA, text="(Discount Rate)")

ent_rate.insert(END, str(3))

lbl_rate.place(x=10, y=135)
ent_rate.place(x=100, y=135)
lbl_SI_rate.place(x=170, y=135)
lbl_des_rate.place(x=220, y=135)

var_ent_life = tk.StringVar()
lbl_life = Label(frame_PV_FA, text="Life")
ent_life = Entry(frame_PV_FA, width=10, textvariable=var_ent_life)
lbl_SI_life = Label(frame_PV_FA, text="Years")
lbl_des_life = Label(frame_PV_FA, text="(Useful Life of the Plant)")
ent_life.config(state="disabled")

lbl_life.place(x=10, y=160)
ent_life.place(x=100, y=160)
lbl_SI_life.place(x=170, y=160)
lbl_des_life.place(x=220, y=160)

lbl_ageing = Label(frame_PV_FA, text="Ageing")
ent_ageing = Entry(frame_PV_FA, width=10)
lbl_SI_ageing = Label(frame_PV_FA, text="%")
lbl_des_ageing = Label(frame_PV_FA, text="( Yearly Degradation of the Plant )")

ent_ageing.insert(END, str(0.5))

lbl_ageing.place(x=10, y=185)
ent_ageing.place(x=100, y=185)
lbl_SI_ageing.place(x=170, y=185)
lbl_des_ageing.place(x=220, y=185)

lbl_ele_sp = Label(frame_PV_FA, text="Electricity selling \n price", font=("Helvatical bold", 8))
ent_ele_sp = Entry(frame_PV_FA, width=10)
lbl_SI_ele_sp = Label(frame_PV_FA, text="$/kWh")
lbl_des_ele_sp = Label(frame_PV_FA, text="(For First 5 years)")
ent_ele_sp2 = Entry(frame_PV_FA, width=10)
lbl_SI_ele_sp2 = Label(frame_PV_FA, text="$/kWh")
lbl_des_ele_sp2 = Label(frame_PV_FA, text="(After 5th year)")

ent_ele_sp.insert(END, str(0.04))
ent_ele_sp2.insert(END, str(0.05))

lbl_ele_sp.place(x=10, y=210)
# ent_ele_sp.place(x=100, y=210)
# lbl_SI_ele_sp.place(x=170, y=210)
# lbl_des_ele_sp.place(x=220, y=210)
# ent_ele_sp2.place(x=100, y=235)
# lbl_SI_ele_sp2.place(x=170, y=235)
# lbl_des_ele_sp2.place(x=220, y=235)


btn_elec_sp = Button(frame_PV_FA, text= "Enter yearly selling price", command=variable_selling_price_pv)
btn_elec_sp.place(x=100, y=210)

var_ent_spec_prod = tk.StringVar()
lbl_spec_prod = Label(frame_PV_FA, text="Specific\nProduction")
ent_spec_prod = Entry(frame_PV_FA, width=10, textvariable=var_ent_spec_prod)
lbl_SI_spec_prod = Label(frame_PV_FA, text="MWh/MW/y", font=("Helvatical ", 7))
lbl_des_spec_prod = Label(frame_PV_FA, text="(Productivity of the Plant for first year)", font=("Helvatical bold", 8))
ent_spec_prod.config(state="disabled")

lbl_spec_prod.place(x=10, y=255)
ent_spec_prod.place(x=100, y=260)
lbl_SI_spec_prod.place(x=170, y=260)
lbl_des_spec_prod.place(x=220, y=260)
#######################################################################################################################
frame_WT_FA = LabelFrame(tab_fin_ana, text="Input Parameters for Wind Turbine (WT) Financial Assessment")
frame_WT_FA.place(x=450, y=10, height=310, width=420)

lbl_ins_cost_WT = Label(frame_WT_FA, text="Ins Cost")
ent_ins_cost_WT = Entry(frame_WT_FA, width=10)
lbl_SI_ins_cost_WT = Label(frame_WT_FA, text="$/kW")
lbl_des_ins_cost_WT = Label(frame_WT_FA, text="(Total Installment cost)")

ent_ins_cost_WT.insert(END, str(1100))

lbl_ins_cost_WT.place(x=10, y=10)
ent_ins_cost_WT.place(x=100, y=10)
lbl_SI_ins_cost_WT.place(x=170, y=10)
lbl_des_ins_cost_WT.place(x=220, y=10)

lbl_om_WT = Label(frame_WT_FA, text="O & M")
ent_om_WT = Entry(frame_WT_FA, width=10)
lbl_SI_om_WT = Label(frame_WT_FA, text="$/kW/y")
lbl_des_om_WT = Label(frame_WT_FA, text="(Operation and Maintenece cost)")

ent_om_WT.insert(END, str(0.0085))

lbl_om_WT.place(x=10, y=35)
ent_om_WT.place(x=100, y=35)
lbl_SI_om_WT.place(x=170, y=35)
lbl_des_om_WT.place(x=220, y=35)

var_ent_size_WT = tk.StringVar()
lbl_size_WT = Label(frame_WT_FA, text="Size")
ent_size_WT = Entry(frame_WT_FA, width=10, textvariable=var_ent_size_WT)
lbl_SI_size_WT = Label(frame_WT_FA, text="MWp")
lbl_des_size_WT = Label(frame_WT_FA, text="(Size of the Plant)")
ent_size_WT.config(state="disabled")

lbl_size_WT.place(x=10, y=60)
ent_size_WT.place(x=100, y=60)
lbl_SI_size_WT.place(x=170, y=60)
lbl_des_size_WT.place(x=220, y=60)

var_ent_prod_WT = tk.StringVar()
lbl_prod_WT = Label(frame_WT_FA, text="Production")
ent_prod_WT = Entry(frame_WT_FA, width=10, textvariable=var_ent_prod_WT)
lbl_SI_prod_WT = Label(frame_WT_FA, text="MWh/y")
lbl_des_prod_WT = Label(frame_WT_FA, text="(Energy Production of the Plant of first year)", font=("Helvatical bold", 7))
ent_prod_WT.config(state="disabled")

lbl_prod_WT.place(x=10, y=85)
ent_prod_WT.place(x=100, y=85)
lbl_SI_prod_WT.place(x=170, y=85)
lbl_des_prod_WT.place(x=220, y=85)

var_ent_grid_inj_WT = tk.StringVar()
lbl_grid_inj_WT = Label(frame_WT_FA, text="Grid Injection")
ent_grid_inj_WT = Entry(frame_WT_FA, width=10, textvariable=var_ent_grid_inj_WT)
lbl_SI_grid_inj_WT = Label(frame_WT_FA, text="MWh/y")
lbl_des_grid_inj_WT = Label(frame_WT_FA, text=" Starting Specific Production (first year)", font=("Helvatical bold", 7))
ent_grid_inj_WT.config(state="disabled")

lbl_grid_inj_WT.place(x=10, y=110)
ent_grid_inj_WT.place(x=100, y=110)
lbl_SI_grid_inj_WT.place(x=170, y=110)
lbl_des_grid_inj_WT.place(x=220, y=110)

lbl_rate_WT = Label(frame_WT_FA, text="Rate")
ent_rate_WT = Entry(frame_WT_FA, width=10)
lbl_SI_rate_WT = Label(frame_WT_FA, text="%")
lbl_des_rate_WT = Label(frame_WT_FA, text="(Discount Rate)")

ent_rate_WT.insert(END, str(10))

lbl_rate_WT.place(x=10, y=135)
ent_rate_WT.place(x=100, y=135)
lbl_SI_rate_WT.place(x=170, y=135)
lbl_des_rate_WT.place(x=220, y=135)

var_ent_life_WT = tk.StringVar()
lbl_life_WT = Label(frame_WT_FA, text="Life")
ent_life_WT = Entry(frame_WT_FA, width=10, textvariable=var_ent_life_WT)
lbl_SI_life_WT = Label(frame_WT_FA, text="Years")
lbl_des_life_WT = Label(frame_WT_FA, text="(Useful Life of the Plant)")
ent_life_WT.config(state="disabled")

lbl_life_WT.place(x=10, y=160)
ent_life_WT.place(x=100, y=160)
lbl_SI_life_WT.place(x=170, y=160)
lbl_des_life_WT.place(x=220, y=160)

lbl_ageing_WT = Label(frame_WT_FA, text="Ageing")
ent_ageing_WT = Entry(frame_WT_FA, width=10)
lbl_SI_ageing_WT = Label(frame_WT_FA, text="%")
lbl_des_ageing_WT = Label(frame_WT_FA, text="( Yearly Degradation of the Plant )")

ent_ageing_WT.insert(END, str(0.5))

lbl_ageing_WT.place(x=10, y=185)
ent_ageing_WT.place(x=100, y=185)
lbl_SI_ageing_WT.place(x=170, y=185)
lbl_des_ageing_WT.place(x=220, y=185)

lbl_ele_sp_WT = Label(frame_WT_FA, text="Electricity selling \n price", font=("Helvatical bold", 8))
ent_ele_sp_WT = Entry(frame_WT_FA, width=10)
lbl_SI_ele_sp_WT = Label(frame_WT_FA, text="$/kWh")
lbl_des_ele_sp_WT = Label(frame_WT_FA, text="(For First 5 years)")
ent_ele_sp2_WT = Entry(frame_WT_FA, width=10)
lbl_SI_ele_sp2_WT = Label(frame_WT_FA, text="$/kWh")
lbl_des_ele_sp2_WT = Label(frame_WT_FA, text="(After 5th year)")

ent_ele_sp_WT.insert(END, str(0.04))
ent_ele_sp2_WT.insert(END, str(0.05))

btn_elec_sp_WT = Button(frame_WT_FA, text= "Enter yearly selling price", command=variable_selling_price_wt)
btn_elec_sp_WT.place(x=100, y=210)


lbl_ele_sp_WT.place(x=10, y=210)
# ent_ele_sp_WT.place(x=100, y=210)
# lbl_SI_ele_sp_WT.place(x=170, y=210)
# lbl_des_ele_sp_WT.place(x=220, y=210)
# ent_ele_sp2_WT.place(x=100, y=235)
# lbl_SI_ele_sp2_WT.place(x=170, y=235)
# lbl_des_ele_sp2_WT.place(x=220, y=235)

var_ent_spec_prod_WT = tk.StringVar()
lbl_spec_prod_WT = Label(frame_WT_FA, text="Specific Production", font=("Helvatical ", 7))
ent_spec_prod_WT = Entry(frame_WT_FA, width=10, textvariable=var_ent_spec_prod_WT)
lbl_SI_spec_prod_WT = Label(frame_WT_FA, text="MWh/MW/y", font=("Helvatical ", 7))
lbl_des_spec_prod_WT = Label(frame_WT_FA, text="(Productivity of the Plant for first year)",
                             font=("Helvatical bold", 8))
ent_spec_prod_WT.config(state="disabled")

lbl_spec_prod_WT.place(x=10, y=260)
ent_spec_prod_WT.place(x=100, y=260)
lbl_SI_spec_prod_WT.place(x=170, y=260)
lbl_des_spec_prod_WT.place(x=220, y=260)
#########Input Parameters for STORAGE  Financial Assessment#############################################################
frame_storage_FA = LabelFrame(tab_fin_ana, text="Input Parameters for STORAGE Financial Assessment",
                              font=("Helvatical ", 7))
frame_storage_FA.place(x=880, y=10, height=85, width=260)

lbl_ins_cost_ST = Label(frame_storage_FA, text="Ins Cost")
ent_ins_cost_ST = Entry(frame_storage_FA, width=10)
lbl_SI_ins_cost_ST = Label(frame_storage_FA, text="$/kWh")
lbl_des_ins_cost_ST = Label(frame_storage_FA, text="(Total Installment \n cost)", font=("Helvatical ", 7))

ent_ins_cost_ST.insert(END, str(300))

lbl_ins_cost_ST.place(x=10, y=10)
ent_ins_cost_ST.place(x=65, y=10)
lbl_SI_ins_cost_ST.place(x=130, y=10)
lbl_des_ins_cost_ST.place(x=170, y=10)

var_ent_size_ST = tk.StringVar()
lbl_size_ST = Label(frame_storage_FA, text="Size")
ent_size_ST = Entry(frame_storage_FA, width=10, textvariable=var_ent_size_ST)
lbl_SI_size_ST = Label(frame_storage_FA, text="MWh")
ent_size_ST.config(state="disabled")

lbl_size_ST.place(x=10, y=35)
ent_size_ST.place(x=65, y=35)
lbl_SI_size_ST.place(x=130, y=35)
#####Self consumption table#############################################################################################
frame_self_consumption = LabelFrame(tab_fin_ana, text="Self Consumption")
frame_self_consumption.place(x=880, y=100, height=220, width=260)

lst_self_consumption = [
    "Yes",
    "No"
]
var_dd_SC = StringVar()
lbl_self_consumption = Label(frame_self_consumption, text="Self\nConsumption", font=("Helvatical ", 7))
lbl_self_consumption.place(x=10, y=10)
drpdwn_self_consumption = OptionMenu(frame_self_consumption, var_dd_SC, *lst_self_consumption)
drpdwn_self_consumption.place(x=85, y=10)
var_dd_SC.set(lst_self_consumption[0])
# tip1=Balloon(frame_self_consumption)
btn_hover_des = Button(frame_self_consumption, text="?")
# tip1.bind_widget(btn_hover_des,balloonmsg="HI")

lbl_value_SC_energy = Label(frame_self_consumption, text="Value of\nself consumed\nEnergy", font=("Helvatical ", 7))
ent_value_SC_energy = Entry(frame_self_consumption, width=10)
lbl_SI_value_SC_energy = Label(frame_self_consumption, text="$/kWh")

ent_value_SC_energy.insert(END, str(0.2))

lbl_value_SC_energy.place(x=10, y=40)
ent_value_SC_energy.place(x=85, y=50)
lbl_SI_value_SC_energy.place(x=150, y=50)

var_ent_SC_energy_PV = tk.StringVar()
lbl_SC_energy_PV = Label(frame_self_consumption, text="Self-consumed\nenergy from\nPV", font=("Helvatical ", 7))
ent_SC_energy_PV = Entry(frame_self_consumption, width=10, textvariable=var_ent_SC_energy_PV)
lbl_SI_SC_energy_PV = Label(frame_self_consumption, text="MWh/y")
ent_SC_energy_PV.config(state="disabled")

lbl_SC_energy_PV.place(x=10, y=80)
ent_SC_energy_PV.place(x=85, y=85)
lbl_SI_SC_energy_PV.place(x=150, y=85)

var_ent_SC_energy_WT = tk.StringVar()
lbl_SC_energy_WT = Label(frame_self_consumption, text="Self-consumed\nenergy from\nWT", font=("Helvatical ", 7))
ent_SC_energy_WT = Entry(frame_self_consumption, width=10, textvariable=var_ent_SC_energy_WT)
lbl_SI_SC_energy_WT = Label(frame_self_consumption, text="MWh/y")
ent_SC_energy_WT.config(state="disabled")

lbl_SC_energy_WT.place(x=10, y=115)
ent_SC_energy_WT.place(x=85, y=120)
lbl_SI_SC_energy_WT.place(x=150, y=120)

lbl_des_self_consumption = Label(frame_self_consumption,
                                 text="*if No is selected, it means that there is not\nself-consumption,"
                                      " and all the energy is sold to the grid.", font=("Helvatical ", 7))
lbl_des_self_consumption.place(x=10, y=160)
#####Output Parameters of RES (PV+WT) and storage Financial Assessment##################################################
frame_op_para_RES_FA = LabelFrame(tab_fin_ana, text="Output Parameters of RES (PV+WT) and storage Financial Assessment")
frame_op_para_RES_FA.place(x=10, y=330, height=210, width=420)

var_ent_irr = tk.StringVar()
lbl_Irr = Label(frame_op_para_RES_FA, text="IRR")
ent_irr = Entry(frame_op_para_RES_FA, width=10, textvariable=var_ent_irr)
lbl_des_irr = Label(frame_op_para_RES_FA,
                    text="Discount rate at which NPV equals zero at\nthe end of project life time",
                    font=("Helvatical ", 7))
ent_irr.config(state="disabled")

lbl_Irr.place(x=10, y=10)
ent_irr.place(x=85, y=10)
lbl_des_irr.place(x=210, y=10)

var_ent_NPV = tk.StringVar()
lbl_NPV = Label(frame_op_para_RES_FA, text="NPV")
ent_NPV = Entry(frame_op_para_RES_FA, width=10, textvariable=var_ent_NPV)
lbl_SI_NPV = Label(frame_op_para_RES_FA, text="M$")
lbl_des_NPV = Label(frame_op_para_RES_FA, text="Total of future incomes counted back\nto investment year",
                    font=("Helvatical ", 7))
ent_NPV.config(state="disabled")

lbl_NPV.place(x=10, y=60)
ent_NPV.place(x=85, y=60)
lbl_SI_NPV.place(x=150, y=60)
lbl_des_NPV.place(x=210, y=60)

var_ent_PBT = tk.StringVar()
lbl_PBT = Label(frame_op_para_RES_FA, text="PBT")
ent_PBT = Entry(frame_op_para_RES_FA, width=10, textvariable=var_ent_PBT)
lbl_SI_PBT = Label(frame_op_para_RES_FA, text="Years")
lbl_des_PBT = Label(frame_op_para_RES_FA, text="Number of years to return the\ninvestment", font=("Helvatical ", 7))
ent_PBT.config(state="disabled")

lbl_PBT.place(x=10, y=110)
ent_PBT.place(x=85, y=110)
lbl_SI_PBT.place(x=150, y=110)
lbl_des_PBT.place(x=210, y=110)

var_ent_LCOE = tk.StringVar()
lbl_LCOE = Label(frame_op_para_RES_FA, text="LCOE")
ent_LCOE = Entry(frame_op_para_RES_FA, width=10, textvariable=var_ent_LCOE)
lbl_SI_LCOE = Label(frame_op_para_RES_FA, text="$/kWh")
lbl_des_LCOE = Label(frame_op_para_RES_FA, text="Price of electricity to cover total\nlife time expenditures",
                     font=("Helvatical ", 7))
ent_LCOE.config(state="disabled")

lbl_LCOE.place(x=10, y=160)
ent_LCOE.place(x=85, y=160)
lbl_SI_LCOE.place(x=150, y=160)
lbl_des_LCOE.place(x=210, y=160)
####net metering or net billing#########################################################################################
frame_NM_and_NB = LabelFrame(tab_fin_ana, text="Net Metering")
frame_NM_and_NB.place(x=450, y=330, height=210, width=320)

lbl_NM_and_NB = Label(frame_NM_and_NB, text="Net metering or Net Billing")
lst_NMandNB = [
    "Yes",
    "No"
]
var_NMandNB = StringVar()
drpdwn_NMandNB = OptionMenu(frame_NM_and_NB, var_NMandNB, *lst_NMandNB)
var_NMandNB.set(lst_NMandNB[1])
lbl_NM_and_NB.place(x=10, y=10)
drpdwn_NMandNB.place(x=200, y=8)

var_ent_annual_surplus = tk.StringVar()
lbl_annual_surplus = Label(frame_NM_and_NB,
                           text="Annual SURPLUS of the annual injection\n(after the balance due to net metering)\n(MWh/y)",
                           font=("Helvatical ", 7))
ent_annual_surplus = Entry(frame_NM_and_NB, width=10, textvariable=var_ent_annual_surplus)
lbl_si_annual_surplus = Label(frame_NM_and_NB, text="MWh/y")
ent_annual_surplus.config(state="disabled")

lbl_annual_surplus.place(x=10, y=55)
ent_annual_surplus.place(x=200, y=60)
lbl_si_annual_surplus.place(x=265, y=60)

var_ent_NM_energy = tk.StringVar()
lbl_NM_energy = Label(frame_NM_and_NB, text="Net metering energy")
ent_NM_energy = Entry(frame_NM_and_NB, width=10, textvariable=var_ent_NM_energy)
lbl_si_NM_energy = Label(frame_NM_and_NB, text="MWh/y")
ent_NM_energy.config(state="disabled")

lbl_NM_energy.place(x=10, y=110)
ent_NM_energy.place(x=200, y=110)
lbl_si_NM_energy.place(x=265, y=110)

lbl_NM_positive_CF = Label(frame_NM_and_NB,
                           text="Net metering additional positive cash flow\nwith respect to the injection in the grid ",
                           font=("Helvatical ", 7))
ent_NM_positive_CF = Entry(frame_NM_and_NB, width=10)
lbl_SI_NM_positive_CF = Label(frame_NM_and_NB, text="$/kWh")

ent_NM_positive_CF.insert(END, str(0.11))

lbl_NM_positive_CF.place(x=10, y=155)
ent_NM_positive_CF.place(x=200, y=160)
lbl_SI_NM_positive_CF.place(x=265, y=160)
######Tax reduction#####################################################################################################
frame_tax_reduction = LabelFrame(tab_fin_ana, text="Tax Reduction")
frame_tax_reduction.place(x=775, y=330, height=210, width=370)

lbl_tax_red_PV = Label(frame_tax_reduction, text="Yearly tax reduction for PV", font=("Helvatical ", 7))
ent_tax_red_PV = Entry(frame_tax_reduction, width=5)
des_tax_red_PV = Label(frame_tax_reduction, text="of the initial investment", font=("Helvatical ", 7))
lbl_tax_red_PV_cost = Label(frame_tax_reduction, text="$", font=("Helvatical ", 7))

ent_tax_red_PV.insert(END, str(1))

lbl_tax_red_PV.place(x=10, y=10)
ent_tax_red_PV.place(x=140, y=10)
des_tax_red_PV.place(x=175, y=10)
lbl_tax_red_PV_cost.place(x=285, y=10)

lbl_val_PV = Label(frame_tax_reduction, text="Number of years of validity\nfor PV", font=("Helvatical ", 7))
ent_val_PV = Entry(frame_tax_reduction, width=5)

ent_val_PV.insert(END, str(10))

lbl_val_PV.place(x=10, y=30)
ent_val_PV.place(x=140, y=30)

lbl_tax_red_WT = Label(frame_tax_reduction, text="Yearly tax reduction for WT", font=("Helvatical ", 7))
ent_tax_red_WT = Entry(frame_tax_reduction, width=5)
des_tax_red_WT = Label(frame_tax_reduction, text="of the initial investment", font=("Helvatical ", 7))
lbl_tax_red_WT_cost = Label(frame_tax_reduction, text="$", font=("Helvatical ", 7))

ent_tax_red_WT.insert(END, str(0))

lbl_tax_red_WT.place(x=10, y=65)
ent_tax_red_WT.place(x=140, y=65)
des_tax_red_WT.place(x=175, y=65)
lbl_tax_red_WT_cost.place(x=285, y=65)

lbl_val_WT = Label(frame_tax_reduction, text="Number of years of validity\nfor WT", font=("Helvatical ", 7))
ent_val_WT = Entry(frame_tax_reduction, width=5)

ent_val_WT.insert(END, str(10))

lbl_val_WT.place(x=10, y=85)
ent_val_WT.place(x=140, y=85)

lbl_tax_red_ST = Label(frame_tax_reduction, text="Yearly tax reduction for\nstorage", font=("Helvatical ", 7))
ent_tax_red_ST = Entry(frame_tax_reduction, width=5)
des_tax_red_ST = Label(frame_tax_reduction, text="of the initial investment", font=("Helvatical ", 7))
lbl_tax_red_ST_cost = Label(frame_tax_reduction, text="$", font=("Helvatical ", 7))

ent_tax_red_ST.insert(END, str(0))

lbl_tax_red_ST.place(x=10, y=120)
ent_tax_red_ST.place(x=140, y=120)
des_tax_red_ST.place(x=175, y=120)
lbl_tax_red_ST_cost.place(x=285, y=120)

lbl_val_ST = Label(frame_tax_reduction, text="Number of years of validity\nfor Storage", font=("Helvatical ", 7))
ent_val_ST = Entry(frame_tax_reduction, width=5)

ent_val_ST.insert(END, str(10))

lbl_val_ST.place(x=10, y=150)
ent_val_ST.place(x=140, y=150)

#######################################################################################################################
btn_all_WT_prod = Button(tab_WT_results, text="Calculate WT production", command=cal_all_WT_prod)
btn_all_WT_prod.place(x=0, y=0)
#######################################################################################################################
# UI of results in main window
frm_results_main_win = LabelFrame(tab_inputs, text="Results of First year")
frm_results_main_win.place(x=820, y=40, height=440, width=440)

lbl_pv_prodcutivity = Label(frm_results_main_win, text="PV productivity")
lbl_pv_prodcutivity.place(x=10, y=10)
lbl_SI_pv_prodcutivity = Label(frm_results_main_win, text="MWh/MW/year")
lbl_SI_pv_prodcutivity.place(x=240, y=10)
ent_pv_prodcutivity = Entry(frm_results_main_win, width=10)
ent_pv_prodcutivity.place(x=340, y=10)
ent_pv_prodcutivity.configure(state="disabled")

lbl_pv_annual_prod = Label(frm_results_main_win, text="Annual PV production")
lbl_pv_annual_prod.place(x=10, y=35)
lbl_SI_pv_annual_prod = Label(frm_results_main_win, text="MWh/year")
lbl_SI_pv_annual_prod.place(x=240, y=35)
ent_pv_annual_prod = Entry(frm_results_main_win, width=10)
ent_pv_annual_prod.place(x=340, y=35)
ent_pv_annual_prod.configure(state="disabled")

lbl_pv_productivity_with_ps = Label(frm_results_main_win, text="Productivity of PV with Peak shaving")
lbl_pv_productivity_with_ps.place(x=10, y=60)
lbl_SI_pv_productivity_with_ps = Label(frm_results_main_win, text="MWh/MW/year")
lbl_SI_pv_productivity_with_ps.place(x=240, y=60)
ent_pv_productivity_with_ps = Entry(frm_results_main_win, width=10)
ent_pv_productivity_with_ps.place(x=340, y=60)
ent_pv_productivity_with_ps.configure(state="disabled")

lbl_Annual_pv_prod_with_ps = Label(frm_results_main_win, text="Annual PV production with Peak shaving")
lbl_Annual_pv_prod_with_ps.place(x=10, y=85)
lbl_SI_Annual_pv_prod_with_ps = Label(frm_results_main_win, text="MWh/year")
lbl_SI_Annual_pv_prod_with_ps.place(x=240, y=85)
ent_Annual_pv_prod_with_ps = Entry(frm_results_main_win, width=10)
ent_Annual_pv_prod_with_ps.place(x=340, y=85)
ent_Annual_pv_prod_with_ps.configure(state="disabled")

lbl_WT_prodcutivity = Label(frm_results_main_win, text="WT productivity")
lbl_WT_prodcutivity.place(x=10, y=110)
lbl_SI_WT_prodcutivity = Label(frm_results_main_win, text="MWh/MW/year")
lbl_SI_WT_prodcutivity.place(x=240, y=110)
ent_WT_prodcutivity = Entry(frm_results_main_win, width=10)
ent_WT_prodcutivity.place(x=340, y=110)
ent_WT_prodcutivity.configure(state="disabled")

lbl_WT_annual_prod = Label(frm_results_main_win, text="Annual WT production")
lbl_WT_annual_prod.place(x=10, y=135)
lbl_SI_WT_annual_prod = Label(frm_results_main_win, text="MWh/year")
lbl_SI_WT_annual_prod.place(x=240, y=135)
ent_WT_annual_prod = Entry(frm_results_main_win, width=10)
ent_WT_annual_prod.place(x=340, y=135)
ent_WT_annual_prod.configure(state="disabled")

lbl_WT_productivity_with_ps = Label(frm_results_main_win, text="Productivity of WT with Peak shaving")
lbl_WT_productivity_with_ps.place(x=10, y=160)
lbl_SI_WT_productivity_with_ps = Label(frm_results_main_win, text="MWh/MW/year")
lbl_SI_WT_productivity_with_ps.place(x=240, y=160)
ent_WT_productivity_with_ps = Entry(frm_results_main_win, width=10)
ent_WT_productivity_with_ps.place(x=340, y=160)
ent_WT_productivity_with_ps.configure(state="disabled")

lbl_Annual_WT_prod_with_ps = Label(frm_results_main_win, text="Annual WT production with peak shaving")
lbl_Annual_WT_prod_with_ps.place(x=10, y=185)
lbl_SI_Annual_WT_prod_with_ps = Label(frm_results_main_win, text="MWh/year")
lbl_SI_Annual_WT_prod_with_ps.place(x=240, y=185)
ent_Annual_WT_prod_with_ps = Entry(frm_results_main_win, width=10)
ent_Annual_WT_prod_with_ps.place(x=340, y=185)
ent_Annual_WT_prod_with_ps.configure(state="disabled")

lbl_dischrged_energy = Label(frm_results_main_win, text="Discharged Energy")
lbl_dischrged_energy.place(x=10, y=210)
lbl_SI_dischrged_energy = Label(frm_results_main_win, text="MWh/year")
lbl_SI_dischrged_energy.place(x=240, y=210)
ent_dischrged_energy = Entry(frm_results_main_win, width=10)
ent_dischrged_energy.place(x=340, y=210)
ent_dischrged_energy.configure(state="disabled")

lbl_chrged_energy = Label(frm_results_main_win, text="Charged Energy")
lbl_chrged_energy.place(x=10, y=235)
lbl_SI_chrged_energy = Label(frm_results_main_win, text="MWh/year")
lbl_SI_chrged_energy.place(x=240, y=235)
ent_chrged_energy = Entry(frm_results_main_win, width=10)
ent_chrged_energy.place(x=340, y=235)
ent_chrged_energy.configure(state="disabled")

lbl_self_consumption = Label(frm_results_main_win, text="Self Consumption")
lbl_self_consumption.place(x=10, y=260)
lbl_SI_self_consumption = Label(frm_results_main_win, text="%")
lbl_SI_self_consumption.place(x=240, y=260)
ent_self_consumption = Entry(frm_results_main_win, width=10)
ent_self_consumption.place(x=340, y=260)
ent_self_consumption.configure(state="disabled")

lbl_self_sufficiency = Label(frm_results_main_win, text="Self Sufficiency")
lbl_self_sufficiency.place(x=10, y=285)
lbl_SI_self_sufficiency = Label(frm_results_main_win, text="%")
lbl_SI_self_sufficiency.place(x=240, y=285)
ent_self_sufficiency = Entry(frm_results_main_win, width=10)
ent_self_sufficiency.place(x=340, y=285)
ent_self_sufficiency.configure(state="disabled")

lbl_absorption_from_grid = Label(frm_results_main_win, text="Absorption from the grid / load")
lbl_absorption_from_grid.place(x=10, y=310)
lbl_SI_absorption_from_grid = Label(frm_results_main_win, text="%")
lbl_SI_absorption_from_grid.place(x=240, y=310)
ent_absorption_from_grid = Entry(frm_results_main_win, width=10)
ent_absorption_from_grid.place(x=340, y=310)
ent_absorption_from_grid.configure(state="disabled")

lbl_injection_to_grid = Label(frm_results_main_win, text="Injection in the grid / load")
lbl_injection_to_grid.place(x=10, y=335)
lbl_SI_injection_to_grid = Label(frm_results_main_win, text="%")
lbl_SI_injection_to_grid.place(x=240, y=335)
ent_injection_to_grid = Entry(frm_results_main_win, width=10)
ent_injection_to_grid.place(x=340, y=335)
ent_injection_to_grid.configure(state="disabled")

lbl_anaual_load = Label(frm_results_main_win, text="Annual Load")
lbl_anaual_load.place(x=10, y=360)
lbl_SI_anaual_load = Label(frm_results_main_win, text="MWh/year")
lbl_SI_anaual_load.place(x=240, y=360)
ent_annual_load = Entry(frm_results_main_win, width=10)
ent_annual_load.place(x=340, y=360)
ent_annual_load.configure(state="disabled")

lst_SI_units = [
    "kWh",
    "MWh",
    "GWh"
]

var_results_SI_units = StringVar()
drp_results_SI_units = OptionMenu(frm_results_main_win, var_results_SI_units, *lst_SI_units)
var_results_SI_units.set(lst_SI_units[1])
drp_results_SI_units.place(x=10, y=390)
lst_result_labels = [lbl_SI_pv_prodcutivity, lbl_SI_pv_annual_prod, lbl_SI_pv_productivity_with_ps,
                     lbl_SI_Annual_pv_prod_with_ps, lbl_SI_WT_prodcutivity,
                     lbl_SI_WT_annual_prod, lbl_SI_WT_productivity_with_ps, lbl_SI_Annual_WT_prod_with_ps,
                     lbl_SI_dischrged_energy, lbl_SI_chrged_energy
    , lbl_SI_anaual_load]

lst_result_entries = [ent_pv_annual_prod, ent_Annual_pv_prod_with_ps, ent_WT_annual_prod, ent_Annual_WT_prod_with_ps,
                      ent_dischrged_energy
    , ent_chrged_energy, ent_annual_load]

var_results_SI_units.trace("w", lambda *args: convert_SI_units(lst_result_labels, lst_result_values, lst_result_entries,
                                                               var_results_SI_units))

#########################################################################################################################
# Optimization part
btn_opt = Button(tab_opt, text="optimize", command=optimization)
btn_opt.place(x=10, y=10)

lbl_opt_selfsufficiency = Label(tab_opt, text="Optimum Self sufficiency(%)")
lbl_opt_selfsufficiency.place(x=10, y=50)
ent_opt_selfsufficiency = Entry(tab_opt, width=10)
ent_opt_selfsufficiency.place(x=170, y=50)
lbl_opt_irr = Label(tab_opt, text="Optimum IRR(%)")
lbl_opt_irr.place(x=250, y=50)
ent_opt_irr = Entry(tab_opt, width=10)
ent_opt_irr.place(x=350, y=50)
lbl_opt_PV_size = Label(tab_opt, text="Optimum PV size(kW)")
lbl_opt_PV_size.place(x=450, y=50)
ent_opt_PV_size = Entry(tab_opt, width=10)
ent_opt_PV_size.place(x=570, y=50)
lbl_opt_WT_size = Label(tab_opt, text="Optimum num WT")
lbl_opt_WT_size.place(x=650, y=50)
ent_opt_WT_size = Entry(tab_opt, width=10)
ent_opt_WT_size.place(x=775, y=50)
lbl_opt_ST_size = Label(tab_opt, text="Optimum ST size(kWh)")
lbl_opt_ST_size.place(x=850, y=50)
ent_opt_ST_size = Entry(tab_opt, width=10)
ent_opt_ST_size.place(x=980, y=50)
#########################################################################################################################
# btn_opt_SS_irr = Button(tab_opt, text="SS & IRR")
# btn_opt_SS_irr.place(x=100, y=10)
# ToolTip(btn_opt_SS_irr, msg="Hover info 1", follow=True)
# btn_opt_results = Button(tab_opt, text="Get values")
# btn_opt_results.place(x=200, y=10)
# ToolTip(btn_opt_results, msg="Hover info 2", follow=True)
frm_input_opt = LabelFrame(tab_opt, text="Inputs for Optimization", height=220, width=250)
frm_input_opt.place(x=10, y=150)

lbl_min_opt = Label(frm_input_opt, text="Min")
lbl_min_opt.place(x=80, y=10)
lbl_max_opt = Label(frm_input_opt, text="Max")
lbl_max_opt.place(x=160, y=10)

lbl_pv_size_opt_min_max = Label(frm_input_opt, text="PV size(kW)")
lbl_pv_size_opt_min_max.place(x=10, y=40)
ent_pv_size_opt_min = Entry(frm_input_opt, width=10)
ent_pv_size_opt_min.place(x=80, y=40)
ent_pv_size_opt_max = Entry(frm_input_opt, width=10)
ent_pv_size_opt_max.place(x=160, y=40)

lbl_wt_size_opt_min_max = Label(frm_input_opt, text="Num WT's")
lbl_wt_size_opt_min_max.place(x=10, y=70)
ent_wt_size_opt_min = Entry(frm_input_opt, width=10)
ent_wt_size_opt_min.place(x=80, y=70)
ent_wt_size_opt_max = Entry(frm_input_opt, width=10)
ent_wt_size_opt_max.place(x=160, y=70)

lbl_st_size_opt_min_max = Label(frm_input_opt, text="ST size(kWh)")
lbl_st_size_opt_min_max.place(x=10, y=100)
ent_st_size_opt_min = Entry(frm_input_opt, width=10)
ent_st_size_opt_min.place(x=80, y=100)
ent_st_size_opt_max = Entry(frm_input_opt, width=10)
ent_st_size_opt_max.place(x=160, y=100)

lbl_irr_opt_min = Label(frm_input_opt, text="IRR(%)")
lbl_irr_opt_min.place(x=10, y=130)
ent_irr_opt_min = Entry(frm_input_opt, width=10)
ent_irr_opt_min.place(x=80, y=130)

lbl_tol_opt_min = Label(frm_input_opt, text="Tolerence")
lbl_tol_opt_min.place(x=10, y=160)
ent_tol_opt = Entry(frm_input_opt, width=10)
ent_tol_opt.place(x=80, y=160)

# UI of results in optimization window
frm_results_opt_win = LabelFrame(tab_opt, text="Results of Optimization")
frm_results_opt_win.place(x=420, y=150, height=440, width=440)

lbl_pv_prodcutivity_opt = Label(frm_results_opt_win, text="PV productivity")
lbl_pv_prodcutivity_opt.place(x=10, y=10)
lbl_SI_pv_prodcutivity_opt = Label(frm_results_opt_win, text="MWh/MW/year")
lbl_SI_pv_prodcutivity_opt.place(x=240, y=10)
ent_pv_prodcutivity_opt = Entry(frm_results_opt_win, width=10)
ent_pv_prodcutivity_opt.place(x=340, y=10)
ent_pv_prodcutivity_opt.configure(state="disabled")

lbl_pv_annual_prod_opt = Label(frm_results_opt_win, text="Annual PV production")
lbl_pv_annual_prod_opt.place(x=10, y=35)
lbl_SI_pv_annual_prod_opt = Label(frm_results_opt_win, text="MWh/year")
lbl_SI_pv_annual_prod_opt.place(x=240, y=35)
ent_pv_annual_prod_opt = Entry(frm_results_opt_win, width=10)
ent_pv_annual_prod_opt.place(x=340, y=35)
ent_pv_annual_prod_opt.configure(state="disabled")

lbl_pv_productivity_with_ps_opt = Label(frm_results_opt_win, text="Productivity of PV with Peak shaving")
lbl_pv_productivity_with_ps_opt.place(x=10, y=60)
lbl_SI_pv_productivity_with_ps_opt = Label(frm_results_opt_win, text="MWh/MW/year")
lbl_SI_pv_productivity_with_ps_opt.place(x=240, y=60)
ent_pv_productivity_with_ps_opt = Entry(frm_results_opt_win, width=10)
ent_pv_productivity_with_ps_opt.place(x=340, y=60)
ent_pv_productivity_with_ps_opt.configure(state="disabled")

lbl_Annual_pv_prod_with_ps_opt = Label(frm_results_opt_win, text="Annual PV production with Peak shaving")
lbl_Annual_pv_prod_with_ps_opt.place(x=10, y=85)
lbl_SI_Annual_pv_prod_with_ps_opt = Label(frm_results_opt_win, text="MWh/year")
lbl_SI_Annual_pv_prod_with_ps_opt.place(x=240, y=85)
ent_Annual_pv_prod_with_ps_opt = Entry(frm_results_opt_win, width=10)
ent_Annual_pv_prod_with_ps_opt.place(x=340, y=85)
ent_Annual_pv_prod_with_ps_opt.configure(state="disabled")

lbl_WT_prodcutivity_opt = Label(frm_results_opt_win, text="WT productivity")
lbl_WT_prodcutivity_opt.place(x=10, y=110)
lbl_SI_WT_prodcutivity_opt = Label(frm_results_opt_win, text="MWh/MW/year")
lbl_SI_WT_prodcutivity_opt.place(x=240, y=110)
ent_WT_prodcutivity_opt = Entry(frm_results_opt_win, width=10)
ent_WT_prodcutivity_opt.place(x=340, y=110)
ent_WT_prodcutivity_opt.configure(state="disabled")

lbl_WT_annual_prod_opt = Label(frm_results_opt_win, text="Annual WT production")
lbl_WT_annual_prod_opt.place(x=10, y=135)
lbl_SI_WT_annual_prod_opt = Label(frm_results_opt_win, text="MWh/year")
lbl_SI_WT_annual_prod_opt.place(x=240, y=135)
ent_WT_annual_prod_opt = Entry(frm_results_opt_win, width=10)
ent_WT_annual_prod_opt.place(x=340, y=135)
ent_WT_annual_prod_opt.configure(state="disabled")

lbl_WT_productivity_with_ps_opt = Label(frm_results_opt_win, text="Productivity of WT with Peak shaving")
lbl_WT_productivity_with_ps_opt.place(x=10, y=160)
lbl_SI_WT_productivity_with_ps_opt = Label(frm_results_opt_win, text="MWh/MW/year")
lbl_SI_WT_productivity_with_ps_opt.place(x=240, y=160)
ent_WT_productivity_with_ps_opt = Entry(frm_results_opt_win, width=10)
ent_WT_productivity_with_ps_opt.place(x=340, y=160)
ent_WT_productivity_with_ps_opt.configure(state="disabled")

lbl_Annual_WT_prod_with_ps_opt = Label(frm_results_opt_win, text="Annual WT production with peak shaving")
lbl_Annual_WT_prod_with_ps_opt.place(x=10, y=185)
lbl_SI_Annual_WT_prod_with_ps_opt = Label(frm_results_opt_win, text="MWh/year")
lbl_SI_Annual_WT_prod_with_ps_opt.place(x=240, y=185)
ent_Annual_WT_prod_with_ps_opt = Entry(frm_results_opt_win, width=10)
ent_Annual_WT_prod_with_ps_opt.place(x=340, y=185)
ent_Annual_WT_prod_with_ps_opt.configure(state="disabled")

lbl_dischrged_energy_opt = Label(frm_results_opt_win, text="Discharged Energy")
lbl_dischrged_energy_opt.place(x=10, y=210)
lbl_SI_dischrged_energy_opt = Label(frm_results_opt_win, text="MWh/year")
lbl_SI_dischrged_energy_opt.place(x=240, y=210)
ent_dischrged_energy_opt = Entry(frm_results_opt_win, width=10)
ent_dischrged_energy_opt.place(x=340, y=210)
ent_dischrged_energy_opt.configure(state="disabled")

lbl_chrged_energy_opt = Label(frm_results_opt_win, text="Charged Energy")
lbl_chrged_energy_opt.place(x=10, y=235)
lbl_SI_chrged_energy_opt = Label(frm_results_opt_win, text="MWh/year")
lbl_SI_chrged_energy_opt.place(x=240, y=235)
ent_chrged_energy_opt = Entry(frm_results_opt_win, width=10)
ent_chrged_energy_opt.place(x=340, y=235)
ent_chrged_energy_opt.configure(state="disabled")

lbl_self_consumption_opt = Label(frm_results_opt_win, text="Self Consumption")
lbl_self_consumption_opt.place(x=10, y=260)
lbl_SI_self_consumption_opt = Label(frm_results_opt_win, text="%")
lbl_SI_self_consumption_opt.place(x=240, y=260)
ent_self_consumption_opt = Entry(frm_results_opt_win, width=10)
ent_self_consumption_opt.place(x=340, y=260)
ent_self_consumption_opt.configure(state="disabled")

lbl_self_sufficiency_opt = Label(frm_results_opt_win, text="Self Sufficiency")
lbl_self_sufficiency_opt.place(x=10, y=285)
lbl_SI_self_sufficiency_opt = Label(frm_results_opt_win, text="%")
lbl_SI_self_sufficiency_opt.place(x=240, y=285)
ent_self_sufficiency_opt = Entry(frm_results_opt_win, width=10)
ent_self_sufficiency_opt.place(x=340, y=285)
ent_self_sufficiency_opt.configure(state="disabled")

lbl_absorption_from_grid_opt = Label(frm_results_opt_win, text="Absorption from the grid / load")
lbl_absorption_from_grid_opt.place(x=10, y=310)
lbl_SI_absorption_from_grid_opt = Label(frm_results_opt_win, text="%")
lbl_SI_absorption_from_grid_opt.place(x=240, y=310)
ent_absorption_from_grid_opt = Entry(frm_results_opt_win, width=10)
ent_absorption_from_grid_opt.place(x=340, y=310)
ent_absorption_from_grid_opt.configure(state="disabled")

lbl_injection_to_grid_opt = Label(frm_results_opt_win, text="Injection in the grid / load")
lbl_injection_to_grid_opt.place(x=10, y=335)
lbl_SI_injection_to_grid_opt = Label(frm_results_opt_win, text="%")
lbl_SI_injection_to_grid_opt.place(x=240, y=335)
ent_injection_to_grid_opt = Entry(frm_results_opt_win, width=10)
ent_injection_to_grid_opt.place(x=340, y=335)
ent_injection_to_grid_opt.configure(state="disabled")

lbl_anaual_load_opt = Label(frm_results_opt_win, text="Annual Load")
lbl_anaual_load_opt.place(x=10, y=360)
lbl_SI_anaual_load_opt = Label(frm_results_opt_win, text="MWh/year")
lbl_SI_anaual_load_opt.place(x=240, y=360)
ent_annual_load_opt = Entry(frm_results_opt_win, width=10)
ent_annual_load_opt.place(x=340, y=360)
ent_annual_load_opt.configure(state="disabled")

#########################################################################################################################
lst_SI_units_opt = [
    "kWh",
    "MWh",
    "GWh"
]

var_results_SI_units_opt = StringVar()
drp_results_SI_units_opt = OptionMenu(frm_results_opt_win, var_results_SI_units_opt, *lst_SI_units_opt)
var_results_SI_units_opt.set(lst_SI_units_opt[1])
drp_results_SI_units_opt.place(x=10, y=390)
lst_result_labels_opt = [lbl_SI_pv_prodcutivity_opt, lbl_SI_pv_annual_prod_opt, lbl_SI_pv_productivity_with_ps_opt,
                         lbl_SI_Annual_pv_prod_with_ps_opt, lbl_SI_WT_prodcutivity_opt,
                         lbl_SI_WT_annual_prod_opt, lbl_SI_WT_productivity_with_ps_opt,
                         lbl_SI_Annual_WT_prod_with_ps_opt, lbl_SI_dischrged_energy_opt, lbl_SI_chrged_energy_opt
    , lbl_SI_anaual_load_opt]

lst_result_entries_opt = [ent_pv_annual_prod_opt, ent_Annual_pv_prod_with_ps_opt, ent_WT_annual_prod_opt,
                          ent_Annual_WT_prod_with_ps_opt, ent_dischrged_energy_opt
    , ent_chrged_energy_opt, ent_annual_load_opt]

var_results_SI_units_opt.trace("w", lambda *args: convert_SI_units(lst_result_labels_opt, lst_result_values_opt,
                                                                   lst_result_entries_opt, var_results_SI_units_opt))

global nb_of_iterations
nb_of_iterations = 0

lbl_opt_status = Label(tab_opt, text="Optimization status: ")
lbl_opt_status.place(x=20, y=380)

lbl_itirations = Label(tab_opt, text="No of iterations = " + str(nb_of_iterations))
lbl_itirations.place(x=20, y=410)
#########################################################################################################################
root.mainloop()